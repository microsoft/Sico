# Copyright (c) 2026 Sico Authors
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from __future__ import annotations

import argparse
import contextlib
import functools
import json
import mimetypes
import os
import posixpath
import re
import statistics
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field, replace
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any


DEFAULT_BASE_URL = "http://localhost:8080"
DEFAULT_EMAIL = os.getenv("SICO_ACCEPTANCE_EMAIL", "operator@sico.local")
DEFAULT_PASSWORD = os.getenv("SICO_ACCEPTANCE_PASSWORD", "operator")
COPILOT_APK_URL = "https://dwp-cdn-ddcqh0dkgnhbchgs.b01.azurefd.net/test/123/copilot.apk"
XLSX_ANDROID_MAX_SECONDS = 720.0
# Tools that "do not execute / no delegation" scenarios must never invoke. The
# single `delegate` tool (kind selects the adapter) replaces the former
# per-adapter `delegate_general` / `delegate_workbook` tools.
FORBIDDEN_DELEGATION_TOOLS = (
    "delegate",
    "run_command",
    "sandbox_acquire",
    "sandbox_release",
    "sandbox_reset",
)
SEED_PREFLIGHT_SCENARIO = "seeded_agents_preflight"
ACCEPTANCE_FIXTURE_DIR = Path(__file__).resolve().parent / "acceptance-fixtures"
WORKBOOK_HEADER_LIMIT = 12
ACTIVE_RUNNING_LIST_STATUSES = {0, 1, 2}
EXPECTED_AGENT_ROLES = ("Assistant", "Android Tester", "3D Artist", "Product Manager", "Marketing")
EXPECTED_SEEDED_AGENT_INSTANCES = {
    2: {"name": "Max", "role": "Android Tester"},
    3: {"name": "Luna", "role": "3D Artist"},
    4: {"name": "Ethan", "role": "Product Manager"},
    5: {"name": "Chloe", "role": "Marketing"},
}


@dataclass(frozen=True)
class Scenario:
    name: str
    agent_instance_id: int
    message: str
    expect_batch: bool = False
    expect_run_report: bool = False
    expect_attachment_parse: bool = False
    expect_no_attachment_parse: bool = False
    expect_file_conversion: bool = False
    expect_no_low_level_android: bool = True
    expect_single_case: bool = False
    allow_failed_batch: bool = False
    allow_duration_overrun: bool = False
    allow_slow_first_event: bool = False
    min_batch_total: int | None = None
    max_batches: int | None = None
    max_parse_document_tool_calls: int | None = None
    forbidden_tool_names: tuple[str, ...] = ()
    expected_plan_text: tuple[str, ...] = ()
    forbidden_plan_text: tuple[str, ...] = ()
    expected_final_text: tuple[str, ...] = ()
    expected_final_any_text: tuple[str, ...] = ()
    max_seconds: float = 180.0
    attachments: tuple[str, ...] = ()


@dataclass
class ScenarioResult:
    name: str
    ok: bool
    duration_s: float
    agent_instance_id: int = 0
    message: str = ""
    first_event_s: float | None = None
    turn_id: int | None = None
    conversation_id: int | None = None
    event_count: int = 0
    final_content_chars: int = 0
    final_content: str = ""
    plan_tool_calls: int = 0
    batches: list[dict[str, Any]] = field(default_factory=list)
    metrics: dict[str, Any] | None = None
    findings: list[str] = field(default_factory=list)


def main() -> int:
    parser = argparse.ArgumentParser(description="Run real local chat acceptance scenarios against a deployed Sico stack.")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument("--email", default=DEFAULT_EMAIL)
    parser.add_argument("--password", default=DEFAULT_PASSWORD)
    parser.add_argument("--json-out", default="")
    parser.add_argument("--scenario", action="append", default=[], help="Run only named scenario(s).")
    parser.add_argument(
        "--max-seconds",
        type=float,
        default=None,
        help="Override each selected scenario timeout. Useful for long benchmark workbooks.",
    )
    parser.add_argument(
        "--xlsx-path",
        action="append",
        default=[],
        help=(
            "Local XLSX file to attach for a real document-driven scenario. "
            "Defaults to scripts/acceptance-fixtures/*.xlsx when omitted."
        ),
    )
    args = parser.parse_args()

    client = Client(args.base_url, args.email, args.password)
    scenarios = _scenarios(args.xlsx_path or _default_xlsx_paths())
    include_seed_preflight = not args.scenario or SEED_PREFLIGHT_SCENARIO in set(args.scenario)
    if args.scenario:
        wanted = set(args.scenario)
        scenarios = [scenario for scenario in scenarios if scenario.name in wanted]
        missing = wanted - {scenario.name for scenario in scenarios} - {SEED_PREFLIGHT_SCENARIO}
        if missing:
            raise SystemExit(f"unknown scenario(s): {', '.join(sorted(missing))}")
    if args.max_seconds is not None:
        scenarios = [replace(scenario, max_seconds=args.max_seconds) for scenario in scenarios]

    results: list[ScenarioResult] = []
    if include_seed_preflight:
        print(f"\n=== {SEED_PREFLIGHT_SCENARIO} ===", flush=True)
        result = run_seeded_agents_preflight(client)
        results.append(result)
        status = "PASS" if result.ok else "FAIL"
        print(
            f"{status} duration={result.duration_s:.2f}s first_event=n/a turn=None conv=None events=0 tool_calls=0 batches=0",
            flush=True,
        )
        for finding in result.findings:
            print(f" - {finding}", flush=True)
    for scenario in scenarios:
        print(f"\n=== {scenario.name} ===", flush=True)
        result = run_scenario(client, scenario)
        results.append(result)
        status = "PASS" if result.ok else "FAIL"
        print(
            f"{status} duration={result.duration_s:.2f}s first_event={_fmt_optional(result.first_event_s)} "
            f"turn={result.turn_id} conv={result.conversation_id} events={result.event_count} "
            f"tool_calls={result.plan_tool_calls} batches={len(result.batches)}",
            flush=True,
        )
        for finding in result.findings:
            print(f" - {finding}", flush=True)

    summary = _summary(results)
    print("\n=== Summary ===")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if args.json_out:
        parent_dir = os.path.dirname(args.json_out)
        if parent_dir:
            os.makedirs(parent_dir, exist_ok=True)
        with open(args.json_out, "w", encoding="utf-8") as handle:
            json.dump(
                {"summary": summary, "results": [_result_payload(item) for item in results]},
                handle,
                ensure_ascii=False,
                indent=2,
            )
    return 0 if summary["failed"] == 0 else 1


def _default_xlsx_paths() -> list[str]:
    if not ACCEPTANCE_FIXTURE_DIR.exists():
        return []
    return [str(path) for path in sorted(ACCEPTANCE_FIXTURE_DIR.glob("*.xls*")) if path.suffix.lower() in {".xlsx", ".xlsm"}]


class Client:
    def __init__(self, base_url: str, email: str, password: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.email = email
        self.password = password
        self.token = self._login()

    def _login(self) -> str:
        payload = self.json_request(
            "/api/sico/rbac/login",
            method="POST",
            payload={"email": self.email, "password": self.password},
        )
        token = payload.get("data", {}).get("tokenInfo", {}).get("accessToken", "")
        if not token:
            raise RuntimeError("login did not return an access token")
        return token

    def json_request(
        self,
        path: str,
        *,
        method: str = "GET",
        payload: dict[str, Any] | None = None,
        query: dict[str, Any] | None = None,
        timeout: int = 60,
    ) -> dict[str, Any]:
        body = None if payload is None else json.dumps(payload).encode("utf-8")
        headers = {}
        token = getattr(self, "token", "")
        if token:
            headers["Authorization"] = f"Bearer {token}"
        if body is not None:
            headers["Content-Type"] = "application/json"
        req = urllib.request.Request(self._url(path, query), data=body, headers=headers, method=method)
        raw = _request_bytes(req, timeout)
        if not raw:
            return {}
        data = json.loads(raw.decode("utf-8"))
        code = data.get("code")
        if code not in (None, 0):
            raise RuntimeError(data.get("msg") or f"server returned code={code}")
        return data

    def post_sse(self, path: str, payload: dict[str, Any], *, timeout: int) -> list[tuple[str, str, float]]:
        body = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            self._url(path, None),
            data=body,
            headers={"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"},
            method="POST",
        )
        started = time.perf_counter()
        events: list[tuple[str, str, float]] = []
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            event_name = "message"
            data_lines: list[str] = []
            for raw_line in resp:
                line = raw_line.decode("utf-8", errors="replace").rstrip("\r\n")
                if not line:
                    if data_lines:
                        events.append((event_name, "\n".join(data_lines), time.perf_counter() - started))
                    event_name = "message"
                    data_lines = []
                    continue
                if line.startswith("event:"):
                    event_name = line.removeprefix("event:").strip() or "message"
                elif line.startswith("data:"):
                    data_lines.append(line.removeprefix("data:").lstrip())
            if data_lines:
                events.append((event_name, "\n".join(data_lines), time.perf_counter() - started))
        return events

    def fetch_text(self, uri: str, *, timeout: int = 30) -> str:
        if uri.startswith("file://"):
            with open(uri.removeprefix("file://"), encoding="utf-8") as handle:
                return handle.read()
        req = urllib.request.Request(self._absolute_url(uri), headers={"Authorization": f"Bearer {self.token}"})
        return _request_bytes(req, timeout).decode("utf-8", errors="replace")

    def _url(self, path: str, query: dict[str, Any] | None) -> str:
        url = self._absolute_url(path)
        if query:
            encoded = urllib.parse.urlencode({key: value for key, value in query.items() if value is not None})
            url = f"{url}?{encoded}"
        return url

    def _absolute_url(self, value: str) -> str:
        if value.startswith("http://") or value.startswith("https://"):
            return value
        return f"{self.base_url}{value if value.startswith('/') else '/' + value}"


def run_scenario(client: Client, scenario: Scenario) -> ScenarioResult:
    started = time.perf_counter()
    result = ScenarioResult(
        name=scenario.name,
        ok=False,
        duration_s=0.0,
        agent_instance_id=scenario.agent_instance_id,
        message=scenario.message,
    )
    try:
        with _served_attachments(scenario.attachments) as attachments:
            events = client.post_sse(
                "/api/sico/conversation/chat",
                {"message": scenario.message, "agentInstanceId": scenario.agent_instance_id, "attachments": attachments},
                timeout=int(scenario.max_seconds),
            )
        result.duration_s = time.perf_counter() - started
        result.event_count = len(events)
        result.first_event_s = _first_event_time(events)
        _inspect_sse(events, result, scenario)
        if result.turn_id is None:
            result.findings.append("SSE did not include a turnId")
            return result
        if result.conversation_id:
            batches = _retry(lambda: _batches(client, result.conversation_id or 0, result.turn_id or 0), attempts=6, delay_s=0.5)
            result.batches = batches
            _inspect_batches(client, result, scenario)
        elif scenario.expect_batch:
            result.findings.append("cannot query batch_summaries because conversationId is missing")
        plan = _safe_json(
            lambda: client.json_request(
                "/api/sico/conversation/plan",
                query={
                    "agentInstanceId": scenario.agent_instance_id,
                    "turnId": result.turn_id,
                    "conversationId": result.conversation_id or 0,
                },
            )
        )
        _inspect_plan(plan, result, scenario)
        if scenario.expect_batch and not result.batches:
            result.findings.append("expected at least one delegated batch, but none was persisted")
        if result.duration_s > scenario.max_seconds and not scenario.allow_duration_overrun:
            result.findings.append(f"scenario exceeded max_seconds={scenario.max_seconds:.0f}")
        result.ok = not result.findings
        return result
    except Exception as exc:  # noqa: BLE001 - acceptance script should report scenario-level failures.
        result.duration_s = time.perf_counter() - started
        result.findings.append(f"exception: {type(exc).__name__}: {exc}")
        return result


def run_seeded_agents_preflight(client: Client) -> ScenarioResult:
    started = time.perf_counter()
    result = ScenarioResult(name=SEED_PREFLIGHT_SCENARIO, ok=False, duration_s=0.0)
    try:
        roles_payload = client.json_request("/api/sico/agent/roles")
        roles_data = roles_payload.get("data", {})
        roles = roles_data.get("roles") or roles_data.get("role") or []
        missing_roles = [role for role in EXPECTED_AGENT_ROLES if role not in roles]
        if missing_roles:
            result.findings.append(f"agent roles endpoint missing role(s): {', '.join(missing_roles)}")

        instances_payload = client.json_request(
            "/api/sico/agent/single_agent_instances",
            query={"page": 1, "pageSize": 50},
        )
        instances = instances_payload.get("data", {}).get("instances") or []
        by_id = {_int_or_zero(item.get("id")): item for item in instances if isinstance(item, dict)}
        selected_instances = []
        for instance_id, expected in EXPECTED_SEEDED_AGENT_INSTANCES.items():
            instance = by_id.get(instance_id)
            if not instance:
                result.findings.append(f"seeded agent instance {instance_id} is missing")
                continue
            selected_instances.append(instance)
            for key, expected_value in expected.items():
                actual_value = str(instance.get(key) or "")
                if actual_value != expected_value:
                    result.findings.append(
                        f"seeded agent instance {instance_id} {key}={actual_value!r}, expected {expected_value!r}"
                    )

        result.final_content = json.dumps({"roles": roles, "seededInstances": selected_instances}, ensure_ascii=False, indent=2)
        result.final_content_chars = len(result.final_content)
        result.duration_s = time.perf_counter() - started
        result.ok = not result.findings
        return result
    except Exception as exc:  # noqa: BLE001 - acceptance script should report scenario-level failures.
        result.duration_s = time.perf_counter() - started
        result.findings.append(f"exception: {type(exc).__name__}: {exc}")
        return result


def _inspect_sse(events: list[tuple[str, str, float]], result: ScenarioResult, scenario: Scenario) -> None:
    if not events:
        result.findings.append("SSE returned no events")
        return
    if not any(name == "done" for name, _data, _elapsed in events):
        result.findings.append("SSE missing done event")
    if any(name == "error" for name, _data, _elapsed in events):
        result.findings.append("SSE emitted error event")
    result.final_content = _collect_sse_content_and_ids(events, result)
    result.final_content_chars = len(result.final_content)
    if "Trajectory:" in result.final_content or "trajectory URL" in result.final_content:
        result.findings.append("final answer still surfaces trajectory instead of keeping it inside the execution summary")
    _inspect_file_conversion_final_content(result, scenario)
    if "Metrics report:" in result.final_content:
        result.findings.append("final answer still appends a separate Metrics report link")
    if re.search(r"\bStaleWorkerError\b|stale worker token", result.final_content, re.IGNORECASE):
        result.findings.append("final answer leaks task-runtime stale-worker internals")
    if re.search(r"\b(transient|skill_runtime|sandbox_unhealthy|sandbox_no_capacity)\b", result.final_content, re.IGNORECASE):
        result.findings.append("final answer leaks internal error classification labels")
    final_content_lowered = result.final_content.lower()
    for marker in scenario.expected_final_text:
        if marker.lower() not in final_content_lowered:
            result.findings.append(f"final answer missing expected text: {marker}")
    if scenario.expected_final_any_text and not any(
        marker.lower() in final_content_lowered for marker in scenario.expected_final_any_text
    ):
        result.findings.append("final answer missing any expected text: " + ", ".join(scenario.expected_final_any_text))
    if scenario.name == "detailed_readme_no_execute" and re.search(
        r"android_tester|python\s+-m\s+android_tester|android capability entrypoint",
        result.final_content,
        re.IGNORECASE,
    ):
        result.findings.append("README/docs answer drifted into unrelated Android skill source details")
    if scenario.name == "explicit_source_debug_allowed" and re.search(
        r"\bcontext\b.*\bread\b.*\bgrep\b.*(?:unavailable|not available|aren['’]t available|are not available)",
        result.final_content,
        re.IGNORECASE | re.DOTALL,
    ):
        result.findings.append("source-debug answer claimed workspace inspection tools were unavailable")
    # Delegated/batch turns run an intent-check LLM call plus a planner LLM call before the
    # first persisted plan event, so their first-event latency is legitimately higher and more
    # variable than plain FAST/INSPECT replies. Scenarios that attach a document incur the same
    # kind of pre-first-event cost: the attachment is ingested and parsed (a parse_document tool
    # call) before any user-visible event. Use a looser budget for both to avoid flaky failures
    # on provider-latency spikes while still catching pathological stalls.
    first_event_budget_s = 45.0 if (scenario.expect_batch or scenario.attachments) else 20.0
    if (
        result.first_event_s is not None
        and result.first_event_s > first_event_budget_s
        and not scenario.allow_slow_first_event
    ):
        result.findings.append(f"first SSE event is slow ({result.first_event_s:.2f}s, budget {first_event_budget_s:.0f}s)")


def _collect_sse_content_and_ids(events: list[tuple[str, str, float]], result: ScenarioResult) -> str:
    content_chunks: list[str] = []
    for _name, data, _elapsed in events:
        if not data:
            continue
        try:
            item = json.loads(data)
        except json.JSONDecodeError:
            continue
        if item.get("turnId") and result.turn_id is None:
            result.turn_id = int(item["turnId"])
        if item.get("conversationId") and result.conversation_id is None:
            result.conversation_id = int(item["conversationId"])
        content = item.get("content")
        if isinstance(content, str):
            content_chunks.append(content)
    return "".join(content_chunks)


def _inspect_file_conversion_final_content(result: ScenarioResult, scenario: Scenario) -> None:
    if not scenario.expect_file_conversion:
        return
    lowered = result.final_content.lower()
    if "policy_deny" in lowered or "execution failed" in lowered:
        result.findings.append("file conversion final answer reported policy denial or execution failure")
    if ".csv" not in lowered and "csv" not in lowered:
        result.findings.append("file conversion final answer did not mention the generated CSV")


def _inspect_plan(plan: dict[str, Any] | None, result: ScenarioResult, scenario: Scenario) -> None:
    if not plan:
        result.findings.append("plan endpoint returned no JSON")
        return
    plan_data = plan.get("data", {}).get("plan", {})
    steps = plan_data.get("steps") or []
    tool_calls = _flatten_tool_calls(steps)
    result.plan_tool_calls = len(tool_calls)
    if scenario.max_parse_document_tool_calls is not None:
        parse_calls = sum(1 for tool_call in tool_calls if _tool_call_builtin_name(tool_call) == "parse_document")
        if parse_calls > scenario.max_parse_document_tool_calls:
            result.findings.append(
                f"plan called parse_document {parse_calls} time(s), expected at most {scenario.max_parse_document_tool_calls}"
            )
    if scenario.forbidden_tool_names:
        called_forbidden = sorted(
            {
                tool_name
                for tool_name in (_tool_call_builtin_name(tool_call) for tool_call in tool_calls)
                if tool_name in scenario.forbidden_tool_names
            }
        )
        if called_forbidden:
            result.findings.append(f"plan called forbidden tool(s): {', '.join(called_forbidden)}")
    _inspect_parse_document_targets(tool_calls, result)
    serialized = json.dumps(plan, ensure_ascii=False)
    _inspect_plan_text(serialized, result, scenario)
    _inspect_plan_lifecycle_statuses(tool_calls, result)
    for step in steps:
        for tool in step.get("toolCalls") or []:
            if not isinstance(tool, dict) or not tool.get("batchCalls"):
                continue
            parent_message = str(tool.get("message") or "")
            if "Trajectory:" in parent_message or "runner:" in parent_message:
                result.findings.append("batch parent message duplicates per-case execution details")


def _inspect_plan_text(serialized: str, result: ScenarioResult, scenario: Scenario) -> None:
    _inspect_plan_required_text(serialized, result, scenario)
    _inspect_plan_forbidden_text(serialized, result, scenario)
    _inspect_plan_lifecycle_text(serialized, result)


def _inspect_plan_required_text(serialized: str, result: ScenarioResult, scenario: Scenario) -> None:
    lowered = serialized.lower()
    for marker in scenario.expected_plan_text:
        if marker.lower() not in lowered:
            result.findings.append(f"plan missing expected text: {marker}")
    if scenario.expect_batch and "run_task" not in serialized:
        result.findings.append("plan does not show run_task for delegated scenario")
    if scenario.expect_run_report and "report:" not in lowered and "run report:" not in lowered:
        result.findings.append("plan missing run report URL")
    if scenario.expect_attachment_parse and "parse_document" not in serialized:
        result.findings.append("plan does not show parse_document for attachment-driven scenario")
    if scenario.expect_no_attachment_parse and "parse_document" in serialized:
        result.findings.append("plan unexpectedly parsed an attachment for a no-attachment scenario")
    if scenario.expect_file_conversion and "file_convert" not in serialized:
        result.findings.append("plan does not show file_convert for delegated file conversion")


def _inspect_plan_forbidden_text(serialized: str, result: ScenarioResult, scenario: Scenario) -> None:
    lowered = serialized.lower()
    for marker in scenario.forbidden_plan_text:
        if marker.lower() in lowered:
            result.findings.append(f"plan includes forbidden text: {marker}")
    if scenario.expect_no_low_level_android and (
        "android_tester.__main__" in serialized or "python -m android_tester run" in serialized
    ):
        result.findings.append("plan leaks low-level android_tester command details")
    if "Command:" in serialized or "Entrypoint:" in serialized:
        result.findings.append("plan leaks low-level Command/Entrypoint labels")
    if 'Summary: {"event"' in serialized or "Summary: {'event'" in serialized:
        result.findings.append("plan leaks raw runner JSON event output as a case summary")
    if "batch report:" in serialized.lower():
        result.findings.append("plan still labels delegate execution summary as a batch report")
    if re.search(r"\bStaleWorkerError\b|stale worker token", serialized, re.IGNORECASE):
        result.findings.append("plan leaks task-runtime stale-worker internals")
    if scenario.expect_single_case and (
        "Progress group 1" in serialized or "up to 1 concurrent" in serialized or "up to 1 emulator task" in serialized
    ):
        result.findings.append("single-case plan still uses batch/concurrency wording")
    if scenario.expect_batch and "Skipped source read for normal execution" in serialized:
        result.findings.append("normal task execution still attempted to read skill/playbook source")


def _inspect_plan_lifecycle_text(serialized: str, result: ScenarioResult) -> None:
    if re.search(r"Prepare workspace[^\n\r)]*running", serialized):
        result.findings.append("workspace preparation row still includes execution-running wording")
    if re.search(r"(?:Prepare|Allocating|Resetting) [^\n\r)]*sandbox[^\n\r)]*Executing", serialized):
        result.findings.append("sandbox preparation row still includes execution-running wording")
    if re.search(r"Collect results and release [^\n\r)]*sandbox[^\n\r)]*released[^\n\r)]*0/\d+ .* finished", serialized):
        result.findings.append("batch finalization row reports released sandboxes before completed cases")
    if re.search(r"尝试次数：\d+/\d+", serialized):
        result.findings.append("plan uses ambiguous attempt count copy")


def _inspect_plan_lifecycle_statuses(tool_calls: list[dict[str, Any]], result: ScenarioResult) -> None:
    for tool_call in tool_calls:
        message = str(tool_call.get("message") or "")
        statuses = [item.get("status") for item in tool_call.get("runningList") or [] if isinstance(item, dict)]
        if message.startswith("Finished ") and any(status in ACTIVE_RUNNING_LIST_STATUSES for status in statuses):
            result.findings.append("finished batch parent still has pending/running lifecycle rows")
        if _is_terminal_task_run_message(message) and any(status in ACTIVE_RUNNING_LIST_STATUSES for status in statuses):
            result.findings.append("terminal child task still has active lifecycle rows")


def _inspect_parse_document_targets(tool_calls: list[dict[str, Any]], result: ScenarioResult) -> None:
    for tool_call in tool_calls:
        if _tool_call_builtin_name(tool_call) != "parse_document":
            continue
        target = _parse_document_target(str(tool_call.get("message") or ""))
        if target and not target.startswith(("attachments/", "download/")):
            result.findings.append(f"parse_document used for non-attachment path: {target}")


def _parse_document_target(message: str) -> str:
    prefixes = (
        "Parsing document: ",
        "Parsed document: ",
        "file not found: ",
    )
    for prefix in prefixes:
        if message.startswith(prefix):
            value = message[len(prefix) :].strip()
            return value.split(" (", 1)[0].strip()
    return ""


def _is_terminal_task_run_message(message: str) -> bool:
    return message.startswith(
        (
            "Android test completed:",
            "Android test failed:",
            "Skill task completed:",
            "Skill task failed:",
            "Skill task cancelled:",
            "Skill task timed out:",
            "Skill task blocked:",
            "Task completed:",
            "Task failed:",
            "Task cancelled:",
            "Task timed out:",
            "Task blocked:",
        )
    )


def _inspect_batches(client: Client, result: ScenarioResult, scenario: Scenario) -> None:
    _inspect_batch_collection_shape(result, scenario)
    for batch in result.batches:
        _inspect_single_batch(client, result, scenario, batch)


def _inspect_batch_collection_shape(result: ScenarioResult, scenario: Scenario) -> None:
    max_batches = scenario.max_batches
    if max_batches is None and scenario.expect_batch:
        max_batches = 1
    if max_batches is not None and len(result.batches) > max_batches:
        result.findings.append(f"expected at most {max_batches} batch(es), got {len(result.batches)}")
    if scenario.min_batch_total is not None:
        largest_total = max((int(batch.get("totalCount") or 0) for batch in result.batches), default=0)
        if largest_total < scenario.min_batch_total:
            result.findings.append(
                f"expected a batch with at least {scenario.min_batch_total} task(s), largest batch had {largest_total}"
            )


def _inspect_single_batch(client: Client, result: ScenarioResult, scenario: Scenario, batch: dict[str, Any]) -> None:
    if batch.get("status") not in {"completed", "partial", "failed", "blocked", "timed_out", "cancelled"}:
        result.findings.append(f"batch {batch.get('batchId')} has non-terminal status {batch.get('status')}")
    if scenario.expect_batch and batch.get("status") not in {"completed", "partial"} and not scenario.allow_failed_batch:
        result.findings.append(f"batch {batch.get('batchId')} ended with {batch.get('status')}")
    if scenario.expect_file_conversion and batch.get("status") != "completed":
        result.findings.append(f"file conversion batch {batch.get('batchId')} ended with {batch.get('status')}")
    # Per-batch result detail now renders inline in the plan progress (runningList / batchCalls);
    # the standalone summary.html report and turn_metrics endpoint were removed from the product.
    # Inline plan progress is validated by _inspect_plan; batch persistence/status/counts are
    # validated above and by _inspect_batch_collection_shape.


def _batches(client: Client, conversation_id: int, turn_id: int) -> list[dict[str, Any]]:
    resp = client.json_request(
        "/api/sico/conversation/batch_summaries",
        query={"conversationId": conversation_id, "turnId": turn_id, "page": 1, "pageSize": 20},
    )
    return resp.get("data", {}).get("items") or []


def _flatten_tool_calls(steps: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []

    def walk(tool: dict[str, Any]) -> None:
        result.append(tool)
        for child in tool.get("batchCalls") or []:
            if isinstance(child, dict):
                walk(child)

    for step in steps:
        for tool in step.get("toolCalls") or []:
            if isinstance(tool, dict):
                walk(tool)
    return result


def _tool_call_builtin_name(tool_call: dict[str, Any]) -> str:
    execution_info = tool_call.get("executionInfo") or {}
    if isinstance(execution_info, dict):
        builtin_name = execution_info.get("builtinToolName")
        if builtin_name:
            return str(builtin_name)
    return str(tool_call.get("toolName") or "")


def _retry(func, *, attempts: int, delay_s: float):
    last_value = None
    for attempt in range(attempts):
        last_value = func()
        if last_value:
            return last_value
        if attempt < attempts - 1:
            time.sleep(delay_s)
    return last_value


def _safe_json(func) -> dict[str, Any] | None:
    try:
        return func()
    except Exception:
        return None


def _first_event_time(events: list[tuple[str, str, float]]) -> float | None:
    for name, _data, elapsed in events:
        if name != "keepalive":
            return elapsed
    return None


def _summary(results: list[ScenarioResult]) -> dict[str, Any]:
    durations = [result.duration_s for result in results]
    return {
        "total": len(results),
        "passed": sum(1 for result in results if result.ok),
        "failed": sum(1 for result in results if not result.ok),
        "avgDurationS": round(statistics.mean(durations), 3) if durations else 0,
        "p95DurationS": round(_percentile(durations, 95), 3) if durations else 0,
        "failedScenarios": [result.name for result in results if not result.ok],
    }


def _result_payload(result: ScenarioResult) -> dict[str, Any]:
    return {
        "name": result.name,
        "ok": result.ok,
        "agentInstanceId": result.agent_instance_id,
        "message": result.message,
        "durationS": round(result.duration_s, 3),
        "firstEventS": None if result.first_event_s is None else round(result.first_event_s, 3),
        "turnId": result.turn_id,
        "conversationId": result.conversation_id,
        "eventCount": result.event_count,
        "finalContentChars": result.final_content_chars,
        "planToolCalls": result.plan_tool_calls,
        "batchCount": len(result.batches),
        "batches": result.batches,
        "metrics": result.metrics,
        "finalContentPreview": result.final_content[-800:],
        "findings": result.findings,
    }


def _percentile(values: list[float], percentile: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    index = min(len(ordered) - 1, max(0, int(round((percentile / 100) * (len(ordered) - 1)))))
    return ordered[index]


def _request_bytes(req: urllib.request.Request, timeout: int) -> bytes:
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read()
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code} {exc.reason}: {details}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"request failed: {exc}") from exc


def _fmt_optional(value: float | None) -> str:
    return "n/a" if value is None else f"{value:.2f}s"


def _int_or_zero(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


@contextlib.contextmanager
def _served_attachments(paths: tuple[str, ...]):
    if not paths:
        yield []
        return
    resolved = [Path(path).expanduser().resolve() for path in paths]
    missing = [str(path) for path in resolved if not path.is_file()]
    if missing:
        raise RuntimeError(f"attachment file(s) not found: {', '.join(missing)}")
    roots = {path.parent for path in resolved}
    if len(roots) != 1:
        raise RuntimeError("all local attachments for one scenario must live in the same directory")
    root = next(iter(roots))

    class QuietHandler(SimpleHTTPRequestHandler):
        def log_message(self, _format: str, *args: Any) -> None:
            return

    handler = functools.partial(QuietHandler, directory=str(root))
    server = ThreadingHTTPServer(("0.0.0.0", 0), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        base_url = f"http://host.docker.internal:{server.server_port}"
        yield [_attachment_payload(path, base_url) for path in resolved]
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def _attachment_payload(path: Path, base_url: str) -> dict[str, Any]:
    content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    uri = f"{base_url}/{urllib.parse.quote(path.name)}"
    return {
        "name": path.name,
        "uri": uri,
        "sasUrl": uri,
        "type": content_type,
        "size": path.stat().st_size,
    }


def _scenarios(xlsx_paths: list[str] | None = None) -> list[Scenario]:
    scenarios = [
        Scenario(
            name="max_tester_hi",
            agent_instance_id=2,
            message="hi",
            max_seconds=90,
        ),
        Scenario(
            name="max_smoke_fast",
            agent_instance_id=2,
            message="Reply exactly: Max Tester smoke ok",
            max_seconds=90,
        ),
        Scenario(
            name="single_delegate_echo",
            agent_instance_id=2,
            message=(
                "Use the delegate tool (kind=general) to execute one isolated local echo test case. "
                "Task title: Acceptance single echo. Echo message: single delegate ok. "
                "Return the execution summary URL if one is produced."
            ),
            expect_batch=True,
            max_seconds=180,
        ),
        Scenario(
            name="parallel_delegate_three_cases",
            agent_instance_id=2,
            message=(
                "Use the delegate tool (kind=general) to execute three independent local echo test cases in one batch. "
                "Case 1 message: alpha ok. Case 2 message: beta ok. Case 3 message: gamma ok. "
                "Do not inspect skill source before delegating."
            ),
            expect_batch=True,
            max_seconds=180,
        ),
        Scenario(
            name="partial_ok_mixed_echo_batch",
            agent_instance_id=2,
            message=(
                "Use the delegate tool (kind=general) to run two independent local echo cases in one batch. "
                "Case A should echo stable pass. Case B should echo stable second pass. "
                "Summarize the structured digest of both case results."
            ),
            expect_batch=True,
            max_seconds=180,
        ),
        Scenario(
            name="android_label_probe",
            agent_instance_id=2,
            message=(
                "Use the Android tester capability card through the delegate tool (kind=general) for one minimal Android UI test case. "
                "Case title: Copilot Android label smoke. Steps: launch Copilot, observe the first screen, then stop. "
                "This is an acceptance probe for plan wording and report generation."
            ),
            expect_batch=True,
            allow_failed_batch=True,
            allow_duration_overrun=True,
            allow_slow_first_event=True,
            max_seconds=420,
        ),
        Scenario(
            name="android_apk_install_launch_uninstall",
            agent_instance_id=2,
            message=(
                f"Use the Android tester capability card through the delegate tool (kind=general). Install the APK from {COPILOT_APK_URL}, "
                "launch Copilot, "
                "observe the first screen, then uninstall Copilot. Return only the execution result "
                "and run report URL."
            ),
            expect_batch=True,
            expect_run_report=True,
            allow_duration_overrun=True,
            allow_slow_first_event=True,
            max_batches=1,
            max_seconds=1500,
        ),
        Scenario(
            name="android_accessible_link_install_launch_uninstall",
            agent_instance_id=2,
            message=(
                "Use the Android tester capability card through the delegate tool (kind=general). Install Copilot from this accessible link: "
                f"{COPILOT_APK_URL}. Treat the link as the installation source, launch Copilot, "
                "observe the first screen, then uninstall Copilot. Return only the execution result "
                "and run report URL."
            ),
            expect_batch=True,
            expect_run_report=True,
            allow_duration_overrun=True,
            allow_slow_first_event=True,
            max_batches=1,
            max_seconds=1500,
        ),
        Scenario(
            name="delegate_first_no_source_read",
            agent_instance_id=2,
            message=(
                "For this executable tester capability, delegate first without reading SKILL.md or source files. "
                "Run one local echo task titled Delegate first check with message delegate first ok."
            ),
            expect_batch=True,
            max_seconds=180,
        ),
        Scenario(
            name="batch_report_metrics_probe",
            agent_instance_id=2,
            message=(
                "Use the delegate tool (kind=general) to run a local echo batch with two cases. "
                "Case timing-summary-1 message: timing one ok. Case timing-summary-2 message: timing two ok. "
                "Summarize the structured digest of both case results."
            ),
            expect_batch=True,
            max_seconds=180,
        ),
        Scenario(
            name="ambiguous_no_context_no_execute",
            agent_instance_id=2,
            message=(
                "Please test it. I have not provided a target, file, URL, or case ID. "
                "If the request is underspecified, ask for the missing target instead of running the delegate tool."
            ),
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="case_id_without_source_no_execute",
            agent_instance_id=2,
            message=(
                "请执行测试用例 SICO-UNKNOWN-001。当前请求没有提供附件、用例内容或其他可访问来源。"
                "如果找不到可靠来源，请询问用户补充用例内容，不要执行测试。"
            ),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            expect_no_attachment_parse=True,
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="detailed_readme_no_execute",
            agent_instance_id=2,
            message=(
                "Read README.md and docs/quickstart.md if useful, then summarize local startup commands. "
                "Do not run commands, do not run tests, and do not delegate this as a task."
            ),
            allow_slow_first_event=True,
            max_seconds=240,
        ),
        Scenario(
            name="explicit_source_debug_allowed",
            agent_instance_id=2,
            message=(
                "Debug the duplicate delegation guard implementation by inspecting workspace source only. "
                "Use context/read/grep if available. Do not execute tests and do not delegate. "
                "If the relevant source is not present in the workspace, say that."
            ),
            allow_slow_first_event=True,
            max_seconds=240,
        ),
        Scenario(
            name="detailed_four_echo_no_sandbox",
            agent_instance_id=2,
            message=(
                "Use the delegate tool (kind=general) for exactly four independent local echo tool cases in one batch, no sandbox. "
                "Messages: matrix one ok, matrix two ok, matrix three ok, matrix four ok. "
                "Summarize the digest of all four case results."
            ),
            expect_batch=True,
            min_batch_total=4,
            max_batches=1,
            max_seconds=240,
        ),
        Scenario(
            name="tool_choice_explanation_no_action",
            agent_instance_id=2,
            message=(
                "Do not execute anything. For a hypothetical request to install an Android APK and verify launch, "
                "briefly explain whether you would choose a pluggable skill, delegation, or generic file tools, and why."
            ),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="web_test_explanation_no_action",
            agent_instance_id=2,
            message=(
                "Do not execute the web test. Analyze how you would validate checkout in Playwright, what evidence you "
                "would collect, and whether delegation would be appropriate."
            ),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="product_research_no_action",
            agent_instance_id=2,
            message=("不要执行，只分析一个产品调研方案：比较 Copilot 类产品时需要看哪些指标、信息来源和输出结构。"),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="luna_3d_brief_no_action",
            agent_instance_id=3,
            message=(
                "Do not execute anything. As Luna the 3D Artist, outline the information you need before creating "
                "a game-ready low-poly robot mascot model. Keep it concise."
            ),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            expected_final_text=("mascot",),
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="luna_3d_skill_guidance_no_action",
            agent_instance_id=3,
            message=(
                "Do not execute anything. As Luna, use the ai-3d-model skill guidance to outline a compact production "
                "brief for a game-ready GLB of a friendly low-poly robot mascot with clean topology constraints."
            ),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            expected_final_text=("GLB", "topology"),
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="luna_local_echo_delegate",
            agent_instance_id=3,
            message=(
                "Use the delegate tool (kind=general) to run one local echo task titled Luna generic result file with message "
                "luna generic result ok. Return the execution summary URL if one is produced."
            ),
            expect_batch=True,
            forbidden_plan_text=("android_tester", "python -m android_tester run"),
            max_batches=1,
            max_seconds=180,
        ),
        Scenario(
            name="ethan_prd_outline_no_action",
            agent_instance_id=4,
            message=(
                "Do not execute anything. As Ethan the Product Manager, draft a concise PRD outline for a Team Inbox "
                "Automation feature, including problem, users, success metrics, and risks."
            ),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            expected_final_text=("PRD", "success"),
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="ethan_frontend_slides_guidance_no_action",
            agent_instance_id=4,
            message=(
                "Do not execute anything. As Ethan, use the frontend-slides skill guidance to outline a compact HTML "
                "slide deck for Team Inbox Automation with three slides: problem, workflow, and rollout."
            ),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            expected_final_text=("slide", "workflow"),
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="ethan_local_echo_delegate",
            agent_instance_id=4,
            message=(
                "Use the delegate tool (kind=general) to run one local echo task titled Ethan generic result file with message "
                "ethan generic result ok. Return the execution summary URL if one is produced."
            ),
            expect_batch=True,
            forbidden_plan_text=("android_tester", "python -m android_tester run"),
            max_batches=1,
            max_seconds=180,
        ),
        Scenario(
            name="chloe_positioning_no_action",
            agent_instance_id=5,
            message=(
                "Do not execute anything. As Chloe the Marketing agent, write concise positioning for a B2B Team Inbox "
                "Automation product: audience, category, differentiated value, and tagline."
            ),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            expected_final_text=("audience", "tagline"),
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="chloe_image_generation_prompt_no_action",
            agent_instance_id=5,
            message=(
                "Do not execute anything. As Chloe, use the image-generator skill guidance to write a production-ready "
                "1024x1024 launch poster prompt for Team Inbox Automation with clean SaaS visual direction."
            ),
            forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
            expected_final_text=("1024x1024", "prompt"),
            allow_slow_first_event=True,
            max_seconds=180,
        ),
        Scenario(
            name="chloe_local_echo_delegate",
            agent_instance_id=5,
            message=(
                "Use the delegate tool (kind=general) to run one local echo task titled Chloe generic result file with message "
                "chloe generic result ok. Return the execution summary URL if one is produced."
            ),
            expect_batch=True,
            forbidden_plan_text=("android_tester", "python -m android_tester run"),
            max_batches=1,
            max_seconds=180,
        ),
        Scenario(
            name="repeatability_echo_run_1",
            agent_instance_id=2,
            message="Use the delegate tool (kind=general) to run one local echo test titled Repeatability one with message repeatability one ok.",
            expect_batch=True,
            max_seconds=180,
        ),
        Scenario(
            name="repeatability_echo_run_2",
            agent_instance_id=2,
            message="Use the delegate tool (kind=general) to run one local echo test titled Repeatability two with message repeatability two ok.",
            expect_batch=True,
            max_seconds=180,
        ),
        Scenario(
            name="concise_final_response",
            agent_instance_id=2,
            message=(
                "Use the delegate tool (kind=general) to run one local echo test titled Concise final response with message concise ok. "
                "Keep the final answer concise."
            ),
            expect_batch=True,
            max_seconds=180,
        ),
    ]
    for index, raw_path in enumerate(xlsx_paths or [], start=1):
        file_name = Path(raw_path).name
        path = Path(raw_path)
        workbook_profile = _xlsx_workbook_profile(path)
        expected_rows = int(workbook_profile.get("runnable_data_rows") or 0) if workbook_profile else _xlsx_data_row_count(path)
        scenarios.append(
            Scenario(
                name=f"xlsx_to_csv_delegate_{index}",
                agent_instance_id=2,
                message=f"请使用 delegate 工具（kind=general）把附件 {file_name} 转换为 CSV，并返回生成文件 URL。",
                expect_batch=True,
                expect_file_conversion=True,
                max_batches=1,
                max_parse_document_tool_calls=0,
                max_seconds=240,
                attachments=(raw_path,),
            )
        )
        if workbook_profile and workbook_profile.get("requires_scope_selection"):
            scenarios.append(
                Scenario(
                    name=f"xlsx_multisheet_scope_prompt_{index}",
                    agent_instance_id=2,
                    message=f"帮我测试 {file_name}",
                    max_batches=0,
                    max_parse_document_tool_calls=1,
                    forbidden_tool_names=FORBIDDEN_DELEGATION_TOOLS,
                    expected_final_any_text=("sheet", "tab", "工作表", "范围", "哪"),
                    max_seconds=240,
                    attachments=(raw_path,),
                )
            )
            selected_sheet = _xlsx_preferred_run_sheet(workbook_profile)
            if selected_sheet:
                sheet_name = str(selected_sheet["name"])
                sheet_rows = int(selected_sheet.get("data_rows") or 0) or None
                download_sheet = _xlsx_named_data_sheet(workbook_profile, "rewritten_download")
                scenarios.append(
                    Scenario(
                        name=f"xlsx_android_followup_sheet_name_only_{index}",
                        agent_instance_id=2,
                        message=f"跑 {sheet_name}",
                        expect_batch=True,
                        expect_run_report=True,
                        expect_no_attachment_parse=True,
                        allow_failed_batch=True,
                        allow_duration_overrun=True,
                        min_batch_total=sheet_rows,
                        max_batches=1,
                        max_seconds=XLSX_ANDROID_MAX_SECONDS,
                    )
                )
                if download_sheet and str(download_sheet.get("name") or "") != sheet_name:
                    download_sheet_name = str(download_sheet["name"])
                    download_sheet_rows = int(download_sheet.get("data_rows") or 0) or None
                    scenarios.append(
                        Scenario(
                            name=f"xlsx_android_followup_download_sheet_name_only_{index}",
                            agent_instance_id=2,
                            message=f"跑 {download_sheet_name}",
                            expect_batch=True,
                            expect_run_report=True,
                            expect_no_attachment_parse=True,
                            allow_failed_batch=True,
                            allow_duration_overrun=True,
                            min_batch_total=download_sheet_rows,
                            max_batches=1,
                            max_seconds=XLSX_ANDROID_MAX_SECONDS,
                        )
                    )
                scenarios.append(
                    Scenario(
                        name=f"xlsx_android_run_explicit_sheet_{index}",
                        agent_instance_id=2,
                        message=f"帮我跑 {file_name} 中 {sheet_name} 工作表里的测试用例",
                        expect_batch=True,
                        expect_run_report=True,
                        allow_failed_batch=True,
                        allow_duration_overrun=True,
                        min_batch_total=sheet_rows,
                        max_batches=1,
                        max_parse_document_tool_calls=1,
                        max_seconds=XLSX_ANDROID_MAX_SECONDS,
                        attachments=(raw_path,),
                    )
                )
                scenarios.append(
                    Scenario(
                        name=f"xlsx_android_repeat_explicit_sheet_{index}",
                        agent_instance_id=2,
                        message=f"请重跑 {file_name} 中 {sheet_name} 工作表里的测试用例",
                        expect_batch=True,
                        expect_run_report=True,
                        expect_no_attachment_parse=True,
                        allow_failed_batch=True,
                        allow_duration_overrun=True,
                        allow_slow_first_event=True,
                        min_batch_total=sheet_rows,
                        max_batches=1,
                        max_seconds=XLSX_ANDROID_MAX_SECONDS,
                    )
                )
            continue
        scenarios.append(
            Scenario(
                name=f"xlsx_android_test_workbook_{index}",
                agent_instance_id=2,
                message=f"帮我测试 {file_name}",
                expect_batch=True,
                expect_run_report=True,
                allow_failed_batch=True,
                allow_duration_overrun=True,
                min_batch_total=expected_rows,
                max_batches=1,
                max_parse_document_tool_calls=1,
                max_seconds=XLSX_ANDROID_MAX_SECONDS,
                attachments=(raw_path,),
            )
        )
        scenarios.append(
            Scenario(
                name=f"xlsx_android_run_workbook_{index}",
                agent_instance_id=2,
                message=f"帮我跑 {file_name} 里的测试用例",
                expect_batch=True,
                expect_run_report=True,
                allow_failed_batch=True,
                allow_duration_overrun=True,
                min_batch_total=expected_rows,
                max_batches=1,
                max_parse_document_tool_calls=1,
                max_seconds=XLSX_ANDROID_MAX_SECONDS,
                attachments=(raw_path,),
            )
        )
        scenarios.append(
            Scenario(
                name=f"xlsx_android_repeat_previous_{index}",
                agent_instance_id=2,
                message=f"请重跑 {file_name} 里的测试用例",
                expect_batch=True,
                expect_run_report=True,
                expect_no_attachment_parse=True,
                allow_failed_batch=True,
                allow_duration_overrun=True,
                allow_slow_first_event=True,
                min_batch_total=expected_rows,
                max_batches=1,
                max_seconds=XLSX_ANDROID_MAX_SECONDS,
            )
        )
    return scenarios


def _xlsx_workbook_profile(path: Path) -> dict[str, Any] | None:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    try:
        from openpyxl import load_workbook
    except Exception:
        return _xlsx_workbook_profile_zip(path)
    with contextlib.suppress(Exception):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheets = []
            for worksheet in workbook.worksheets:
                sheets.append(_xlsx_worksheet_profile(worksheet.title, worksheet.iter_rows(values_only=True)))
            return _xlsx_profile_from_sheets(sheets)
        finally:
            workbook.close()
    return _xlsx_workbook_profile_zip(path)


def _xlsx_profile_from_sheets(sheets: list[dict[str, Any]]) -> dict[str, Any]:
    data_sheets = [sheet for sheet in sheets if sheet["kind"] == "data"]
    master_sheets = [sheet for sheet in sheets if sheet["kind"] == "master"]
    summary_sheets = [sheet for sheet in sheets if sheet["kind"] == "summary"]
    executable_sheets = [
        sheet for sheet in sheets if sheet["kind"] in {"data", "master"} and int(sheet.get("data_rows") or 0) > 0
    ]
    source_data_rows = _xlsx_sum_sheet_rows(data_sheets)
    master_data_rows = _xlsx_sum_sheet_rows(master_sheets)
    return {
        "sheets": sheets,
        "total_data_rows": _xlsx_sum_sheet_rows(sheets),
        "runnable_data_rows": source_data_rows if source_data_rows > 0 else master_data_rows,
        "source_data_rows": source_data_rows,
        "master_data_rows": master_data_rows,
        "summary_data_rows": _xlsx_sum_sheet_rows(summary_sheets),
        "requires_scope_selection": len(executable_sheets) > 1,
        "multiple_data_sheets": len(data_sheets) > 1,
    }


def _xlsx_workbook_profile_zip(path: Path) -> dict[str, Any] | None:
    with contextlib.suppress(Exception):
        with zipfile.ZipFile(path) as archive:
            shared_strings = _xlsx_shared_strings(archive)
            sheets = [
                _xlsx_worksheet_profile_xml(sheet_name, archive, sheet_path, shared_strings)
                for sheet_name, sheet_path in _xlsx_sheet_entries(archive)
            ]
            if sheets:
                return _xlsx_profile_from_sheets(sheets)
    return None


def _xlsx_sheet_entries(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    relationships = _xlsx_workbook_relationships(archive)
    entries: list[tuple[str, str]] = []
    for sheet in workbook_root.iter():
        if _xml_local_name(sheet.tag) != "sheet":
            continue
        sheet_name = sheet.attrib.get("name", "").strip()
        relationship_id = next((value for key, value in sheet.attrib.items() if _xml_local_name(key) == "id"), "")
        target = relationships.get(relationship_id, "")
        if not sheet_name or not target:
            continue
        entries.append((sheet_name, _xlsx_workbook_target_path(target)))
    return entries


def _xlsx_workbook_relationships(archive: zipfile.ZipFile) -> dict[str, str]:
    relationships_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationships: dict[str, str] = {}
    for relationship in relationships_root.iter():
        if _xml_local_name(relationship.tag) != "Relationship":
            continue
        relationship_id = relationship.attrib.get("Id", "")
        target = relationship.attrib.get("Target", "")
        relationship_type = relationship.attrib.get("Type", "")
        if relationship_id and target and relationship_type.endswith("/worksheet"):
            relationships[relationship_id] = target
    return relationships


def _xlsx_workbook_target_path(target: str) -> str:
    if target.startswith("/"):
        return posixpath.normpath(target.lstrip("/"))
    if target.startswith("xl/"):
        return posixpath.normpath(target)
    return posixpath.normpath(posixpath.join("xl", target))


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    with contextlib.suppress(KeyError):
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
        strings: list[str] = []
        for item in root.iter():
            if _xml_local_name(item.tag) != "si":
                continue
            parts = [element.text or "" for element in item.iter() if _xml_local_name(element.tag) == "t"]
            strings.append("".join(parts).strip())
        return strings
    return []


def _xlsx_worksheet_profile_xml(
    name: str,
    archive: zipfile.ZipFile,
    sheet_path: str,
    shared_strings: list[str],
) -> dict[str, Any]:
    non_empty_rows = 0
    headers: list[str] = []
    with archive.open(sheet_path) as stream:
        for _, element in ET.iterparse(stream, events=("end",)):
            if _xml_local_name(element.tag) != "row":
                continue
            values = [_xlsx_cell_text(cell, shared_strings) for cell in element if _xml_local_name(cell.tag) == "c"]
            values = [value for value in values if value]
            if values:
                non_empty_rows += 1
                if not headers:
                    headers = values[:WORKBOOK_HEADER_LIMIT]
            element.clear()
    data_rows = max(0, non_empty_rows - 1)
    return {
        "name": name,
        "kind": _classify_xlsx_sheet(name, headers, data_rows),
        "data_rows": data_rows,
        "headers": headers,
    }


def _xlsx_cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return "".join(element.text or "" for element in cell.iter() if _xml_local_name(element.tag) == "t").strip()
    raw_value = next((element.text or "" for element in cell.iter() if _xml_local_name(element.tag) == "v"), "").strip()
    if not raw_value:
        return ""
    if cell_type == "s":
        with contextlib.suppress(ValueError, IndexError):
            return shared_strings[int(raw_value)].strip()
    return raw_value


def _xlsx_worksheet_profile(name: str, rows: Any) -> dict[str, Any]:
    non_empty_rows = 0
    headers: list[str] = []
    for row in rows:
        values = [_stringify_xlsx_cell(cell) for cell in row]
        values = [value for value in values if value]
        if not values:
            continue
        non_empty_rows += 1
        if not headers:
            headers = values[:WORKBOOK_HEADER_LIMIT]
    data_rows = max(0, non_empty_rows - 1)
    return {
        "name": name,
        "kind": _classify_xlsx_sheet(name, headers, data_rows),
        "data_rows": data_rows,
        "headers": headers,
    }


def _stringify_xlsx_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _classify_xlsx_sheet(name: str, headers: list[str], data_rows: int) -> str:
    if data_rows <= 0:
        return "empty"
    lowered_name = name.strip().lower()
    lowered_headers = {header.strip().lower() for header in headers if header.strip()}
    if lowered_name in {"summary", "readme", "overview"} or (lowered_headers and lowered_headers <= {"metric", "value"}):
        return "summary"
    if lowered_name in {"master", "all", "combined"} or {"source_file", "source_row"}.issubset(lowered_headers):
        return "master"
    return "data"


def _xlsx_sum_sheet_rows(sheets: list[dict[str, Any]]) -> int:
    return sum(int(sheet.get("data_rows") or 0) for sheet in sheets)


def _xlsx_preferred_run_sheet(profile: dict[str, Any]) -> dict[str, Any] | None:
    sheets = [sheet for sheet in profile.get("sheets") or [] if isinstance(sheet, dict) and int(sheet.get("data_rows") or 0) > 0]
    for kind in ("data", "master"):
        candidates = [sheet for sheet in sheets if sheet.get("kind") == kind]
        if candidates:
            return min(candidates, key=lambda sheet: int(sheet.get("data_rows") or 0))
    return None


def _xlsx_named_data_sheet(profile: dict[str, Any], name: str) -> dict[str, Any] | None:
    wanted = name.strip().lower()
    for sheet in profile.get("sheets") or []:
        if not isinstance(sheet, dict) or sheet.get("kind") != "data":
            continue
        if str(sheet.get("name") or "").strip().lower() == wanted and int(sheet.get("data_rows") or 0) > 0:
            return sheet
    return None


def _xlsx_data_row_count(path: Path) -> int | None:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    openpyxl_count = _xlsx_data_row_count_openpyxl(path)
    if openpyxl_count is not None:
        return openpyxl_count
    return _xlsx_data_row_count_zip(path)


def _xlsx_data_row_count_openpyxl(path: Path) -> int | None:
    profile = _xlsx_workbook_profile(path)
    if profile is None:
        return None
    return int(profile.get("runnable_data_rows") or 0)


def _xlsx_data_row_count_zip(path: Path) -> int | None:
    with contextlib.suppress(Exception):
        total_rows = 0
        with zipfile.ZipFile(path) as archive:
            sheet_names = sorted(name for name in archive.namelist() if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name))
            for sheet_name in sheet_names:
                non_empty_rows = 0
                with archive.open(sheet_name) as stream:
                    for _, element in ET.iterparse(stream, events=("end",)):
                        if _xml_local_name(element.tag) == "row":
                            if _xlsx_row_has_value(element):
                                non_empty_rows += 1
                            element.clear()
                if non_empty_rows > 1:
                    total_rows += non_empty_rows - 1
        return total_rows
    return None


def _xlsx_row_has_value(row: ET.Element) -> bool:
    for element in row.iter():
        if _xml_local_name(element.tag) in {"v", "t"} and element.text and element.text.strip():
            return True
    return False


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


if __name__ == "__main__":
    sys.exit(main())
