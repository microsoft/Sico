from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from pydantic import ValidationError

from app.biz.chat.preparation import (
    DelegatePreparationService,
    LlmTaskPlanner,
    NeedsClarification,
    PlannedDecision,
    PlannerOutput,
    PreparationError,
    Rejected,
)
from app.biz.chat.preparation.catalogue import CapabilityCatalogue, WorkspaceCapabilityCatalogue
from app.biz.chat.preparation.request import MAX_DELEGATE_SOURCES, MAX_DELEGATE_WORK_ITEMS, parse_delegate_request
from app.biz.chat.preparation.service import _truncate_description
from app.biz.chat.preparation.tabular import (
    ArgumentBinder,
    BindingRule,
)
from app.biz.chat.preparation.tabular.planner import (
    BindingDecision,
    LlmTabularPlanner,
    TablePlanDecision,
    TabularPlanningContext,
    TabularPlannerOutput,
    _planner_payload,
)
from app.biz.task_runtime.capabilities.descriptors import CapabilityDescriptor
from app.biz.task_runtime.domain.models import CapabilityDispatch, PreparedTaskBatch
from app.biz.task_runtime.sub_agent.profile import ALL_CAPABILITIES, ProfileDescriptor
from app.biz.task_runtime.workspace.rerun_sources import compact_rerun_source_payload, delegate_request_from_rerun_source
from app.biz.source import (
    GenericRowNormalizer,
    NormalizerSelector,
    SourceAccessContext,
    SourceError,
    TabularScope,
    TestCaseNormalizer,
    WorkspaceSourceService,
)
from app.biz.source.persistence.repository import WorkspaceSourceRepository
from app.tools.common import ToolContext

TabularParseError = SourceError


class _SourceSelectionHarness:
    """Test harness exercising the canonical workspace source boundary."""

    def parse(self, context: ToolContext, source_ref: str, **scope):
        tabular_scope = TabularScope(
            sheet_names=tuple(scope.pop("sheet_names", ())),
            row_start=scope.pop("row_start", None),
            row_end=scope.pop("row_end", None),
            case_ids=tuple(scope.pop("case_ids", ())),
        )
        return WorkspaceSourceService().select(
            SourceAccessContext(
                username=context.username,
                agent_instance_id=int(context.agent_instance_id or 0),
                conversation_id=int(context.conversation_id or 0),
            ),
            source_ref,
            scope=tabular_scope,
            **scope,
        )


@pytest.fixture
def context(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> ToolContext:
    workspace = tmp_path / "workspace"
    (workspace / "attachments").mkdir(parents=True)
    monkeypatch.setattr(
        "app.biz.source.service.CHAT_FS.get_workspace_path",
        lambda *args, **kwargs: workspace,
    )
    return ToolContext.model_construct(username="alice", agent_instance_id=7, project_id=13, conversation_id=11)


def test_parser_preserves_excel_cell_types(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Accounts"
    sheet.append(["Username", "Retries", "Enabled"])
    sheet.append(["alice", 3, True])
    workbook.save(path)

    document = _SourceSelectionHarness().parse(context, "attachments/accounts.xlsx")

    row = document.sheets[0].rows[0]
    assert row.values == {"Username": "alice", "Retries": 3, "Enabled": True}
    assert row.display_values == {"Username": "alice", "Retries": "3", "Enabled": "True"}


def test_selected_skill_description_respects_small_remaining_budget() -> None:
    assert _truncate_description("reporting instructions", 5) == "repor"


def test_parser_rejects_ambiguous_attachment_basename(context: ToolContext, tmp_path: Path) -> None:
    attachments = tmp_path / "workspace" / "attachments"
    (attachments / "first").mkdir()
    (attachments / "second").mkdir()
    (attachments / "first" / "rows.csv").write_text("Name\nAlice\n", encoding="utf-8")
    (attachments / "second" / "rows.csv").write_text("Name\nBob\n", encoding="utf-8")

    with pytest.raises(TabularParseError) as excinfo:
        _SourceSelectionHarness().parse(context, "rows.csv")

    assert excinfo.value.code == "tabular_source_ambiguous"
    assert excinfo.value.details["candidates"] == [
        "attachments/first/rows.csv",
        "attachments/second/rows.csv",
    ]


def test_parser_auto_selects_only_runnable_sheet(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "multi.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Cases"
    first.append(["Case ID", "Steps", "Expected Result"])
    first.append(["TC-1", "Open", "Home"])
    second = workbook.create_sheet("Summary")
    second.append(["Metric", "Value"])
    second.append(["Count", 1])
    workbook.save(path)

    document = _SourceSelectionHarness().parse(context, "attachments/multi.xlsx")

    assert [sheet.name for sheet in document.sheets] == ["Cases"]


def test_parser_rejects_blank_scope_for_multi_sheet_workbook(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "multi.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["Name"])
    first.append(["Alice"])
    second = workbook.create_sheet("Second")
    second.append(["Name"])
    second.append(["Bob"])
    workbook.save(path)

    with pytest.raises(TabularParseError) as excinfo:
        _SourceSelectionHarness().parse(context, "attachments/multi.xlsx", sheet_names=(" ",))

    assert excinfo.value.code == "tabular_sheet_scope_required"


def test_parser_reports_malformed_archive_row_metadata(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "rows.jsonl"
    path.write_text(
        json.dumps({"sheet_name": "Rows", "data_row_index": "one", "values": {"Name": "Alice"}}) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(TabularParseError) as excinfo:
        _SourceSelectionHarness().parse(context, "rows.jsonl")

    assert excinfo.value.code == "tabular_invalid_rows"
    assert excinfo.value.details["source_ref"] == "rows.jsonl"


def test_parser_preserves_missing_header_error(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "empty.csv"
    path.write_text("\n", encoding="utf-8")

    with pytest.raises(TabularParseError) as excinfo:
        _SourceSelectionHarness().parse(context, "attachments/empty.csv")

    assert excinfo.value.code == "tabular_missing_headers"


@pytest.mark.parametrize("suffix", ["xlsx", "jsonl"])
def test_parser_rejects_oversized_source(
    context: ToolContext,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    suffix: str,
) -> None:
    path = tmp_path / "workspace" / "attachments" / f"large.{suffix}"
    path.write_bytes(b"x" * 32)
    monkeypatch.setattr("app.biz.source.tabular.reader._TABULAR_MAX_BYTES", 10)

    with pytest.raises(TabularParseError) as excinfo:
        _SourceSelectionHarness().parse(context, f"attachments/large.{suffix}")

    assert excinfo.value.code == "tabular_source_too_large"


@pytest.mark.parametrize("suffix", ["xlsx", "jsonl"])
def test_parser_rejects_source_row_limit(
    context: ToolContext,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    suffix: str,
) -> None:
    path = tmp_path / "workspace" / "attachments" / f"rows.{suffix}"
    if suffix == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name"])
        sheet.append(["Alice"])
        workbook.save(path)
    else:
        path.write_text(
            "\n".join(json.dumps({"values": {"Name": name}}) for name in ("Alice", "Bob")),
            encoding="utf-8",
        )
    monkeypatch.setattr("app.biz.source.tabular.reader._TABULAR_MAX_SOURCE_ROWS", 1)

    with pytest.raises(TabularParseError) as excinfo:
        _SourceSelectionHarness().parse(context, f"attachments/rows.{suffix}")

    assert excinfo.value.code == "tabular_source_row_limit"


def test_parser_stops_when_selected_rows_exceed_budget(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_text("Name\nAlice\nBob\nCarol\n", encoding="utf-8")

    with pytest.raises(TabularParseError) as excinfo:
        _SourceSelectionHarness().parse(context, "attachments/rows.csv", max_rows=2)

    assert excinfo.value.code == "tabular_row_limit"
    assert excinfo.value.details == {"selected_rows": 3, "max_rows": 2}


def test_parser_applies_row_limit_across_workbook_sheets(
    context: ToolContext,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "multi.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["Name"])
    first.append(["Alice"])
    second = workbook.create_sheet("Second")
    second.append(["Name"])
    second.append(["Bob"])
    workbook.save(path)
    monkeypatch.setattr("app.biz.source.tabular.reader._TABULAR_MAX_SOURCE_ROWS", 3)

    with pytest.raises(TabularParseError) as excinfo:
        _SourceSelectionHarness().parse(context, "attachments/multi.xlsx", sheet_names=("First", "Second"))

    assert excinfo.value.code == "tabular_source_row_limit"


def test_explicit_sheet_scope_still_enforces_full_source_row_limit(
    context: ToolContext,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "multi.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["Name"])
    first.append(["Alice"])
    second = workbook.create_sheet("Second")
    second.append(["Name"])
    second.append(["Bob"])
    workbook.save(path)
    monkeypatch.setattr("app.biz.source.tabular.reader._TABULAR_MAX_SOURCE_ROWS", 2)

    with pytest.raises(TabularParseError) as excinfo:
        _SourceSelectionHarness().parse(context, "attachments/multi.xlsx", sheet_names=("First",))

    assert excinfo.value.code == "tabular_source_row_limit"


@pytest.mark.parametrize(
    ("suffix", "delimiter"),
    [("csv", ","), ("tsv", "\t")],
)
def test_parser_supports_delimited_formats(
    context: ToolContext,
    tmp_path: Path,
    suffix: str,
    delimiter: str,
) -> None:
    path = tmp_path / "workspace" / "attachments" / f"rows.{suffix}"
    path.write_text(f"Name{delimiter}Value\nalpha{delimiter}42\n", encoding="utf-8")

    document = _SourceSelectionHarness().parse(context, f"attachments/rows.{suffix}")

    assert document.format == suffix
    assert document.sheets[0].rows[0].values == {"Name": "alpha", "Value": "42"}


def test_parser_replaces_stray_undecodable_csv_bytes(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_bytes(b"Name\ncaf\xff\n")

    document = _SourceSelectionHarness().parse(context, "attachments/rows.csv")

    assert document.sheets[0].rows[0].values["Name"] == b"caf\xff".decode("utf-8", errors="replace")


def test_parser_skips_whitespace_only_data_rows(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_text("Name,Enabled\nAlice,true\n   ,   \nBob,false\n", encoding="utf-8")

    document = _SourceSelectionHarness().parse(context, "attachments/rows.csv")

    assert [row.values["Name"] for row in document.sheets[0].rows] == ["Alice", "Bob"]
    assert [row.source_row for row in document.sheets[0].rows] == [2, 4]


def test_parser_reads_legacy_case_jsonl_as_tabular_rows(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "case_sources.jsonl"
    path.write_text(
        json.dumps(
            {
                "sheet_name": "Cases",
                "data_row_index": 1,
                "source_row": 2,
                "values": {"Case ID": "TC-1", "Steps": "Open app", "Expected Result": "Home"},
            }
        )
        + "\n",
        encoding="utf-8",
    )

    document = _SourceSelectionHarness().parse(context, "case_sources.jsonl")

    assert document.sheets[0].name == "Cases"
    assert document.sheets[0].rows[0].source_row == 2
    assert document.sheets[0].rows[0].values["Case ID"] == "TC-1"


def test_normalizer_selects_testcase_only_with_strong_header_evidence(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    cases = tmp_path / "workspace" / "attachments" / "cases.csv"
    cases.write_text("Case ID,Steps,Expected Result\nTC-1,Open app,Home\n", encoding="utf-8")
    sales = tmp_path / "workspace" / "attachments" / "sales.csv"
    sales.write_text("Customer,Amount,Region\nAlice,1200,US\n", encoding="utf-8")
    parser = _SourceSelectionHarness()
    selector = NormalizerSelector()

    case_sheet = parser.parse(context, "attachments/cases.csv").sheets[0]
    sales_sheet = parser.parse(context, "attachments/sales.csv").sheets[0]

    assert isinstance(selector.select(case_sheet), TestCaseNormalizer)
    assert isinstance(selector.select(sales_sheet), GenericRowNormalizer)


def test_binder_maps_headers_and_builtin_sources_once_for_every_row(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("User Name,Retries\nalice,3\nbob,4\n", encoding="utf-8")
    document = _SourceSelectionHarness().parse(context, "attachments/accounts.csv")
    sheet = document.sheets[0]
    rows = GenericRowNormalizer().normalize(document, sheet)
    descriptor = CapabilityDescriptor(
        capability_id="skill:accounts.import",
        parameter_schema={
            "type": "object",
            "properties": {
                "username": {
                    "type": "string",
                    "x-sico-binding": {"aliases": ["User Name"]},
                },
                "retries": {"type": "integer"},
                "input_file": {"type": "string"},
            },
            "required": ["username", "retries", "input_file"],
        },
        required_sandbox=(),
        workspace_access="read_only",
        effect="mutate",
    )
    binder = ArgumentBinder()

    plan = binder.infer_plan(descriptor, sheet)
    assert not isinstance(plan, NeedsClarification)
    bound = binder.bind(descriptor, rows, plan)

    assert isinstance(bound, tuple)
    input_file = str(bound[0].arguments["input_file"])
    assert [item.arguments for item in bound] == [
        {
            "username": "alice",
            "retries": 3,
            "input_file": input_file,
        },
        {
            "username": "bob",
            "retries": 4,
            "input_file": input_file,
        },
    ]
    assert bound[0].evidence == {
        "username": "column:User Name",
        "retries": "column:Retries",
        "input_file": "document_path",
    }
    assert input_file.startswith("sico-source://objects/")


def test_binder_requires_clarification_for_unmapped_required_parameter(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_text("Name\nAlice\n", encoding="utf-8")
    sheet = _SourceSelectionHarness().parse(context, "attachments/rows.csv").sheets[0]
    descriptor = CapabilityDescriptor(
        capability_id="skill:accounts.import",
        parameter_schema={
            "type": "object",
            "properties": {"tenant_id": {"type": "string"}},
            "required": ["tenant_id"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )

    outcome = ArgumentBinder().infer_plan(descriptor, sheet)

    assert isinstance(outcome, NeedsClarification)
    assert outcome.code == "tabular_binding_required"
    assert outcome.details["missing_parameters"] == ["tenant_id"]


def test_explicit_binding_overrides_inference(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_text("Login\nalice\n", encoding="utf-8")
    document = _SourceSelectionHarness().parse(context, "attachments/rows.csv")
    sheet = document.sheets[0]
    descriptor = CapabilityDescriptor(
        capability_id="skill:accounts.import",
        parameter_schema={
            "type": "object",
            "properties": {"username": {"type": "string"}},
            "required": ["username"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    binder = ArgumentBinder()

    plan = binder.infer_plan(
        descriptor,
        sheet,
        {"username": BindingRule(source="column", column="Login")},
    )
    rows = GenericRowNormalizer().normalize(document, sheet)
    bound = binder.bind(descriptor, rows, plan)

    assert isinstance(bound, tuple)
    assert bound[0].arguments == {"username": "alice"}


def test_explicit_binding_rejects_unknown_optional_column(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_text("Name\nAlice\n", encoding="utf-8")
    sheet = _SourceSelectionHarness().parse(context, "attachments/rows.csv").sheets[0]
    descriptor = CapabilityDescriptor(
        capability_id="skill:accounts.import",
        parameter_schema={"type": "object", "properties": {"nickname": {"type": "string"}}},
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )

    outcome = ArgumentBinder().infer_plan(
        descriptor,
        sheet,
        {"nickname": BindingRule(source="column", column="Typo")},
    )

    assert isinstance(outcome, Rejected)
    assert outcome.code == "tabular_invalid_bindings"
    assert outcome.details["unknown_columns"] == ["Typo"]


def test_binder_maps_task_name_to_normalized_title(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "cases.csv"
    path.write_text("Case ID,Title,Steps,Expected Result\nTC-1,Login,Open,Home\n", encoding="utf-8")
    document = _SourceSelectionHarness().parse(context, "attachments/cases.csv")
    sheet = document.sheets[0]
    rows = TestCaseNormalizer().normalize(document, sheet)
    descriptor = CapabilityDescriptor(
        capability_id="skill:tests.run",
        parameter_schema={
            "type": "object",
            "properties": {"task_name": {"type": "string"}, "instructions": {"type": "string"}},
            "required": ["task_name", "instructions"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    binder = ArgumentBinder()

    plan = binder.infer_plan(descriptor, sheet)
    bound = binder.bind(descriptor, rows, plan)

    assert isinstance(bound, tuple)
    assert bound[0].arguments["task_name"] == "Login"
    assert bound[0].arguments["instructions"].startswith("Tabular source:")


def test_caller_literal_password_passes_through_binding(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("Username\nalice\n", encoding="utf-8")
    document = _SourceSelectionHarness().parse(context, "attachments/accounts.csv")
    sheet = document.sheets[0]
    rows = GenericRowNormalizer().normalize(document, sheet)
    descriptor = CapabilityDescriptor(
        capability_id="skill:accounts.login",
        parameter_schema={
            "type": "object",
            "properties": {"username": {"type": "string"}, "password": {"type": "string"}},
            "required": ["username", "password"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    binder = ArgumentBinder()

    plan = binder.infer_plan(
        descriptor,
        sheet,
        {"password": BindingRule(source="literal", value="plain-text-fixture")},
    )
    bound = binder.bind(descriptor, rows, plan)

    assert isinstance(bound, tuple)
    assert bound[0].arguments == {"username": "alice", "password": "plain-text-fixture"}


def test_tabular_planner_schema_rejects_model_generated_literal() -> None:
    with pytest.raises(ValidationError):
        BindingDecision(parameter="password", source="literal")


@pytest.mark.asyncio
async def test_tabular_planner_rejects_invented_binding_parameter(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_text("Login\nalice\n", encoding="utf-8")
    document = _SourceSelectionHarness().parse(context, "attachments/rows.csv")
    sheet = document.sheets[0]
    descriptor = CapabilityDescriptor(
        capability_id="skill:accounts.import",
        parameter_schema={
            "type": "object",
            "properties": {"username": {"type": "string"}},
            "required": ["username"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    rows = GenericRowNormalizer().normalize(document, sheet)

    async def invalid_plan(batch_goal, contexts):
        return TabularPlannerOutput(
            tables=[
                TablePlanDecision(
                    table_id=sheet.table_id,
                    capability_id=descriptor.capability_id,
                    bindings=[BindingDecision(parameter="invented", source="column", column="Login")],
                )
            ]
        )

    with pytest.raises(PreparationError) as excinfo:
        await LlmTabularPlanner(invalid_plan).plan(
            "Import accounts",
            (TabularPlanningContext(document, sheet, rows, (descriptor,)),),
        )

    assert excinfo.value.code == "tabular_planner_invalid_output"


@pytest.mark.asyncio
async def test_tabular_planner_can_request_capability_clarification(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "cases.csv"
    path.write_text("Steps\nOpen the browser and search for weather\n", encoding="utf-8")
    document = _SourceSelectionHarness().parse(context, "attachments/cases.csv")
    sheet = document.sheets[0]
    rows = GenericRowNormalizer().normalize(document, sheet)
    descriptors = (
        _account_descriptor("skill:desktop-tester.run_desktop_tester"),
        _account_descriptor("skill:android-tester.run_android_test_case"),
    )

    async def clarify(*args):
        return TabularPlannerOutput(
            tables=[
                TablePlanDecision(
                    table_id=sheet.table_id,
                    outcome="needs_clarification",
                    clarification="Should these cases run on desktop or Android?",
                )
            ]
        )

    outcome = await LlmTabularPlanner(clarify).plan(
        "Run browser tests",
        (TabularPlanningContext(document, sheet, rows, descriptors),),
    )

    assert isinstance(outcome, NeedsClarification)
    assert outcome.code == "tabular_capability_clarification"
    assert outcome.missing == ("target execution capability",)


def test_tabular_planner_payload_deduplicates_capability_schemas(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_text("Login\nalice\n", encoding="utf-8")
    document = _SourceSelectionHarness().parse(context, "attachments/rows.csv")
    sheet = document.sheets[0]
    rows = GenericRowNormalizer().normalize(document, sheet)
    descriptor = CapabilityDescriptor(
        capability_id="skill:accounts.import",
        parameter_schema={"type": "object", "properties": {"username": {"type": "string"}}},
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    contexts = (
        TabularPlanningContext(document, sheet, rows, (descriptor,), context_id="table-1"),
        TabularPlanningContext(document, sheet, rows, (descriptor,), context_id="table-2"),
    )

    payload = _planner_payload("Import", contexts)

    assert [item["capability_id"] for item in payload["capabilities"]] == [descriptor.capability_id]
    assert [item["candidate_capability_ids"] for item in payload["tables"]] == [
        [descriptor.capability_id],
        [descriptor.capability_id],
    ]


@pytest.mark.asyncio
async def test_tabular_planner_requests_explicit_scope_above_table_limit(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_text("Login\nalice\n", encoding="utf-8")
    document = _SourceSelectionHarness().parse(context, "attachments/rows.csv")
    sheet = document.sheets[0]
    rows = GenericRowNormalizer().normalize(document, sheet)
    descriptor = CapabilityDescriptor(
        capability_id="skill:accounts.import",
        parameter_schema={"type": "object", "properties": {"username": {"type": "string"}}},
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    contexts = tuple(
        TabularPlanningContext(document, sheet, rows, (descriptor,), context_id=f"table-{index}")
        for index in range(51)
    )

    outcome = await LlmTabularPlanner(lambda *args: pytest.fail("scope limit must run before planner")).plan(
        "Import",
        contexts,
    )

    assert isinstance(outcome, NeedsClarification)
    assert outcome.code == "tabular_planner_scope_limit"


class _Catalogue(CapabilityCatalogue):
    def __init__(self, descriptors: tuple[CapabilityDescriptor, ...]) -> None:
        self.descriptors = descriptors

    async def list_descriptors(self, context, query):
        return tuple(descriptor for descriptor in self.descriptors if query.matches(descriptor))


class _Profiles:
    def list_profiles(self, query):
        return ()

    def resolve(self, profile_id):
        return None


def _echo_descriptor() -> CapabilityDescriptor:
    return CapabilityDescriptor(
        capability_id="builtin:echo",
        parameter_schema={
            "type": "object",
            "properties": {"message": {"type": "string"}},
            "required": ["message"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="read",
    )


def _account_descriptor(capability_id: str = "skill:accounts.import") -> CapabilityDescriptor:
    return CapabilityDescriptor(
        capability_id=capability_id,
        parameter_schema={
            "type": "object",
            "properties": {
                "username": {"type": "string", "x-sico-binding": {"aliases": ["User Name"]}},
                "input_file": {"type": "string"},
            },
            "required": ["username", "input_file"],
        },
        required_sandbox=(),
        workspace_access="read_only",
        effect="mutate",
    )


@pytest.mark.asyncio
async def test_service_prepares_mixed_sources_and_multiple_tabular_files(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    first = tmp_path / "workspace" / "attachments" / "first.csv"
    first.write_text("User Name\nalice\n", encoding="utf-8")
    second = tmp_path / "workspace" / "attachments" / "second.tsv"
    second.write_text("User Name\nbob\n", encoding="utf-8")

    async def unexpected_task_planner(*args):  # pragma: no cover - every item is prebound
        raise AssertionError("task planner must not run for fully prebound mixed sources")

    service = DelegatePreparationService(
        LlmTaskPlanner(unexpected_task_planner),
        _Profiles(),
        _Catalogue((_echo_descriptor(), _account_descriptor())),
    )
    request = {
        "batch_goal": "Echo a marker and import both account tables",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "documents": [
                    {"source_ref": "attachments/first.csv"},
                    {"source_ref": "attachments/second.tsv"},
                ],
            },
            {
                "type": "instructions",
                "capability_ids": ["builtin:echo"],
                "items": [
                    {
                        "goal": "Echo the marker",
                        "capability_id": "builtin:echo",
                        "params": {"message": "start"},
                    }
                ],
            },
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert len(outcome.batch.tasks) == 3
    first_args, second_args, instruction_args = [task.args for task in outcome.batch.tasks]
    assert first_args["username"] == "alice"
    assert second_args["username"] == "bob"
    assert instruction_args == {"message": "start"}
    repository = WorkspaceSourceRepository(tmp_path / "workspace")
    first_input = repository.resolve_object_ref(str(first_args["input_file"]))
    second_input = repository.resolve_object_ref(str(second_args["input_file"]))
    assert first_input is not None
    assert second_input is not None
    assert first_input.read_text(encoding="utf-8") == "User Name\nalice\n"
    assert second_input.read_text(encoding="utf-8") == "User Name\nbob\n"
    assert outcome.batch.tasks[0].metadata["tabular"]["source_object_ref"] == first_args["input_file"]
    assert outcome.batch.tasks[1].metadata["tabular"]["source_object_ref"] == second_args["input_file"]
    assert all(isinstance(task.dispatch, CapabilityDispatch) for task in outcome.batch.tasks)
    assert outcome.batch_metadata == {"source_count": 2}


@pytest.mark.asyncio
async def test_service_preserves_caller_literal_password_in_task_args(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("Username\nalice\n", encoding="utf-8")
    descriptor = CapabilityDescriptor(
        capability_id="skill:accounts.login",
        parameter_schema={
            "type": "object",
            "properties": {
                "username": {"type": "string"},
                "password": {"type": "string", "sensitive": True},
            },
            "required": ["username", "password"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("fully bound table must not call the task planner")),
        _Profiles(),
        _Catalogue((descriptor,)),
    )
    request = {
        "batch_goal": "Log in",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.login"],
                "documents": [{"source_ref": "attachments/accounts.csv"}],
                "parameter_bindings": {
                    "password": {"source": "literal", "value": "plain-text-fixture"},
                },
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert outcome.batch.tasks[0].args == {"username": "alice", "password": "plain-text-fixture"}


@pytest.mark.asyncio
async def test_tabular_batch_preserves_selected_skill_reporting_context(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("User Name\nalice\n", encoding="utf-8")
    skill_dir = tmp_path / "account-skill"
    skill_dir.mkdir()
    skill_description = "# Account import\n\nReporting: include the imported account count."
    (skill_dir / "SKILL.md").write_text(skill_description, encoding="utf-8")

    class _Card:
        def __init__(self) -> None:
            self.skill_dir = str(skill_dir)

    class _Loader:
        def resolve(self, name):
            assert name == "accounts.import"
            return _Card()

    context.skill_loader = _Loader()
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("tabular row is prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
    )
    request = {
        "batch_goal": "Import accounts",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "documents": [{"source_ref": "attachments/accounts.csv"}],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert outcome.adapter_state["skill_description"] == skill_description


@pytest.mark.asyncio
async def test_builtin_only_request_does_not_load_skills_or_profiles(context: ToolContext) -> None:
    class _FailingLoader:
        def list_cards(self, *, visibility):
            pytest.fail("builtin-only request must not load the skill catalogue")

    class _FailingProfiles:
        def list_profiles(self, query):
            pytest.fail("fully prebound request must not enumerate profiles")

        def resolve(self, profile_id):
            return None

    context.skill_loader = _FailingLoader()
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("prebound request must not call the planner")),
        _FailingProfiles(),
        WorkspaceCapabilityCatalogue(),
    )
    request = {
        "batch_goal": "Echo",
        "sources": [
            {
                "type": "instructions",
                "allow_sub_agent": False,
                "capability_ids": ["builtin:echo"],
                "items": [
                    {
                        "goal": "Echo",
                        "capability_id": "echo",
                        "params": {"message": "hello"},
                    }
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert outcome.batch.tasks[0].args == {"message": "hello"}


@pytest.mark.asyncio
async def test_item_level_builtin_capability_drives_discovery(context: ToolContext) -> None:
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("prebound builtin must not call the planner")),
        _Profiles(),
        WorkspaceCapabilityCatalogue(),
    )
    request = {
        "batch_goal": "Echo",
        "sources": [
            {
                "type": "instructions",
                "allow_sub_agent": False,
                "items": [
                    {
                        "goal": "Echo",
                        "capability_id": "echo",
                        "params": {"message": "hello"},
                    }
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert outcome.batch.tasks[0].capability_id == "builtin:echo"


@pytest.mark.asyncio
async def test_unused_profile_allow_list_is_validated_for_capability_bound_source(context: ToolContext) -> None:
    profile = ProfileDescriptor("default", "General tasks", ALL_CAPABILITIES)

    class _ProfilesWithDefault:
        def list_profiles(self, query):
            return (profile,)

        def resolve(self, profile_id):
            return None

    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("prebound builtin must not call the planner")),
        _ProfilesWithDefault(),
        WorkspaceCapabilityCatalogue(),
    )
    request = {
        "batch_goal": "Echo",
        "sources": [
            {
                "type": "instructions",
                "capability_ids": ["builtin:echo"],
                "profile_ids": ["default"],
                "items": [
                    {
                        "goal": "Echo",
                        "capability_id": "builtin:echo",
                        "params": {"message": "hello"},
                    }
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)


@pytest.mark.asyncio
async def test_service_uses_one_tabular_planner_call_for_ambiguous_capabilities(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("Login\nalice\n", encoding="utf-8")
    calls = []

    async def choose_capability(batch_goal, contexts):
        calls.append(tuple(context.sheet.table_id for context in contexts))
        assert [descriptor.capability_id for descriptor in contexts[0].descriptors] == [
            "skill:accounts.primary",
            "skill:accounts.backup",
        ]
        return TabularPlannerOutput(
            tables=[
                TablePlanDecision(
                    table_id=contexts[0].planning_id,
                    capability_id="skill:accounts.primary",
                )
            ]
        )

    incompatible = CapabilityDescriptor(
        capability_id="skill:accounts.incompatible",
        parameter_schema={
            "type": "object",
            "properties": {"account_name": {"type": "string"}},
            "required": ["account_name"],
        },
        required_sandbox=(),
        workspace_access="read_only",
        effect="mutate",
    )
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("all tabular rows are prebound")),
        _Profiles(),
        _Catalogue(
            (
                _account_descriptor("skill:accounts.primary"),
                _account_descriptor("skill:accounts.backup"),
                incompatible,
            )
        ),
        tabular_planner=LlmTabularPlanner(choose_capability),
    )
    request = {
        "batch_goal": "Import accounts",
        "sources": [
            {
                "type": "tabular",
                "documents": [{"source_ref": "attachments/accounts.csv"}],
                "parameter_bindings": {"username": {"source": "column", "column": "Login"}},
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert len(calls) == 1
    assert outcome.batch.tasks[0].capability_id == "skill:accounts.primary"
    assert outcome.batch.tasks[0].args["username"] == "alice"


@pytest.mark.asyncio
async def test_service_merges_overlapping_scopes_before_table_planning(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("Login\nalice\nbob\ncarol\n", encoding="utf-8")
    calls = []

    async def choose_capability(batch_goal, contexts):
        calls.append(tuple(context.sheet.table_id for context in contexts))
        return TabularPlannerOutput(
            tables=[
                TablePlanDecision(
                    table_id=contexts[0].planning_id,
                    capability_id="skill:accounts.primary",
                )
            ]
        )

    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("all tabular rows are prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor("skill:accounts.primary"), _account_descriptor("skill:accounts.backup"))),
        tabular_planner=LlmTabularPlanner(choose_capability),
    )
    request = {
        "batch_goal": "Import overlapping account scopes once",
        "sources": [
            {
                "type": "tabular",
                "documents": [
                    {"source_ref": "attachments/accounts.csv", "row_start": 1, "row_end": 2},
                    {"source_ref": "attachments/accounts.csv", "row_start": 2, "row_end": 3},
                ],
                "parameter_bindings": {"username": {"source": "column", "column": "Login"}},
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert len(calls) == 1
    assert len(calls[0]) == 1
    assert [task.args["username"] for task in outcome.batch.tasks] == ["alice", "bob", "carol"]


@pytest.mark.asyncio
async def test_overlapping_scopes_only_budget_remaining_plus_duplicate_rows(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("User Name\nalice\nbob\ncarol\n", encoding="utf-8")

    class _RecordingSources(WorkspaceSourceService):
        def __init__(self) -> None:
            super().__init__()
            self.max_rows: list[int | None] = []

        def select(self, *args, **kwargs):
            self.max_rows.append(kwargs.get("max_rows"))
            return super().select(*args, **kwargs)

    sources = _RecordingSources()
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("all tabular rows are prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
        source_service=sources,
    )
    request = {
        "batch_goal": "Import overlapping account scopes once",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "max_rows": 3,
                "documents": [
                    {"source_ref": "attachments/accounts.csv", "row_start": 1, "row_end": 2},
                    {"source_ref": "attachments/accounts.csv", "row_start": 2, "row_end": 3},
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert sources.max_rows == [3, 2]


@pytest.mark.asyncio
async def test_bare_attachment_ref_does_not_reuse_same_named_knowledge_rows_for_budget(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    workspace = tmp_path / "workspace"
    knowledge = workspace / "knowledge" / "1" / "accounts.csv"
    attachment = workspace / "attachments" / "accounts.csv"
    knowledge.parent.mkdir(parents=True)
    attachment.parent.mkdir(parents=True, exist_ok=True)
    knowledge.write_text("User Name\nknowledge-one\nknowledge-two\n", encoding="utf-8")
    attachment.write_text("User Name\nattachment-one\nattachment-two\n", encoding="utf-8")

    class _RecordingSources(WorkspaceSourceService):
        def __init__(self) -> None:
            super().__init__()
            self.max_rows: list[int | None] = []

        def select(self, *args, **kwargs):
            self.max_rows.append(kwargs.get("max_rows"))
            return super().select(*args, **kwargs)

    sources = _RecordingSources()
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("over-limit source must not plan")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
        source_service=sources,
    )
    request = {
        "batch_goal": "Keep same-named sources separate",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "max_rows": 2,
                "documents": [
                    {"source_ref": "knowledge/1/accounts.csv", "row_start": 1, "row_end": 1},
                    {"source_ref": "accounts.csv"},
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, NeedsClarification)
    assert outcome.code == "tabular_row_limit"
    assert sources.max_rows == [2, 1]


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("second_scope", "expected_limits", "expected_code"),
    [
        ({"sheet_names": ["table"], "row_start": 1, "row_end": 2}, [2, 2], None),
        ({"row_start": 2, "row_end": 3}, [2, 1], "tabular_row_limit"),
    ],
)
async def test_full_source_budget_allows_duplicates_but_rejects_new_rows(
    context: ToolContext,
    tmp_path: Path,
    second_scope: dict[str, object],
    expected_limits: list[int],
    expected_code: str | None,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("User Name\nalice\nbob\ncarol\n", encoding="utf-8")

    class _RecordingSources(WorkspaceSourceService):
        def __init__(self) -> None:
            super().__init__()
            self.max_rows: list[int | None] = []

        def select(self, *args, **kwargs):
            self.max_rows.append(kwargs.get("max_rows"))
            return super().select(*args, **kwargs)

    sources = _RecordingSources()
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("all selected rows are prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
        source_service=sources,
    )
    request = {
        "batch_goal": "Import bounded overlapping account scopes",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "max_rows": 2,
                "documents": [
                    {"source_ref": "attachments/accounts.csv", "row_start": 1, "row_end": 2},
                    {"source_ref": "attachments/accounts.csv", **second_scope},
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert sources.max_rows == expected_limits
    if expected_code is None:
        assert isinstance(outcome, PreparedTaskBatch)
        assert [task.args["username"] for task in outcome.batch.tasks] == ["alice", "bob"]
    else:
        assert isinstance(outcome, NeedsClarification)
        assert outcome.code == expected_code


@pytest.mark.asyncio
async def test_service_merges_row_range_and_exact_case_scopes(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("Case ID,User Name\nTC-1,alice\nTC-2,bob\nTC-3,carol\n", encoding="utf-8")
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("all tabular rows are prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
    )
    request = {
        "batch_goal": "Import mixed scopes once",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "documents": [
                    {"source_ref": "attachments/accounts.csv", "row_start": 1, "row_end": 2},
                    {"source_ref": "attachments/accounts.csv", "case_ids": ["TC-3"]},
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert [task.args["username"] for task in outcome.batch.tasks] == ["alice", "bob", "carol"]
    assert [task.metadata["source"]["document_index"] for task in outcome.batch.tasks] == [1, 1, 2]


@pytest.mark.asyncio
async def test_bindable_fallback_does_not_bypass_semantic_capability_planning(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("Login\nalice\n", encoding="utf-8")
    fallback = CapabilityDescriptor(
        capability_id="skill:generic.noop",
        parameter_schema={"type": "object", "properties": {}},
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    correct = CapabilityDescriptor(
        capability_id="skill:accounts.import",
        parameter_schema={
            "type": "object",
            "properties": {"username": {"type": "string"}},
            "required": ["username"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    calls = []

    async def choose_capability(batch_goal, contexts):
        calls.append(contexts)
        assert [descriptor.capability_id for descriptor in contexts[0].descriptors] == [
            "skill:generic.noop",
            "skill:accounts.import",
        ]
        return TabularPlannerOutput(
            tables=[
                TablePlanDecision(
                    table_id=contexts[0].planning_id,
                    capability_id="skill:accounts.import",
                    bindings=[BindingDecision(parameter="username", source="column", column="Login")],
                )
            ]
        )

    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("tabular rows are prebound")),
        _Profiles(),
        _Catalogue((fallback, correct)),
        tabular_planner=LlmTabularPlanner(choose_capability),
    )
    request = {
        "batch_goal": "Import accounts",
        "sources": [{"type": "tabular", "documents": [{"source_ref": "attachments/accounts.csv"}]}],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert len(calls) == 1
    assert outcome.batch.tasks[0].capability_id == "skill:accounts.import"
    assert outcome.batch.tasks[0].args == {"username": "alice"}


@pytest.mark.asyncio
async def test_service_plans_all_ambiguous_tabular_sources_in_one_call(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    first = tmp_path / "workspace" / "attachments" / "first.csv"
    second = tmp_path / "workspace" / "attachments" / "second.csv"
    first.write_text("Login\nalice\n", encoding="utf-8")
    second.write_text("Login\nbob\n", encoding="utf-8")
    calls = []

    async def choose_capabilities(batch_goal, contexts):
        calls.append(tuple(context.planning_id for context in contexts))
        return TabularPlannerOutput(
            tables=[
                TablePlanDecision(
                    table_id=context.planning_id,
                    capability_id="skill:accounts.primary",
                )
                for context in contexts
            ]
        )

    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("all tabular rows are prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor("skill:accounts.primary"), _account_descriptor("skill:accounts.backup"))),
        tabular_planner=LlmTabularPlanner(choose_capabilities),
    )
    request = {
        "batch_goal": "Import both sources",
        "sources": [
            {
                "type": "tabular",
                "documents": [{"source_ref": "attachments/first.csv"}],
                "parameter_bindings": {"username": {"source": "column", "column": "Login"}},
            },
            {
                "type": "tabular",
                "documents": [{"source_ref": "attachments/second.csv"}],
                "parameter_bindings": {"username": {"source": "column", "column": "Login"}},
            },
            {
                "type": "tabular",
                "documents": [{"source_ref": "attachments/first.csv"}],
                "parameter_bindings": {"username": {"source": "column", "column": "Login"}},
            },
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert len(calls) == 1
    assert len(calls[0]) == 3
    assert len(set(calls[0])) == 3
    assert [task.args["username"] for task in outcome.batch.tasks] == ["alice", "bob", "alice"]


@pytest.mark.asyncio
async def test_service_gives_identical_files_unique_task_ids(context: ToolContext, tmp_path: Path) -> None:
    content = "User Name\nalice\n"
    first = tmp_path / "workspace" / "attachments" / "first.csv"
    second = tmp_path / "workspace" / "attachments" / "second.csv"
    first.write_text(content, encoding="utf-8")
    second.write_text(content, encoding="utf-8")
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("tabular rows are prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
    )
    request = {
        "batch_goal": "Import both copies",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "documents": [
                    {"source_ref": "attachments/first.csv"},
                    {"source_ref": "attachments/second.csv"},
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert len(outcome.batch.tasks) == 2
    assert len({task.task_id for task in outcome.batch.tasks}) == 2


@pytest.mark.asyncio
async def test_service_keeps_same_named_identical_files_as_distinct_inputs(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    content = "User Name\nalice\n"
    first = tmp_path / "workspace" / "attachments" / "first" / "accounts.csv"
    second = tmp_path / "workspace" / "attachments" / "second" / "accounts.csv"
    first.parent.mkdir()
    second.parent.mkdir()
    first.write_text(content, encoding="utf-8")
    second.write_text(content, encoding="utf-8")
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("tabular rows are prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
    )
    request = {
        "batch_goal": "Import both copies",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "documents": [
                    {"source_ref": "attachments/first/accounts.csv"},
                    {"source_ref": "attachments/second/accounts.csv"},
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert len(outcome.batch.tasks) == 2
    assert len({task.task_id for task in outcome.batch.tasks}) == 2


@pytest.mark.asyncio
async def test_service_prepares_from_snapshot_after_original_attachment_is_removed(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("User Name\nalice\n", encoding="utf-8")
    WorkspaceSourceService().index_path(
        tmp_path / "workspace",
        "attachments/accounts.csv",
        path,
    )
    path.unlink()
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("tabular rows are prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
    )
    request = {
        "batch_goal": "Import archived accounts",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "documents": [{"source_ref": "attachments/accounts.csv"}],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert outcome.batch.tasks[0].args["username"] == "alice"
    input_file = str(outcome.batch.tasks[0].args["input_file"])
    assert input_file.startswith("sico-source://objects/")
    assert outcome.batch.tasks[0].metadata["tabular"]["source_object_ref"] == input_file
    object_path = WorkspaceSourceRepository(tmp_path / "workspace").resolve_object_ref(input_file)
    assert object_path is not None
    assert object_path.read_bytes() == b"User Name\nalice\n"


@pytest.mark.asyncio
@pytest.mark.parametrize("replace_source", [False, True])
async def test_compact_rerun_rematerializes_matching_snapshot_only_source(
    context: ToolContext,
    tmp_path: Path,
    replace_source: bool,
) -> None:
    workspace = tmp_path / "workspace"
    source_ref = "attachments/accounts.csv"
    path = workspace / source_ref
    path.write_text("User Name\nalice\n", encoding="utf-8")
    source_service = WorkspaceSourceService()
    original = source_service.index_path(workspace, source_ref, path)
    repository = WorkspaceSourceRepository(workspace)
    object_ref = repository.object_ref(original)
    rerun_source = compact_rerun_source_payload(
        {
            "reason": "Repeat account import",
            "tasks": [
                {
                    "task_id": "account-1",
                    "title": "Import alice",
                    "instructions": "Import alice",
                    "dispatch": {"type": "capability", "capability_id": "skill:accounts.import"},
                    "args": {"username": "alice", "input_file": object_ref},
                    "metadata": {
                        "tabular": {
                            "source_ref": source_ref,
                            "source_object_ref": object_ref,
                        }
                    },
                }
            ],
        }
    )
    path.unlink()
    if replace_source:
        path.write_text("User Name\nbob\n", encoding="utf-8")
        source_service.index_path(workspace, source_ref, path)
    request = delegate_request_from_rerun_source(rerun_source)
    assert request is not None
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("compact rerun is prebound")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
        source_service=source_service,
    )

    outcome = await service.prepare(context, json.dumps(request))

    if replace_source:
        assert isinstance(outcome, NeedsClarification)
        assert outcome.code == "source_content_changed"
        return
    assert isinstance(outcome, PreparedTaskBatch)
    task = outcome.batch.tasks[0]
    assert task.args == {"username": "alice", "input_file": object_ref}
    assert task.metadata["tabular"] == {
        "source_ref": source_ref,
        "source_object_ref": object_ref,
    }


@pytest.mark.asyncio
async def test_service_requires_every_requested_case_id(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "cases.csv"
    path.write_text("Case ID,Steps,Expected Result\nTC-1,Open,Home\n", encoding="utf-8")
    descriptor = CapabilityDescriptor(
        capability_id="skill:tests.run",
        parameter_schema={
            "type": "object",
            "properties": {"case_id": {"type": "string"}},
            "required": ["case_id"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("missing case must not plan")),
        _Profiles(),
        _Catalogue((descriptor,)),
    )
    request = {
        "batch_goal": "Run selected cases",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:tests.run"],
                "documents": [{"source_ref": "attachments/cases.csv", "case_ids": ["TC-1", "TC-404"]}],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, NeedsClarification)
    assert outcome.code == "tabular_case_ids_not_found"
    assert outcome.details["missing_case_ids"] == ["TC-404"]


@pytest.mark.asyncio
async def test_service_finds_exact_case_id_beyond_row_budget_prefix(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "cases.csv"
    path.write_text(
        "Case ID\n" + "\n".join(f"TC-{index}" for index in range(1, 502)) + "\n",
        encoding="utf-8",
    )
    descriptor = CapabilityDescriptor(
        capability_id="skill:tests.run",
        parameter_schema={
            "type": "object",
            "properties": {"case_id": {"type": "string"}},
            "required": ["case_id"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("exact tabular row is prebound")),
        _Profiles(),
        _Catalogue((descriptor,)),
    )
    request = {
        "batch_goal": "Run one exact case",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:tests.run"],
                "documents": [{"source_ref": "attachments/cases.csv", "case_ids": ["TC-501"]}],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert [task.args["case_id"] for task in outcome.batch.tasks] == ["TC-501"]


@pytest.mark.asyncio
async def test_service_preserves_requested_case_id_order(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "cases.csv"
    path.write_text("Case ID\nTC-1\nTC-2\nTC-3\n", encoding="utf-8")
    descriptor = CapabilityDescriptor(
        capability_id="skill:tests.run",
        parameter_schema={
            "type": "object",
            "properties": {"case_id": {"type": "string"}},
            "required": ["case_id"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("ordered tabular rows are prebound")),
        _Profiles(),
        _Catalogue((descriptor,)),
    )
    request = {
        "batch_goal": "Run ordered cases",
        "max_concurrency": 1,
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:tests.run"],
                "documents": [
                    {"source_ref": "attachments/cases.csv", "case_ids": ["ＴＣ-３", "tc-3", "TC-1"]},
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert [task.args["case_id"] for task in outcome.batch.tasks] == ["TC-3", "TC-1"]


@pytest.mark.asyncio
async def test_generic_rows_preserve_case_id_for_filtering(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "owners.csv"
    path.write_text("Case ID,Owner\nTC-1,Alice\nTC-2,Bob\n", encoding="utf-8")
    descriptor = CapabilityDescriptor(
        capability_id="skill:cases.assign",
        parameter_schema={
            "type": "object",
            "properties": {"case_id": {"type": "string"}},
            "required": ["case_id"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("generic row is prebound")),
        _Profiles(),
        _Catalogue((descriptor,)),
    )
    request = {
        "batch_goal": "Assign selected case",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:cases.assign"],
                "documents": [{"source_ref": "attachments/owners.csv", "case_ids": ["TC-2"]}],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert len(outcome.batch.tasks) == 1
    assert outcome.batch.tasks[0].args == {"case_id": "TC-2"}


@pytest.mark.asyncio
async def test_tabular_source_without_capabilities_rejects_without_planner(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "owners.csv"
    path.write_text("Case ID\nTC-1\n", encoding="utf-8")

    async def unexpected_tabular_planner(*args):
        pytest.fail("an empty capability set must not call the tabular planner")

    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("an empty capability set must not call the task planner")),
        _Profiles(),
        _Catalogue(()),
        tabular_planner=LlmTabularPlanner(unexpected_tabular_planner),
    )
    request = {
        "batch_goal": "Assign cases",
        "sources": [{"type": "tabular", "documents": [{"source_ref": "attachments/owners.csv"}]}],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, Rejected)
    assert outcome.code == "preparation_no_available_execution"


@pytest.mark.asyncio
async def test_unavailable_persisted_source_is_operational_failure(context: ToolContext) -> None:
    class _UnavailableSources:
        def select(self, *args, **kwargs):
            raise SourceError(
                "persisted source object is unavailable",
                code="source_object_unavailable",
                details={"source_ref": "attachments/cases.csv"},
            )

    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("unavailable source must not plan")),
        _Profiles(),
        _Catalogue((_account_descriptor(),)),
        source_service=_UnavailableSources(),
    )
    request = {
        "batch_goal": "Import accounts",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "documents": [{"source_ref": "attachments/cases.csv"}],
            }
        ],
    }

    with pytest.raises(PreparationError) as excinfo:
        await service.prepare(context, json.dumps(request))

    assert excinfo.value.code == "source_object_unavailable"


@pytest.mark.asyncio
async def test_tabular_source_can_explicitly_allow_internal_capability(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "owners.csv"
    path.write_text("Case ID\nTC-1\n", encoding="utf-8")
    descriptor = CapabilityDescriptor(
        capability_id="skill:internal-cases.assign",
        parameter_schema={
            "type": "object",
            "properties": {"case_id": {"type": "string"}},
            "required": ["case_id"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
        visibility="internal",
    )
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("the internal capability is prebound")),
        _Profiles(),
        _Catalogue((descriptor,)),
    )
    request = {
        "batch_goal": "Assign cases",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:internal-cases.assign"],
                "documents": [{"source_ref": "attachments/owners.csv"}],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert outcome.batch.tasks[0].capability_id == "skill:internal-cases.assign"


@pytest.mark.asyncio
async def test_instruction_source_profile_allow_list_rejects_other_profile() -> None:
    profile = ProfileDescriptor("default", "General tasks", ALL_CAPABILITIES)

    class _VisibleProfiles:
        def list_profiles(self, query):
            return (profile, ProfileDescriptor("research", "Research tasks", ALL_CAPABILITIES))

        def resolve(self, profile_id):
            return None

    async def choose_disallowed_profile(batch_goal, unresolved, prebound, catalogue, profiles):
        return PlannerOutput(
            items=[PlannedDecision(item_id="instruction-01-001", dispatch_type="sub_agent", profile_id="research")]
        )

    service = DelegatePreparationService(
        LlmTaskPlanner(choose_disallowed_profile),
        _VisibleProfiles(),
        _Catalogue(()),
    )
    request = {
        "batch_goal": "Investigate",
        "sources": [
            {
                "type": "instructions",
                "items": [{"goal": "Investigate the issue"}],
                "profile_ids": ["default"],
            }
        ],
    }

    with pytest.raises(PreparationError) as excinfo:
        await service.prepare(
            context=ToolContext.model_construct(username="alice", agent_instance_id=0, project_id=0),
            request_json=json.dumps(request),
        )

    assert excinfo.value.code == "task_planner_invalid_output"


@pytest.mark.asyncio
async def test_duplicate_case_ids_get_unique_task_ids(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "owners.csv"
    long_case_id = "CASE-" + "x" * 200
    path.write_text(f"Case ID\nTC-1\nTC-1\n{long_case_id}\n", encoding="utf-8")
    descriptor = CapabilityDescriptor(
        capability_id="skill:cases.assign",
        parameter_schema={
            "type": "object",
            "properties": {"case_id": {"type": "string"}},
            "required": ["case_id"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("tabular rows are prebound")),
        _Profiles(),
        _Catalogue((descriptor,)),
    )
    request = {
        "batch_goal": "Assign duplicate cases",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:cases.assign"],
                "documents": [{"source_ref": "attachments/owners.csv"}],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)
    assert len(outcome.batch.tasks) == 3
    assert len({task.task_id for task in outcome.batch.tasks}) == 3
    assert all(len(task.task_id) <= 128 for task in outcome.batch.tasks)


@pytest.mark.asyncio
async def test_service_reports_invalid_transform_as_row_clarification(context: ToolContext, tmp_path: Path) -> None:
    path = tmp_path / "workspace" / "attachments" / "rows.csv"
    path.write_text("Retries\nnot-a-number\n", encoding="utf-8")
    descriptor = CapabilityDescriptor(
        capability_id="skill:jobs.run",
        parameter_schema={
            "type": "object",
            "properties": {"retries": {"type": "integer"}},
            "required": ["retries"],
        },
        required_sandbox=(),
        workspace_access="none",
        effect="mutate",
    )
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("invalid row must not plan")),
        _Profiles(),
        _Catalogue((descriptor,)),
    )
    request = {
        "batch_goal": "Run jobs",
        "sources": [
            {
                "type": "tabular",
                "capability_ids": ["skill:jobs.run"],
                "documents": [{"source_ref": "attachments/rows.csv"}],
                "parameter_bindings": {
                    "retries": {
                        "source": "column",
                        "column": "Retries",
                        "transform": "string_to_integer",
                    }
                },
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, NeedsClarification)
    assert outcome.code == "tabular_row_validation_failed"


@pytest.mark.asyncio
async def test_service_rejects_old_kind_options_contract(context: ToolContext) -> None:
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("invalid request must not plan")),
        _Profiles(),
        _Catalogue(()),
    )

    outcome = await service.prepare(context, '{"kind":"general","options_json":"{}"}')

    assert isinstance(outcome, Rejected)
    assert outcome.code == "delegate_request_invalid"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "internal_ref",
    [
        ".source-repository",
        "./.source-repository/objects/deadbeef/source.csv",
        "sico-source://objects/deadbeef/source.csv",
    ],
)
async def test_instruction_source_rejects_internal_source_object_reference(
    context: ToolContext,
    internal_ref: str,
) -> None:
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("invalid source ref must not plan")),
        _Profiles(),
        _Catalogue((_echo_descriptor(),)),
    )
    request = {
        "batch_goal": "Replay internal object",
        "sources": [
            {
                "type": "instructions",
                "capability_ids": ["builtin:echo"],
                "items": [
                    {
                        "goal": "Echo path",
                        "capability_id": "builtin:echo",
                        "params": {"message": internal_ref},
                    }
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, Rejected)
    assert outcome.code == "preparation_internal_source_ref"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "business_ref",
    ["sources/objects/catalog.json", ".sources/catalog.json", "business/.source-repository/catalog.json"],
)
async def test_instruction_source_allows_non_internal_path_segments(
    context: ToolContext,
    business_ref: str,
) -> None:
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("direct item must not plan")),
        _Profiles(),
        _Catalogue((_echo_descriptor(),)),
    )
    request = {
        "batch_goal": "Echo a business path",
        "sources": [
            {
                "type": "instructions",
                "capability_ids": ["builtin:echo"],
                "items": [
                    {
                        "goal": "Echo path",
                        "capability_id": "builtin:echo",
                        "params": {"message": business_ref},
                    }
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, PreparedTaskBatch)


@pytest.mark.asyncio
async def test_instruction_source_rejects_internal_source_path_inside_command(context: ToolContext) -> None:
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("internal source command must not plan")),
        _Profiles(),
        _Catalogue((_echo_descriptor(),)),
    )
    request = {
        "batch_goal": "Read internal source",
        "sources": [
            {
                "type": "instructions",
                "items": [
                    {
                        "goal": "Run cat .source-repository/objects/deadbeef/source.csv",
                        "params": {"command": "cat .source-repository/objects/deadbeef/source.csv"},
                    }
                ],
            }
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, Rejected)
    assert outcome.code == "preparation_internal_source_ref"


def test_delegate_request_rejects_too_many_sources() -> None:
    request = {
        "batch_goal": "Too many sources",
        "sources": [
            {"type": "instructions", "items": [{"goal": f"Task {index}"}]}
            for index in range(MAX_DELEGATE_SOURCES + 1)
        ],
    }

    outcome = parse_delegate_request(json.dumps(request))

    assert isinstance(outcome, Rejected)
    assert outcome.code == "delegate_request_limit"


def test_delegate_request_rejects_blank_case_id() -> None:
    request = {
        "batch_goal": "Run selected cases",
        "sources": [
            {
                "type": "tabular",
                "documents": [{"source_ref": "attachments/cases.csv", "case_ids": [" "]}],
            }
        ],
    }

    outcome = parse_delegate_request(json.dumps(request))

    assert isinstance(outcome, Rejected)
    assert outcome.code == "delegate_request_invalid"


def test_delegate_request_rejects_canonically_duplicate_document_scopes() -> None:
    request = {
        "batch_goal": "Do not select the same scope twice",
        "sources": [
            {
                "type": "tabular",
                "documents": [
                    {"source_ref": "attachments/cases.csv", "sheet_names": [" Cases "], "case_ids": ["ＴＣ-１"]},
                    {"source_ref": "/attachments\\cases.csv", "sheet_names": ["cases"], "case_ids": ["tc-1"]},
                ],
            }
        ],
    }

    outcome = parse_delegate_request(json.dumps(request))

    assert isinstance(outcome, Rejected)
    assert outcome.code == "delegate_request_invalid"


@pytest.mark.asyncio
async def test_service_stops_before_planning_when_mixed_request_exceeds_work_item_limit(
    context: ToolContext,
    tmp_path: Path,
) -> None:
    path = tmp_path / "workspace" / "attachments" / "accounts.csv"
    path.write_text("User Name\nalice\nbob\n", encoding="utf-8")
    service = DelegatePreparationService(
        LlmTaskPlanner(lambda *args: pytest.fail("over-limit request must not call the task planner")),
        _Profiles(),
        _Catalogue((_echo_descriptor(), _account_descriptor())),
    )
    request = {
        "batch_goal": "Reject oversized mixed batch",
        "sources": [
            {
                "type": "instructions",
                "capability_ids": ["builtin:echo"],
                "items": [
                    {
                        "goal": f"Echo {index}",
                        "capability_id": "builtin:echo",
                        "params": {"message": str(index)},
                    }
                    for index in range(MAX_DELEGATE_WORK_ITEMS - 1)
                ],
            },
            {
                "type": "tabular",
                "capability_ids": ["skill:accounts.import"],
                "documents": [{"source_ref": "attachments/accounts.csv"}],
            },
        ],
    }

    outcome = await service.prepare(context, json.dumps(request))

    assert isinstance(outcome, NeedsClarification)
    assert outcome.code == "delegate_work_item_limit"
    assert outcome.details["selected_work_items"] == MAX_DELEGATE_WORK_ITEMS + 1
    assert outcome.details["max_work_items"] == MAX_DELEGATE_WORK_ITEMS
