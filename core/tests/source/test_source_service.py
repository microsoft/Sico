from __future__ import annotations

import json
from concurrent.futures import ThreadPoolExecutor
from dataclasses import replace
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.biz.source import (
    SourceAccessContext,
    SourceError,
    TabularScope,
    WorkspaceSourceService,
    canonical_case_id,
    compact_manifest_payload,
)
from app.biz.source.persistence.repository import WorkspaceSourceRepository
from app.biz.source.persistence import repository as source_repository


@pytest.fixture
def source_access(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> tuple[SourceAccessContext, Path]:
    workspace = tmp_path / "workspace"
    (workspace / "attachments").mkdir(parents=True)
    monkeypatch.setattr("app.biz.source.service.CHAT_FS.get_workspace_path", lambda *args, **kwargs: workspace)
    return SourceAccessContext(username="alice", agent_instance_id=7, conversation_id=11), workspace


def test_excel_snapshot_preserves_typed_values(source_access: tuple[SourceAccessContext, Path]) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "accounts.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Accounts"
    sheet.append(["Username", "Retries", "Enabled", "Created"])
    created = datetime(2026, 8, 14, 9, 30)
    sheet.append(["alice", 3, True, created])
    workbook.save(path)

    service = WorkspaceSourceService()
    service.index_path(workspace, "attachments/accounts.xlsx", path)
    document = service.select(access, "attachments/accounts.xlsx")

    assert document.sheets[0].rows[0].values == {
        "Username": "alice",
        "Retries": 3,
        "Enabled": True,
        "Created": created,
    }


def test_content_snapshot_deduplicates_objects_but_preserves_refs(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workspace = tmp_path / "workspace"
    first = workspace / "attachments" / "first.csv"
    second = workspace / "knowledge" / "second.csv"
    first.parent.mkdir(parents=True)
    second.parent.mkdir(parents=True)
    first.write_text("Name\nAlice\n", encoding="utf-8")
    second.write_bytes(first.read_bytes())
    service = WorkspaceSourceService()

    first_manifest = service.index_path(workspace, "attachments/first.csv", first)
    second_manifest = service.index_path(workspace, "knowledge/second.csv", second)
    repository = WorkspaceSourceRepository(workspace)

    assert first_manifest.content_hash == second_manifest.content_hash
    first_loaded = repository.find("attachments/first.csv")
    second_loaded = repository.find("knowledge/second.csv")
    assert first_loaded.source_ref == "attachments/first.csv"
    assert second_loaded.source_ref == "knowledge/second.csv"
    assert first_loaded.source_id != second_loaded.source_id
    index_reads = 0
    original_index = repository._index

    def counting_index():
        nonlocal index_reads
        index_reads += 1
        return original_index()

    monkeypatch.setattr(repository, "_index", counting_index)
    assert len(repository.list_manifests()) == 2
    assert index_reads == 1


def test_source_object_hash_cache_invalidates_on_file_change(tmp_path: Path) -> None:
    path = tmp_path / "source.csv"
    path.write_bytes(b"old")
    source_repository._cached_file_hash.cache_clear()

    first = source_repository._file_hash(path)
    second = source_repository._file_hash(path)
    first_info = source_repository._cached_file_hash.cache_info()
    path.write_bytes(b"new content")
    changed = source_repository._file_hash(path)

    assert first == second
    assert changed != first
    assert first_info.hits == 1
    assert source_repository._cached_file_hash.cache_info().misses == 2


def test_equal_bytes_with_different_formats_use_distinct_snapshots(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    csv_path = workspace / "attachments" / "rows.csv"
    tsv_path = workspace / "attachments" / "rows.tsv"
    csv_path.parent.mkdir(parents=True)
    content = "Name,Retries\nAlice,3\n"
    csv_path.write_text(content, encoding="utf-8")
    tsv_path.write_text(content, encoding="utf-8")
    service = WorkspaceSourceService()

    csv_manifest = service.index_path(workspace, "attachments/rows.csv", csv_path)
    tsv_manifest = service.index_path(workspace, "attachments/rows.tsv", tsv_path)
    repository = WorkspaceSourceRepository(workspace)

    assert csv_manifest.content_hash == tsv_manifest.content_hash
    assert csv_manifest.format == "csv"
    assert tsv_manifest.format == "tsv"
    csv_document = repository.load_document(csv_manifest, ())
    assert csv_document.sheets[0].headers == ("Name", "Retries")
    assert csv_document.materialized_ref == ""
    assert repository.load_document(tsv_manifest, ()).sheets[0].headers == ("Name,Retries",)


def test_equal_csv_content_uses_path_independent_sheet_identity(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    first = workspace / "attachments" / "first.csv"
    second = workspace / "attachments" / "second.csv"
    first.parent.mkdir(parents=True)
    first.write_text("Name\nAlice\n", encoding="utf-8")
    second.write_bytes(first.read_bytes())
    service = WorkspaceSourceService()
    service.index_path(workspace, "attachments/first.csv", first)
    service.index_path(workspace, "attachments/second.csv", second)
    repository = WorkspaceSourceRepository(workspace)

    assert repository.find("attachments/first.csv").sheets[0].name == "table"
    assert repository.find("attachments/second.csv").sheets[0].name == "table"


def test_scope_auto_selects_data_but_rejects_multiple_data_sheets(
    source_access: tuple[SourceAccessContext, Path],
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "cases.xlsx"
    workbook = Workbook()
    cases = workbook.active
    cases.title = "Cases"
    cases.append(["Case ID", "Steps"])
    cases.append(["TC-1", "Open app"])
    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Count", 1])
    workbook.save(path)
    service = WorkspaceSourceService()
    service.index_path(workspace, "attachments/cases.xlsx", path)

    selected = service.select(access, "attachments/cases.xlsx")
    assert [sheet.name for sheet in selected.sheets] == ["Cases"]

    second = workbook.create_sheet("Regression")
    second.append(["Case ID", "Steps"])
    second.append(["TC-2", "Open settings"])
    workbook.save(path)

    with pytest.raises(SourceError) as excinfo:
        service.select(access, "attachments/cases.xlsx")

    assert excinfo.value.code == "tabular_sheet_scope_required"
    assert excinfo.value.details["runnable"] == ["Cases", "Regression"]


def test_explicit_sheet_cache_miss_persists_full_snapshot_and_object(
    source_access: tuple[SourceAccessContext, Path],
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "multi.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["Name"])
    first.append(["Alice"])
    second = workbook.create_sheet("Second")
    second.append(["Name"])
    second.append(["Bob"])
    workbook.save(path)

    selected = WorkspaceSourceService().select(
        access,
        "attachments/multi.xlsx",
        scope=TabularScope(sheet_names=("First",)),
    )
    manifest = WorkspaceSourceRepository(workspace).find("attachments/multi.xlsx")

    assert [sheet.name for sheet in selected.sheets] == ["First"]
    assert manifest is not None
    assert [sheet.name for sheet in manifest.sheets] == ["First", "Second"]
    repository = WorkspaceSourceRepository(workspace)
    assert not repository.root.is_relative_to(workspace)
    assert (repository.root / manifest.object_path).is_file()
    object_ref = repository.object_ref(manifest)
    assert object_ref.startswith("sico-source://objects/")
    assert repository.resolve_object_ref(object_ref) == repository.root / manifest.object_path


def test_case_id_uniquely_selects_sheet(source_access: tuple[SourceAccessContext, Path]) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "cases.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Smoke"
    first.append(["Case ID", "Steps"])
    first.append(["SMOKE-1", "Open app"])
    second = workbook.create_sheet("Regression")
    second.append(["Case ID", "Steps"])
    second.append(["REG-1", "Open settings"])
    workbook.save(path)
    service = WorkspaceSourceService()
    service.index_path(workspace, "attachments/cases.xlsx", path)

    selected = service.select(
        access,
        "attachments/cases.xlsx",
        scope=TabularScope(case_ids=("REG-1",)),
    )

    assert [sheet.name for sheet in selected.sheets] == ["Regression"]
    assert selected.sheets[0].rows[0].values["Case ID"] == "REG-1"


def test_case_id_matching_uses_one_unicode_canonical_form(
    source_access: tuple[SourceAccessContext, Path],
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "cases.csv"
    path.write_text("Case ID,Steps\nＴＣ-１,Open app\n", encoding="utf-8")
    service = WorkspaceSourceService()
    manifest = service.index_path(workspace, "attachments/cases.csv", path)

    selected = service.select(
        access,
        "attachments/cases.csv",
        scope=TabularScope(case_ids=("tc-1",)),
    )

    assert canonical_case_id("ＴＣ-１") == canonical_case_id("tc-1")
    assert manifest.case_ids == ("ＴＣ-１",)
    assert selected.sheets[0].rows[0].values["Case ID"] == "ＴＣ-１"


def test_case_ids_select_multiple_uniquely_matching_sheets(
    source_access: tuple[SourceAccessContext, Path],
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "cases.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Smoke"
    first.append(["Case ID", "Steps"])
    first.append(["SMOKE-1", "Open app"])
    second = workbook.create_sheet("Regression")
    second.append(["Case ID", "Steps"])
    second.append(["REG-1", "Open settings"])
    workbook.save(path)
    service = WorkspaceSourceService()
    service.index_path(workspace, "attachments/cases.xlsx", path)

    selected = service.select(
        access,
        "attachments/cases.xlsx",
        scope=TabularScope(case_ids=("REG-1", "SMOKE-1")),
    )

    assert [sheet.name for sheet in selected.sheets] == ["Smoke", "Regression"]


def test_duplicate_case_id_across_sheets_requires_scope(
    source_access: tuple[SourceAccessContext, Path],
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "cases.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Smoke"
    first.append(["Case ID", "Steps"])
    first.append(["QA-1", "Open app"])
    second = workbook.create_sheet("Regression")
    second.append(["Case ID", "Steps"])
    second.append(["QA-1", "Open settings"])
    workbook.save(path)
    service = WorkspaceSourceService()
    service.index_path(workspace, "attachments/cases.xlsx", path)

    with pytest.raises(SourceError) as excinfo:
        service.select(access, "attachments/cases.xlsx", scope=TabularScope(case_ids=("QA-1",)))

    assert excinfo.value.code == "tabular_sheet_scope_required"
    assert excinfo.value.details["ambiguous_case_ids"] == ["QA-1"]


def test_text_projection_attaches_to_existing_tabular_snapshot(
    source_access: tuple[SourceAccessContext, Path],
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "cases.csv"
    path.write_text("Case ID\nQA-1\n", encoding="utf-8")
    service = WorkspaceSourceService()
    original = service.index_path(workspace, "attachments/cases.csv", path)

    updated = service.index_text(access, "attachments/cases.csv", path, "QA-1 details", "Case summary")

    assert updated.content_hash == original.content_hash
    assert updated.sheets == original.sheets
    assert updated.case_ids == ("QA-1",)
    assert updated.summary == "Case summary"
    assert updated.content_chars == len("QA-1 details")


def test_text_projection_deduplicates_case_ids_without_changing_display(
    source_access: tuple[SourceAccessContext, Path],
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "notes.txt"
    path.write_text("qa-1 and QA-1", encoding="utf-8")

    manifest = WorkspaceSourceService().index_text(
        access,
        "attachments/notes.txt",
        path,
        "qa-1 and QA-1",
        "Case notes",
    )

    assert manifest.case_ids == ("qa-1",)


def test_text_first_projection_does_not_poison_tabular_index(
    source_access: tuple[SourceAccessContext, Path],
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "rows.csv"
    path.write_text("Name\nAlice\n", encoding="utf-8")
    service = WorkspaceSourceService()
    text_manifest = service.index_text(access, "attachments/rows.csv", path, "Name Alice", "Summary")

    indexed = service.index_path(workspace, "attachments/rows.csv", path)

    assert text_manifest.sheets == ()
    assert len(indexed.sheets) == 1
    assert indexed.sheets[0].headers == ("Name",)


def test_snapshot_preserves_structured_json_cells(source_access: tuple[SourceAccessContext, Path]) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "rows.jsonl"
    path.write_text(
        json.dumps({"values": {"Config": {"enabled": True}, "Tags": ["a", "b"]}}) + "\n",
        encoding="utf-8",
    )
    service = WorkspaceSourceService()
    service.index_path(workspace, "attachments/rows.jsonl", path)

    row = service.select(access, "attachments/rows.jsonl").sheets[0].rows[0]

    assert row.values["Config"] == {"enabled": True}
    assert row.values["Tags"] == ["a", "b"]


def test_concurrent_indexing_preserves_every_source_ref(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    attachments = workspace / "attachments"
    attachments.mkdir(parents=True)
    paths = []
    for index in range(8):
        path = attachments / f"rows-{index}.csv"
        path.write_text(f"Name\nUser {index}\n", encoding="utf-8")
        paths.append(path)

    def index_path(path: Path) -> None:
        WorkspaceSourceService().index_path(workspace, f"attachments/{path.name}", path)

    with ThreadPoolExecutor(max_workers=8) as executor:
        tuple(executor.map(index_path, paths))

    assert len(WorkspaceSourceRepository(workspace).list_manifests()) == len(paths)


def test_corrupt_manifest_is_skipped_and_rebuilt(source_access: tuple[SourceAccessContext, Path]) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "rows.csv"
    path.write_text("Name\nAlice\n", encoding="utf-8")
    service = WorkspaceSourceService()
    service.index_path(workspace, "attachments/rows.csv", path)
    repository = WorkspaceSourceRepository(workspace)
    index = json.loads((repository.root / "index.json").read_text(encoding="utf-8"))
    manifest_path = repository.root / index["attachments/rows.csv"]["manifest_path"]
    manifest_path.write_text("{broken", encoding="utf-8")

    assert WorkspaceSourceRepository(workspace).list_manifests() == ()
    selected = service.select(access, "attachments/rows.csv")

    assert selected.sheets[0].rows[0].values == {"Name": "Alice"}


def test_repository_rejects_manifest_path_outside_source_root(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    source_root = WorkspaceSourceRepository(workspace).root
    source_root.mkdir(parents=True)
    workspace.mkdir()
    outside = workspace / "outside.json"
    outside.write_text("{}", encoding="utf-8")
    (source_root / "index.json").write_text(
        json.dumps({"attachments/rows.csv": {"manifest_path": "../outside.json"}}),
        encoding="utf-8",
    )

    assert WorkspaceSourceRepository(workspace).find("attachments/rows.csv") is None


def test_repository_rejects_tampered_object_and_snapshot_paths(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    path = workspace / "attachments" / "rows.csv"
    path.parent.mkdir(parents=True)
    path.write_text("Name\nAlice\n", encoding="utf-8")
    service = WorkspaceSourceService()
    manifest = service.index_path(workspace, "attachments/rows.csv", path)
    repository = WorkspaceSourceRepository(workspace)
    bad_object = replace(manifest, object_path="attachments/rows.csv")
    bad_sheet = replace(manifest.sheets[0], snapshot_path="../../outside.jsonl")
    bad_snapshot = replace(manifest, sheets=(bad_sheet,))

    assert repository.object_ref(bad_object) == ""
    with pytest.raises(ValueError, match="escapes root"):
        repository.load_document(bad_snapshot, ())


def test_replace_refs_removes_only_requested_namespace(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    attachment = workspace / "attachments" / "rows.csv"
    first = workspace / "knowledge" / "1" / "first.csv"
    second = workspace / "knowledge" / "2" / "second.csv"
    for index, path in enumerate((attachment, first, second)):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(f"Name\nUser {index}\n", encoding="utf-8")
    service = WorkspaceSourceService()
    attachment_manifest = service.index_path(workspace, "attachments/rows.csv", attachment)
    first_manifest = service.index_path(workspace, "knowledge/1/first.csv", first)
    second_manifest = service.index_path(workspace, "knowledge/2/second.csv", second)
    repository = WorkspaceSourceRepository(workspace)

    repository.replace_refs("knowledge/", (second_manifest,))

    assert repository.find(attachment_manifest.source_ref) is not None
    assert repository.find(first_manifest.source_ref) is None
    assert repository.find(second_manifest.source_ref) is not None
    active_snapshot = repository.root / "snapshots" / (
        f"{second_manifest.content_hash}.{second_manifest.format}.p{second_manifest.parser_version}"
    )
    orphan_snapshot = repository.root / "snapshots" / (
        f"{first_manifest.content_hash}.{first_manifest.format}.p{first_manifest.parser_version}"
    )
    assert not (active_snapshot / ".orphaned-at").exists()
    assert (orphan_snapshot / ".orphaned-at").exists()


def test_missing_table_snapshot_is_rebuilt(source_access: tuple[SourceAccessContext, Path]) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "rows.csv"
    path.write_text("Name\nAlice\n", encoding="utf-8")
    service = WorkspaceSourceService()
    manifest = service.index_path(workspace, "attachments/rows.csv", path)
    repository = WorkspaceSourceRepository(workspace)
    index = json.loads((repository.root / "index.json").read_text(encoding="utf-8"))
    manifest_path = repository.root / index["attachments/rows.csv"]["manifest_path"]
    (manifest_path.parent / manifest.sheets[0].snapshot_path).unlink()

    selected = service.select(access, "attachments/rows.csv")

    assert selected.sheets[0].rows[0].values == {"Name": "Alice"}


def test_snapshot_remains_readable_after_original_source_is_removed(
    source_access: tuple[SourceAccessContext, Path],
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "rows.csv"
    path.write_text("Name\nAlice\n", encoding="utf-8")
    service = WorkspaceSourceService()
    service.index_path(workspace, "attachments/rows.csv", path)
    path.unlink()

    selected = service.select(access, "attachments/rows.csv")

    assert selected.sheets[0].rows[0].values == {"Name": "Alice"}


def test_text_projection_reindexes_after_parser_version_change(
    source_access: tuple[SourceAccessContext, Path],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "notes.txt"
    path.write_text("QA-1 details", encoding="utf-8")
    service = WorkspaceSourceService()
    original = service.index_text(access, "attachments/notes.txt", path, "QA-1 details", "Original")
    monkeypatch.setattr("app.biz.source.service.PARSER_VERSION", original.parser_version + 1)

    updated = service.index_text(access, "attachments/notes.txt", path, "QA-1 details", "Updated")

    assert updated.parser_version == original.parser_version + 1
    assert updated.summary == "Updated"


def test_text_projection_restores_missing_raw_object(source_access: tuple[SourceAccessContext, Path]) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "notes.txt"
    content = "QA-1 details"
    path.write_text(content, encoding="utf-8")
    service = WorkspaceSourceService()
    manifest = service.index_text(access, "attachments/notes.txt", path, content, "Summary")
    repository = WorkspaceSourceRepository(workspace)
    object_path = repository.root / manifest.object_path
    object_path.unlink()

    updated = service.index_text(access, "attachments/notes.txt", path, content, "Summary")

    assert updated.object_path == manifest.object_path
    assert object_path.read_text(encoding="utf-8") == content


def test_live_source_repairs_same_size_corrupt_raw_object(source_access: tuple[SourceAccessContext, Path]) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "rows.csv"
    original = b"Name\nAlice\n"
    path.write_bytes(original)
    service = WorkspaceSourceService()
    manifest = service.index_path(workspace, "attachments/rows.csv", path)
    repository = WorkspaceSourceRepository(workspace)
    object_path = repository.root / manifest.object_path
    object_path.write_bytes(b"X" * len(original))

    service.select(access, "attachments/rows.csv")

    assert object_path.read_bytes() == original


def test_compact_manifest_bounds_headers_and_case_ids(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    path = workspace / "attachments" / "wide.csv"
    path.parent.mkdir(parents=True)
    headers = ["Case ID", *(f"Column {index}" for index in range(60))]
    rows = [headers, *[[f"QA-{index}", *(str(index) for _ in range(60))] for index in range(30)]]
    path.write_text("\n".join(",".join(row) for row in rows), encoding="utf-8")
    manifest = WorkspaceSourceService().index_path(workspace, "attachments/wide.csv", path)

    payload = compact_manifest_payload(manifest)
    sheet = payload["sheets"][0]

    assert sheet["header_count"] == 61
    assert len(sheet["headers"]) == 50
    assert sheet["omitted_header_count"] == 11
    assert sheet["case_id_count"] == 30
    assert len(sheet["case_ids"]) == 20
    assert sheet["omitted_case_id_count"] == 10
    assert sheet["runnable"] is True


def test_reader_rejects_excessive_columns(source_access: tuple[SourceAccessContext, Path]) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "wide.csv"
    path.write_text(",".join(f"Column {index}" for index in range(501)) + "\n", encoding="utf-8")

    with pytest.raises(SourceError) as excinfo:
        WorkspaceSourceService().select(access, "attachments/wide.csv")

    assert excinfo.value.code == "tabular_column_limit"


def test_reader_rejects_oversized_cell(
    source_access: tuple[SourceAccessContext, Path],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "large-cell.csv"
    path.write_text("Value\n123456\n", encoding="utf-8")
    monkeypatch.setattr("app.biz.source.tabular.reader._TABULAR_MAX_CELL_CHARS", 5)

    with pytest.raises(SourceError) as excinfo:
        WorkspaceSourceService().select(access, "attachments/large-cell.csv")

    assert excinfo.value.code == "tabular_cell_limit"


def test_reader_rejects_decoded_cell_budget(
    source_access: tuple[SourceAccessContext, Path],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    access, workspace = source_access
    path = workspace / "attachments" / "many-cells.csv"
    path.write_text("Name,Value\nAlice,1\nBob,2\n", encoding="utf-8")
    monkeypatch.setattr("app.biz.source.tabular.reader._TABULAR_MAX_DECODED_CELLS", 5)

    with pytest.raises(SourceError) as excinfo:
        WorkspaceSourceService().select(access, "attachments/many-cells.csv")

    assert excinfo.value.code == "tabular_decoded_size_limit"


def test_orphan_snapshot_is_deleted_only_after_grace_period(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workspace = tmp_path / "workspace"
    path = workspace / "knowledge" / "1" / "rows.csv"
    path.parent.mkdir(parents=True)
    path.write_text("Name\nAlice\n", encoding="utf-8")
    manifest = WorkspaceSourceService().index_path(workspace, "knowledge/1/rows.csv", path)
    repository = WorkspaceSourceRepository(workspace)
    snapshot = repository.root / "snapshots" / (
        f"{manifest.content_hash}.{manifest.format}.p{manifest.parser_version}"
    )
    monkeypatch.setattr("app.biz.source.persistence.repository._ORPHAN_RETENTION_SECONDS", 0)

    repository.replace_refs("knowledge/", ())
    assert snapshot.exists()
    repository.replace_refs("knowledge/", ())

    assert not snapshot.exists()
