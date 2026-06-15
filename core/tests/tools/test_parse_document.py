# Copyright (c) 2026 Sico Authors
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from __future__ import annotations

import asyncio
import json
from pathlib import Path
from types import SimpleNamespace

import pytest

# Side-effect import: registers the workbook post-parse hook with
# ``parse_document`` so this module's tests can assert workbook archiving and
# ``workbook_manifest`` response fields when run in isolation.
import app.biz.chat.adapters.workbook  # noqa: F401
from app.schemas.conversation.plan import Plan
from app.tools.common import ToolContext
from app.tools.parse_document import (
    _PARSE_DOCUMENT_CACHE,
    _PARSE_DOCUMENT_CACHE_MAX_ENTRIES,
    _PARSE_DOCUMENT_LOCK_REFS,
    _PARSE_DOCUMENT_LOCKS,
    _parse_document_func,
    _prune_parse_document_cache,
)


class _FakeExtractor:
    def __init__(self) -> None:
        self.calls = 0

    async def extract(self, _path: str):
        self.calls += 1
        return "STCAQA-567 full document text", "short summary"


class _FakePlanEditor:
    def __init__(self):
        self.plan: Plan | None = None
        self.next_id = 0
        self.messages: dict[int, str] = {}
        self.updated_tool_call_ids: set[int] = set()

    async def get_plan(self):
        return self.plan

    async def update_plan(self, plan: Plan) -> None:
        self.plan = plan

    async def create_tool_call(self, name, initial_message, execution_info=None, parent_tool_call_id=None, sub_call_index=0):
        self.next_id += 1
        self.messages[self.next_id] = initial_message
        return self.next_id

    async def update_tool_call_message(self, tool_call_id: int, message: str):
        self.messages[tool_call_id] = message
        return None

    async def update_tool_call(self, tool_call_id: int, updater):
        tool_call = SimpleNamespace()
        updater(tool_call)
        self.updated_tool_call_ids.add(tool_call_id)
        return tool_call


@pytest.mark.asyncio
async def test_parse_document_publishes_plan_progress(tmp_path: Path, monkeypatch) -> None:
    document = tmp_path / "cases.xlsx"
    document.write_bytes(b"xlsx")
    plan_editor = _FakePlanEditor()
    ctx = ToolContext.model_construct(
        username="alice@example.com",
        agent_id="agent",
        agent_instance_id=1,
        turn_id=7,
        project_id=1,
        conversation_id=42,
        response_queue=asyncio.Queue(),
        plan_editor=plan_editor,
    )
    invocation_ctx = SimpleNamespace(kwargs={"tool_context": ctx})

    extractor = _FakeExtractor()
    monkeypatch.setattr("app.tools.parse_document._get_extractor", lambda: extractor)
    monkeypatch.setattr("app.tools.parse_document.CHAT_FS.resolve_workspace_file", lambda *_args: document)
    monkeypatch.setattr("app.tools.parse_document.CHAT_FS.get_workspace_path", lambda *_args: tmp_path / "workspace_7")
    monkeypatch.setattr(
        "app.tools.parse_document.CHAT_FS.write_file",
        lambda _agent_instance_id, _username, _path, _content: None,
    )

    result = await _parse_document_func(invocation_ctx, file_path="attachments/cases.xlsx")

    assert result["error_message"] == ""
    assert extractor.calls == 1
    assert plan_editor.plan is not None
    assert plan_editor.plan.title == "Document Preparation"
    assert plan_editor.messages[1].startswith("Parsed document: attachments/cases.xlsx")
    assert "extracted 29 characters" in plan_editor.messages[1]
    assert 1 not in plan_editor.updated_tool_call_ids

    manifests = list((tmp_path / "workspace_7" / "case_sources" / "parsed_documents").glob("*.json"))
    assert len(manifests) == 1
    payload = json.loads(manifests[0].read_text(encoding="utf-8"))
    assert payload["case_ids"] == ["STCAQA-567"]
    archived = manifests[0].with_name(payload["archived_markdown_path"])
    assert archived.read_text(encoding="utf-8") == "STCAQA-567 full document text"


@pytest.mark.asyncio
async def test_parse_document_rejects_plain_workspace_path_without_plan() -> None:
    plan_editor = _FakePlanEditor()
    ctx = ToolContext.model_construct(
        username="alice@example.com",
        agent_id="agent",
        agent_instance_id=1,
        turn_id=9,
        project_id=1,
        conversation_id=42,
        response_queue=asyncio.Queue(),
        plan_editor=plan_editor,
    )
    invocation_ctx = SimpleNamespace(kwargs={"tool_context": ctx})

    result = await _parse_document_func(invocation_ctx, file_path="README.md")

    assert result["error_message"].startswith("parse_document only accepts")
    assert plan_editor.next_id == 0


@pytest.mark.asyncio
async def test_parse_document_reuses_same_turn_result_without_duplicate_plan(tmp_path: Path, monkeypatch) -> None:
    _PARSE_DOCUMENT_CACHE.clear()
    _PARSE_DOCUMENT_LOCKS.clear()
    _PARSE_DOCUMENT_LOCK_REFS.clear()
    document = tmp_path / "cases.xlsx"
    document.write_bytes(b"xlsx")
    plan_editor = _FakePlanEditor()
    ctx = ToolContext.model_construct(
        username="alice@example.com",
        agent_id="agent",
        agent_instance_id=1,
        turn_id=8,
        project_id=1,
        conversation_id=42,
        response_queue=asyncio.Queue(),
        plan_editor=plan_editor,
    )
    invocation_ctx = SimpleNamespace(kwargs={"tool_context": ctx})
    extractor = _FakeExtractor()

    monkeypatch.setattr("app.tools.parse_document._get_extractor", lambda: extractor)
    monkeypatch.setattr("app.tools.parse_document.CHAT_FS.resolve_workspace_file", lambda *_args: document)
    monkeypatch.setattr("app.tools.parse_document.CHAT_FS.get_workspace_path", lambda *_args: tmp_path / "workspace_8a")
    monkeypatch.setattr(
        "app.tools.parse_document.CHAT_FS.write_file",
        lambda _agent_instance_id, _username, _path, _content: None,
    )

    first, second = await asyncio.gather(
        _parse_document_func(invocation_ctx, file_path="attachments/cases.xlsx"),
        _parse_document_func(invocation_ctx, file_path="attachments/cases.xlsx"),
    )

    assert first == second
    assert extractor.calls == 1
    assert plan_editor.next_id == 1
    assert list(plan_editor.messages) == [1]
    assert _PARSE_DOCUMENT_LOCKS == {}
    assert _PARSE_DOCUMENT_LOCK_REFS == {}


@pytest.mark.asyncio
async def test_parse_document_cache_is_scoped_by_conversation(tmp_path: Path, monkeypatch) -> None:
    _PARSE_DOCUMENT_CACHE.clear()
    _PARSE_DOCUMENT_LOCKS.clear()
    _PARSE_DOCUMENT_LOCK_REFS.clear()
    document = tmp_path / "cases.xlsx"
    document.write_bytes(b"xlsx")
    extractor = _FakeExtractor()

    monkeypatch.setattr("app.tools.parse_document._get_extractor", lambda: extractor)
    monkeypatch.setattr("app.tools.parse_document.CHAT_FS.resolve_workspace_file", lambda *_args: document)
    monkeypatch.setattr("app.tools.parse_document.CHAT_FS.get_workspace_path", lambda *_args: tmp_path / "workspace_8b")
    monkeypatch.setattr(
        "app.tools.parse_document.CHAT_FS.write_file",
        lambda _agent_instance_id, _username, _path, _content: None,
    )

    first_ctx = ToolContext.model_construct(
        username="alice@example.com",
        agent_id="agent",
        agent_instance_id=1,
        turn_id=8,
        project_id=1,
        conversation_id=42,
        response_queue=asyncio.Queue(),
        plan_editor=_FakePlanEditor(),
    )
    second_ctx = ToolContext.model_construct(
        username="alice@example.com",
        agent_id="agent",
        agent_instance_id=1,
        turn_id=8,
        project_id=1,
        conversation_id=43,
        response_queue=asyncio.Queue(),
        plan_editor=_FakePlanEditor(),
    )

    first = await _parse_document_func(
        SimpleNamespace(kwargs={"tool_context": first_ctx}),
        file_path="attachments/cases.xlsx",
    )
    second = await _parse_document_func(
        SimpleNamespace(kwargs={"tool_context": second_ctx}),
        file_path="attachments/cases.xlsx",
    )

    assert first["error_message"] == ""
    assert second["error_message"] == ""
    assert extractor.calls == 2


def test_parse_document_cache_prunes_oldest_entries() -> None:
    _PARSE_DOCUMENT_CACHE.clear()

    for index in range(_PARSE_DOCUMENT_CACHE_MAX_ENTRIES + 2):
        _PARSE_DOCUMENT_CACHE[(1, "alice", index, 1, f"attachments/{index}.xlsx")] = {"error_message": ""}

    _prune_parse_document_cache()

    assert len(_PARSE_DOCUMENT_CACHE) == _PARSE_DOCUMENT_CACHE_MAX_ENTRIES
    assert (1, "alice", 0, 1, "attachments/0.xlsx") not in _PARSE_DOCUMENT_CACHE
    assert (1, "alice", 1, 1, "attachments/1.xlsx") not in _PARSE_DOCUMENT_CACHE
    assert (1, "alice", 2, 1, "attachments/2.xlsx") in _PARSE_DOCUMENT_CACHE


@pytest.mark.asyncio
async def test_parse_document_returns_workbook_manifest_for_multi_sheet_file(tmp_path: Path, monkeypatch) -> None:
    from openpyxl import Workbook

    _PARSE_DOCUMENT_CACHE.clear()
    _PARSE_DOCUMENT_LOCKS.clear()
    document = tmp_path / "multi_sheet_cases.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    summary.append(["metric", "value"])
    summary.append(["total_rows", 3])
    master = workbook.create_sheet("master")
    master.append(["source_file", "source_row", "ID", "Title"])
    master.append(["cases_a.csv", 1, "A-1", "case one"])
    cases_a = workbook.create_sheet("cases_a")
    cases_a.append(["ID", "Title", "Steps"])
    cases_a.append(["A-1", "case one", "open app"])
    cases_b = workbook.create_sheet("cases_b")
    cases_b.append(["ID", "Title", "Steps"])
    cases_b.append(["B-1", "case two", "open settings"])
    cases_b.append(["B-2", "case three", "save"])
    workbook.save(document)

    plan_editor = _FakePlanEditor()
    ctx = ToolContext.model_construct(
        username="alice@example.com",
        agent_id="agent",
        agent_instance_id=1,
        turn_id=10,
        project_id=1,
        conversation_id=42,
        response_queue=asyncio.Queue(),
        plan_editor=plan_editor,
    )
    invocation_ctx = SimpleNamespace(kwargs={"tool_context": ctx})

    extractor = _FakeExtractor()
    monkeypatch.setattr("app.tools.parse_document._get_extractor", lambda: extractor)
    monkeypatch.setattr("app.tools.parse_document.CHAT_FS.resolve_workspace_file", lambda *_args: document)
    monkeypatch.setattr("app.tools.parse_document.CHAT_FS.get_workspace_path", lambda *_args: tmp_path / "workspace_10")
    monkeypatch.setattr(
        "app.tools.parse_document.CHAT_FS.write_file",
        lambda _agent_instance_id, _username, _path, _content: None,
    )

    result = await _parse_document_func(invocation_ctx, file_path="attachments/multi_sheet_cases.xlsx")

    assert result["error_message"] == ""
    manifest = result["workbook_manifest"]
    assert manifest["sheet_count"] == 4
    assert manifest["total_data_rows"] == 5
    assert manifest["runnable_data_rows"] == 3
    assert manifest["source_data_rows"] == 3
    assert manifest["master_data_rows"] == 1
    assert manifest["summary_data_rows"] == 1
    assert manifest["multiple_data_sheets"] is True
    assert manifest["requires_scope_selection"] is True
    assert manifest["contains_master_sheet"] is True
    assert manifest["scope_confirmation_hint"].startswith("Workbook has multiple runnable sheets")
    assert [(sheet["name"], sheet["kind"], sheet["data_rows"]) for sheet in manifest["sheets"]] == [
        ("summary", "summary", 1),
        ("master", "master", 1),
        ("cases_a", "data", 1),
        ("cases_b", "data", 2),
    ]
    assert "detected 3 runnable data rows" in plan_editor.messages[1]
    assert "workbook sheets: summary (1 data rows, summary); master (1 data rows, master)" in plan_editor.messages[1]

    manifests = list((tmp_path / "workspace_10" / "case_sources" / "parsed_documents").glob("*.json"))
    payload = json.loads(manifests[0].read_text(encoding="utf-8"))
    assert payload["data_rows"] == 3
    assert payload["workbook_manifest"]["multiple_data_sheets"] is True
    assert payload["workbook_manifest"]["runnable_data_rows"] == 3
    assert [(source["sheet_name"], source["case_count"]) for source in payload["workbook_case_sources"]] == [
        ("master", 1),
        ("cases_a", 1),
        ("cases_b", 2),
    ]
    cases_b_source = next(source for source in payload["workbook_case_sources"] if source["sheet_name"] == "cases_b")
    cases_b_path = manifests[0].with_name(cases_b_source["case_source_path"])
    cases_b_lines = [json.loads(line) for line in cases_b_path.read_text(encoding="utf-8").splitlines()]
    assert [case["case_id"] for case in cases_b_lines] == ["B-1", "B-2"]
    assert "open settings" in cases_b_lines[0]["instructions"]
