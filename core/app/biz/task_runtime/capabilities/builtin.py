"""The builtin capability provider: ``echo`` / ``file_convert`` / ``run_command``.

These are the runtime's own payloads, declared once in
:mod:`~app.biz.task_runtime.capabilities.tool_catalog` and projected here into descriptors and
handlers. ``echo`` and ``file_convert`` run in-process; ``run_command`` lowers to
a :class:`CommandSpec` and lets the injected :class:`CommandBackend` decide
*where* it runs (host / docker / k8s).
"""

from __future__ import annotations

import contextlib
import csv
import json
import os
import re
from pathlib import Path
from typing import Any

from ..storage.artifact_store import ArtifactStore
from ..domain.models import ArtifactRef, ErrorClass, TaskResult, TaskRun, TaskStatus
from ..execution.naming import sanitize_dns_label
from ..domain.results import build_user_input_result
from ..domain.time import now_ms as _now_ms
from .tool_catalog import (
    ECHO_TOOL_NAME,
    FILE_CONVERT_TOOL_NAME,
    RUN_COMMAND_TOOL_NAME,
)
from ..execution.command.contracts import (
    CommandBackend,
    CommandMount,
    CommandResult,
    CommandSpec,
    readonly_input_mounts,
    truncate_stream,
)
from .catalogue import builtin_descriptors
from .descriptors import (
    CapabilityBinding,
    CapabilityContext,
    CapabilityDescriptor,
    CatalogueQuery,
    ResolveContext,
)
from .ids import BUILTIN_PROVIDER_ID, builtin_tool_of

_WORKSPACE_MOUNT_NAME = "workspace"
_RESULT_MOUNT_NAME = "result"
_COMMAND_SUMMARY_HEAD = 80
_COMMAND_STDOUT_HEAD = 1000
_COMMAND_STDERR_HEAD = 500
_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
_WORKSPACE_SPREADSHEET_RE = re.compile(r"(?:attachments|download)/[^\n\r\"'`]+?\.(?:xlsx|xlsm)", re.IGNORECASE)


class BuiltinCapabilityProvider:
    """Serves the runtime's own payloads as capabilities."""

    provider_id = BUILTIN_PROVIDER_ID

    def __init__(self, *, artifact_store: ArtifactStore, command_backend: CommandBackend) -> None:
        if artifact_store is None:
            raise ValueError("artifact_store is required")
        if command_backend is None:
            raise ValueError("command_backend is required")
        self.artifact_store = artifact_store
        self.command_backend = command_backend
        self._descriptors = {descriptor.capability_id: descriptor for descriptor in builtin_descriptors()}

    async def list_descriptors(self, query: CatalogueQuery) -> tuple[CapabilityDescriptor, ...]:
        return tuple(descriptor for descriptor in self._descriptors.values() if query.matches(descriptor))

    async def resolve(self, capability_id: str, context: ResolveContext) -> CapabilityBinding | None:
        descriptor = self._descriptors.get(capability_id)
        if descriptor is None:
            return None
        return CapabilityBinding(descriptor=descriptor, handler=_BuiltinHandler(self, builtin_tool_of(capability_id)))


class _BuiltinHandler:
    """Dispatches one builtin payload to its implementation."""

    def __init__(self, provider: BuiltinCapabilityProvider, tool_name: str) -> None:
        self._provider = provider
        self._tool_name = tool_name

    async def execute(self, context: CapabilityContext) -> TaskResult:
        if self._tool_name == ECHO_TOOL_NAME:
            return _run_echo(context)
        if self._tool_name == FILE_CONVERT_TOOL_NAME:
            return _run_file_convert(context, self._provider.artifact_store)
        if self._tool_name == RUN_COMMAND_TOOL_NAME:
            return await _run_command(context, self._provider.command_backend)
        return build_user_input_result(context.run, f"Unsupported builtin capability payload: {self._tool_name}")


# ---------------------------------------------------------------------------
# echo
# ---------------------------------------------------------------------------


def _run_echo(context: CapabilityContext) -> TaskResult:
    run = context.run
    now_ms = _now_ms()
    message = str(context.arguments.get("message") or run.spec.instructions or run.spec.title)
    return TaskResult(
        run_id=run.run_id,
        task_id=run.spec.task_id,
        status=TaskStatus.COMPLETED,
        title=run.spec.title,
        summary=message,
        output=message,
        started_at=now_ms,
        ended_at=now_ms,
        duration_ms=0,
    )


# ---------------------------------------------------------------------------
# run_command
# ---------------------------------------------------------------------------


async def _run_command(context: CapabilityContext, backend: CommandBackend) -> TaskResult:
    """Run a shell command via the selected :class:`CommandBackend`.

    This is the single ``run_command`` implementation: the handler builds a
    :class:`CommandSpec` and hands it to a per-run :class:`CommandSession`, so
    *where* the command runs (local/docker/k8s) is decided entirely by the
    backend. A sub-agent reaches the same code through its capability invoker
    rather than re-implementing command execution.
    """
    run = context.run
    command = str(context.arguments.get("command") or "").strip()
    if not command:
        return build_user_input_result(run, "run_command requires a non-empty args.command")
    spec = _command_spec(context, command)
    session = backend.open_session(
        pod_name=_command_pod_name(run),
        image=str(context.arguments.get("image") or ""),
    )
    try:
        outcome = await session.run(spec)
    finally:
        await session.aclose()
    return _command_result_to_task_result(context, command, outcome)


def _command_spec(context: CapabilityContext, command: str) -> CommandSpec:
    # Commands start in the shared workspace for natural relative reads. Whether
    # that mount is writable is the descriptor's call, not the executor's;
    # durable outputs always belong under SICO_RESULT_DIR.
    run = context.run
    result_str = str(context.result_dir)
    workspace_str = str(context.workspace)
    mounts = [
        CommandMount(name=_RESULT_MOUNT_NAME, host_path=result_str, mount_path=result_str),
        CommandMount(
            name=_WORKSPACE_MOUNT_NAME,
            host_path=workspace_str,
            mount_path=workspace_str,
            read_only=not context.descriptor.workspace_is_writable,
        ),
        *readonly_input_mounts(context.input_paths),
    ]
    return CommandSpec(
        argv=["sh", "-lc", command],
        cwd=workspace_str,
        env={**_command_env(run), "SICO_WORKSPACE_DIR": workspace_str, "SICO_RESULT_DIR": result_str},
        mounts=mounts,
        timeout_seconds=_command_timeout_seconds(context),
        metadata={
            "agent_instance_id": str(run.agent_instance_id),
            "user_label": sanitize_dns_label(run.username, max_len=63),
        },
    )


def _command_env(run: TaskRun) -> dict[str, str]:
    """Per-run env overlay for ``run_command``.

    Only SICO_* identity vars are declared; the local backend merges these over
    ``os.environ`` and container backends forward them as ``-e``/pod env, so the
    set is intentionally small (no host environment leakage into containers).
    """
    return {
        "SICO_TASK_RUN_ID": run.run_id,
        "SICO_AGENT_INSTANCE_ID": str(run.agent_instance_id),
        "SICO_PROJECT_ID": str(run.project_id),
        "SICO_APP_NAME": _sico_app_name(),
    }


def _sico_app_name() -> str:
    return os.getenv("SICO_APP_NAME", "sico").strip() or "sico"


def _command_timeout_seconds(context: CapabilityContext) -> int:
    requested = context.arguments.get("timeout")
    if requested:
        with contextlib.suppress(TypeError, ValueError):
            return max(0, int(requested))
    return context.run.execution_policy.timeout_seconds


def _command_pod_name(run: TaskRun) -> str:
    return sanitize_dns_label(f"task-{run.run_id}", max_len=63)


def _command_summary(command: str, outcome: CommandResult, status: TaskStatus) -> str:
    head = command if len(command) <= _COMMAND_SUMMARY_HEAD else command[: _COMMAND_SUMMARY_HEAD - 3] + "..."
    if status == TaskStatus.COMPLETED:
        lines = [f"`{head}` finished with exit code {outcome.return_code}"]
    elif status == TaskStatus.TIMED_OUT:
        lines = [f"`{head}` timed out"]
    elif outcome.system_error:
        lines = [f"`{head}` failed to run: {outcome.system_error}"]
    else:
        lines = [f"`{head}` failed with exit code {outcome.return_code}"]
    # Fold the command's output into the summary so the caller sees it via the
    # batch digest (which only carries ``summary``), not just the exit code.
    stdout = truncate_stream(outcome.stdout, _COMMAND_STDOUT_HEAD)
    if stdout:
        lines.append(f"stdout:\n{stdout}")
    if status != TaskStatus.COMPLETED:
        stderr = truncate_stream(outcome.stderr, _COMMAND_STDERR_HEAD)
        if stderr:
            lines.append(f"stderr:\n{stderr}")
    return "\n".join(lines)


def _command_result_to_task_result(context: CapabilityContext, command: str, outcome: CommandResult) -> TaskResult:
    run = context.run
    finished_at = _now_ms()
    timed_out = outcome.return_code == -1 and "timed out" in outcome.system_error.lower()
    if timed_out:
        status = TaskStatus.TIMED_OUT
    elif outcome.system_error or outcome.return_code != 0:
        status = TaskStatus.FAILED
    else:
        status = TaskStatus.COMPLETED
    error_class: ErrorClass | None = None
    error_message = ""
    if status == TaskStatus.TIMED_OUT:
        error_class = ErrorClass.TIMEOUT
        error_message = outcome.system_error or f"command timed out after {_command_timeout_seconds(context)}s"
    elif status == TaskStatus.FAILED:
        error_class = ErrorClass.TRANSIENT if outcome.system_error else ErrorClass.SKILL_RUNTIME
        error_message = outcome.system_error or outcome.stderr or f"command exited with {outcome.return_code}"
    return TaskResult(
        run_id=run.run_id,
        task_id=run.spec.task_id,
        status=status,
        title=run.spec.title,
        summary=_command_summary(command, outcome, status),
        output=outcome.stdout,
        error_class=error_class,
        error_message=error_message,
        sandbox=run.sandbox,
        started_at=context.started_at,
        ended_at=finished_at,
        duration_ms=max(0, finished_at - context.started_at),
    )


# ---------------------------------------------------------------------------
# file_convert
# ---------------------------------------------------------------------------


def _run_file_convert(context: CapabilityContext, artifact_store: ArtifactStore) -> TaskResult:
    run = context.run
    try:
        # Inputs are read from the shared workspace; converted outputs are
        # written to this run's writable result directory.
        read_root = context.workspace
        result_dir = context.result_dir
        requests = _file_conversion_requests(context)
        target_format = str(context.arguments.get("target_format") or "csv").lower().lstrip(".")
        if target_format != "csv":
            raise ValueError(f"file_convert only supports target_format=csv, got: {target_format}")
        output_dir = _relative_dir(str(context.arguments.get("output_dir") or "output/csv"))
        artifacts: list[ArtifactRef] = []
        files: list[dict[str, Any]] = []
        for input_path in requests:
            source = _workspace_file(read_root, input_path)
            target = _workspace_file(result_dir, f"{output_dir}/{Path(input_path).stem}.csv", write=True)
            sheet_name, row_count = _write_excel_csv(source, target, sheet_name=context.arguments.get("sheet"))
            artifact = _put_artifact(artifact_store, run.run_id, target, read_root)
            artifacts.append(artifact)
            files.append(
                {
                    "input": input_path,
                    "output": target.relative_to(result_dir).as_posix(),
                    "artifact_uri": artifact.uri,
                    "sheet": sheet_name,
                    "rows": row_count,
                }
            )
        finished_at = _now_ms()
        return TaskResult(
            run_id=run.run_id,
            task_id=run.spec.task_id,
            status=TaskStatus.COMPLETED,
            title=run.spec.title,
            summary=_file_conversion_summary(files),
            output=json.dumps({"files": files}, ensure_ascii=False),
            primary_artifact=artifacts[0] if artifacts else None,
            artifacts=artifacts,
            started_at=context.started_at,
            ended_at=finished_at,
            duration_ms=max(0, finished_at - context.started_at),
        )
    except (FileNotFoundError, ValueError, RuntimeError) as exc:
        return build_user_input_result(run, str(exc))


def _put_artifact(artifact_store: ArtifactStore, run_id: str, path: Path, workspace: Path) -> ArtifactRef:
    artifact = artifact_store.put(run_id, path.name, path, artifact_type="file", role="primary")
    artifact.filepath = _workspace_relative_path(path, workspace)
    return artifact


def _file_conversion_requests(context: CapabilityContext) -> list[str]:
    args = context.arguments
    raw_paths = args.get("input_paths") or args.get("files") or args.get("source_paths")
    if raw_paths is None:
        raw_path = args.get("input_path") or args.get("file_path") or args.get("source_path")
        raw_paths = [raw_path] if raw_path else []
    if isinstance(raw_paths, str):
        paths = [raw_paths.strip()] if raw_paths.strip() else []
    elif isinstance(raw_paths, list):
        paths = [str(path).strip() for path in raw_paths if str(path).strip()]
    else:
        paths = []
    if not paths:
        spec = context.run.spec
        text = f"{spec.title}\n{spec.instructions}"
        paths = [match.group(0).strip() for match in _WORKSPACE_SPREADSHEET_RE.finditer(text)]
    if not paths:
        raise ValueError("file_convert requires args.input_paths with workspace-relative Excel paths")
    return _dedupe_file_conversion_paths(paths)


def _dedupe_file_conversion_paths(paths: list[str]) -> list[str]:
    deduped: list[str] = []
    seen: set[str] = set()
    for path in paths:
        normalized = path.replace("\\", "/").strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        deduped.append(normalized)
    return deduped


def _workspace_file(workspace: Path, relative_path: str, *, write: bool = False) -> Path:
    root = workspace.resolve()
    target = (root / relative_path).resolve()
    if not target.is_relative_to(root):
        raise ValueError("file_convert paths must stay within the delegated workspace")
    if write:
        target.parent.mkdir(parents=True, exist_ok=True)
        return target
    if not target.is_file():
        raise FileNotFoundError(f"file_convert input not found: {relative_path}")
    return target


def _relative_dir(value: str) -> str:
    normalized = value.replace("\\", "/").strip().strip("/")
    if not normalized or normalized.startswith("../") or "/../" in normalized:
        raise ValueError("file_convert output_dir must be workspace-relative")
    return normalized


def _write_excel_csv(source: Path, target: Path, *, sheet_name: Any = None) -> tuple[str, int]:
    if source.suffix.lower() not in _EXCEL_EXTENSIONS:
        raise ValueError(f"file_convert only supports Excel .xlsx/.xlsm inputs, got: {source.name}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency is present in normal core installs.
        raise RuntimeError("file_convert requires openpyxl to convert Excel workbooks") from exc

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"file_convert could not read Excel workbook: {source.name}") from exc
    try:
        requested_sheet = str(sheet_name).strip() if sheet_name is not None else ""
        if requested_sheet:
            if requested_sheet not in workbook.sheetnames:
                raise ValueError(f"file_convert sheet not found: {requested_sheet}")
            worksheet = workbook[requested_sheet]
        else:
            worksheet = workbook.worksheets[0]
        row_count = 0
        with target.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
                row_count += 1
        return worksheet.title, row_count
    finally:
        workbook.close()


def _file_conversion_summary(files: list[dict[str, Any]]) -> str:
    lines = [f"Converted {len(files)} Excel file(s) to CSV:"]
    for item in files:
        lines.append(f"- {item['input']} -> {item['output']} ({item['rows']} rows from sheet {item['sheet']})")
    return "\n".join(lines)


def _workspace_relative_path(path: Path, workspace: Path) -> str:
    try:
        return path.resolve().relative_to(workspace.resolve()).as_posix()
    except ValueError:
        return ""


__all__ = ["BuiltinCapabilityProvider"]
