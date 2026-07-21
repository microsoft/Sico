# Copyright (c) 2026 Sico Authors
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from __future__ import annotations

import contextlib
import csv
import json
import os
import re
from pathlib import Path
from typing import Any

from ..artifact_store import ArtifactStore
from ..models import ArtifactRef, ErrorClass, TaskResult, TaskRun, TaskStatus
from ..naming import sanitize_dns_label
from ..results import build_user_input_result
from ..store import RunStore
from ..time_utils import now_ms as _now_ms
from ..tool_catalog import ECHO_TOOL_NAME, FILE_CONVERT_TOOL_NAME, RUN_COMMAND_TOOL_NAME
from ..workspace import workspace_layout
from .command_backend import CommandBackend, CommandMount, CommandResult, CommandSpec, truncate_stream


_WORKSPACE_MOUNT_NAME = "workspace"
_RESULT_MOUNT_NAME = "result"
_COMMAND_SUMMARY_HEAD = 80
_COMMAND_STDOUT_HEAD = 1000
_COMMAND_STDERR_HEAD = 500
_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
_WORKSPACE_SPREADSHEET_RE = re.compile(r"(?:attachments|download)/[^\n\r\"'`]+?\.(?:xlsx|xlsm)", re.IGNORECASE)


class ToolExecutor:
    def __init__(
        self,
        *,
        worker_id: str = "local-task-runtime",
        artifact_store: ArtifactStore,
        sandbox_backend: CommandBackend,
    ) -> None:
        if artifact_store is None:
            raise ValueError("artifact_store is required")
        if sandbox_backend is None:
            raise ValueError("sandbox_backend is required")
        self.worker_id = worker_id
        self.artifact_store = artifact_store
        self._sandbox_backend = sandbox_backend

    @property
    def sandbox_backend(self) -> CommandBackend:
        """The backend that decides *where* ``run_command`` runs.

        Supplied by the task-runtime factory so backend selection happens at
        construction time and execution fails early if the dependency is absent.
        """
        return self._sandbox_backend

    async def run(self, run: TaskRun, store: RunStore) -> TaskResult:
        token = await store.claim_run(run.run_id, self.worker_id)
        result = await self.run_tool(run)
        await store.write_result(run.run_id, result, token)
        return result

    async def run_tool(self, run: TaskRun) -> TaskResult:
        if run.spec.tool_name == ECHO_TOOL_NAME:
            return self._run_echo_tool(run)
        if run.spec.tool_name == FILE_CONVERT_TOOL_NAME:
            return self._run_file_convert_tool(run)
        if run.spec.tool_name == RUN_COMMAND_TOOL_NAME:
            return await self._run_command_tool(run)
        return build_user_input_result(run, f"Unsupported local tool payload: {run.spec.tool_name}")

    async def _run_command_tool(self, run: TaskRun) -> TaskResult:
        """Run a shell command via the selected :class:`CommandBackend`.

        This is the single ``run_command`` implementation: the executor builds a
        :class:`CommandSpec` and hands it to a per-run :class:`CommandSession`,
        so *where* the command runs (local/docker/k8s) is decided entirely by
        the backend. A sub-agent reaches the same code via its capability
        invoker rather than re-implementing command execution.
        """
        command = str(run.spec.args.get("command") or "").strip()
        if not command:
            return build_user_input_result(run, "run_command requires a non-empty args.command")
        started_at = _now_ms()
        spec = self._command_spec(run, command)
        session = self.sandbox_backend.open_session(
            pod_name=_command_pod_name(run),
            image=str(run.spec.args.get("image") or ""),
        )
        try:
            outcome = await session.run(spec)
        finally:
            await session.aclose()
        return _command_result_to_task_result(run, command, outcome, started_at)

    def _command_spec(self, run: TaskRun, command: str) -> CommandSpec:
        # Commands start in the shared workspace for natural relative reads.
        # The workspace mount is read-only in container backends; durable outputs
        # must be written under SICO_RESULT_DIR.
        result_dir = _run_dir(run) / "result"
        result_dir.mkdir(parents=True, exist_ok=True)
        result_str = str(result_dir)
        workspace_dir = _workspace_dir(run)
        workspace_dir.mkdir(parents=True, exist_ok=True)
        workspace_str = str(workspace_dir)
        mounts = [
            CommandMount(name=_RESULT_MOUNT_NAME, host_path=result_str, mount_path=result_str),
            CommandMount(name=_WORKSPACE_MOUNT_NAME, host_path=workspace_str, mount_path=workspace_str, read_only=True),
        ]
        return CommandSpec(
            argv=["sh", "-lc", command],
            cwd=workspace_str,
            env={**_command_env(run), "SICO_WORKSPACE_DIR": workspace_str, "SICO_RESULT_DIR": result_str},
            mounts=mounts,
            timeout_seconds=_command_timeout_seconds(run),
            metadata={
                "agent_instance_id": str(run.agent_instance_id),
                "user_label": sanitize_dns_label(run.username, max_len=63),
            },
        )

    def _run_echo_tool(self, run: TaskRun) -> TaskResult:
        now_ms = _now_ms()
        message = str(run.spec.args.get("message") or run.spec.instructions or run.spec.title)
        return TaskResult(
            run_id=run.run_id,
            task_id=run.spec.task_id,
            status=TaskStatus.COMPLETED,
            title=run.spec.title,
            summary=message,
            output=message,
            started_at=now_ms,
            ended_at=now_ms,
            duration_ms=0,
        )

    def _run_file_convert_tool(self, run: TaskRun) -> TaskResult:
        started_at = _now_ms()
        try:
            # Inputs are read from the shared (read-only) workspace; converted
            # outputs are written to this run's writable result directory.
            read_root = _workspace_dir(run)
            result_dir = _run_dir(run) / "result"
            result_dir.mkdir(parents=True, exist_ok=True)
            requests = _file_conversion_requests(run)
            target_format = str(run.spec.args.get("target_format") or "csv").lower().lstrip(".")
            if target_format != "csv":
                raise ValueError(f"file_convert only supports target_format=csv, got: {target_format}")
            output_dir = _relative_dir(str(run.spec.args.get("output_dir") or "output/csv"))
            artifacts: list[ArtifactRef] = []
            files: list[dict[str, Any]] = []
            for input_path in requests:
                source = _workspace_file(read_root, input_path)
                target = _workspace_file(result_dir, f"{output_dir}/{Path(input_path).stem}.csv", write=True)
                sheet_name, row_count = _write_excel_csv(source, target, sheet_name=run.spec.args.get("sheet"))
                artifact = self._put_artifact(run.run_id, target, read_root)
                artifacts.append(artifact)
                files.append(
                    {
                        "input": input_path,
                        "output": target.relative_to(result_dir).as_posix(),
                        "artifact_uri": artifact.uri,
                        "sheet": sheet_name,
                        "rows": row_count,
                    }
                )
            finished_at = _now_ms()
            summary = _file_conversion_summary(files)
            return TaskResult(
                run_id=run.run_id,
                task_id=run.spec.task_id,
                status=TaskStatus.COMPLETED,
                title=run.spec.title,
                summary=summary,
                output=json.dumps({"files": files}, ensure_ascii=False),
                primary_artifact=artifacts[0] if artifacts else None,
                artifacts=artifacts,
                started_at=started_at,
                ended_at=finished_at,
                duration_ms=max(0, finished_at - started_at),
            )
        except (FileNotFoundError, ValueError, RuntimeError) as exc:
            return build_user_input_result(run, str(exc))

    def _put_artifact(self, run_id: str, path: Path, workspace: Path) -> ArtifactRef:
        filepath = _workspace_relative_path(path, workspace)
        artifact = self.artifact_store.put(run_id, path.name, path, artifact_type="file", role="primary")
        artifact.filepath = filepath
        return artifact


def _file_conversion_requests(run: TaskRun) -> list[str]:
    raw_paths = run.spec.args.get("input_paths") or run.spec.args.get("files") or run.spec.args.get("source_paths")
    if raw_paths is None:
        raw_path = run.spec.args.get("input_path") or run.spec.args.get("file_path") or run.spec.args.get("source_path")
        raw_paths = [raw_path] if raw_path else []
    if isinstance(raw_paths, str):
        paths = [raw_paths.strip()] if raw_paths.strip() else []
    elif isinstance(raw_paths, list):
        paths = [str(path).strip() for path in raw_paths if str(path).strip()]
    else:
        paths = []
    if not paths:
        text = f"{run.spec.title}\n{run.spec.instructions}"
        paths = [match.group(0).strip() for match in _WORKSPACE_SPREADSHEET_RE.finditer(text)]
    if not paths:
        raise ValueError("file_convert requires args.input_paths with workspace-relative Excel paths")
    return _dedupe_file_conversion_paths(paths)


def _dedupe_file_conversion_paths(paths: list[str]) -> list[str]:
    deduped: list[str] = []
    seen: set[str] = set()
    for path in paths:
        normalized = path.replace("\\", "/").strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        deduped.append(normalized)
    return deduped


def _workspace_file(workspace: Path, relative_path: str, *, write: bool = False) -> Path:
    root = workspace.resolve()
    target = (root / relative_path).resolve()
    if not target.is_relative_to(root):
        raise ValueError("file_convert paths must stay within the delegated workspace")
    if write:
        target.parent.mkdir(parents=True, exist_ok=True)
        return target
    if not target.is_file():
        raise FileNotFoundError(f"file_convert input not found: {relative_path}")
    return target


def _relative_dir(value: str) -> str:
    normalized = value.replace("\\", "/").strip().strip("/")
    if not normalized or normalized.startswith("../") or "/../" in normalized:
        raise ValueError("file_convert output_dir must be workspace-relative")
    return normalized


def _write_excel_csv(source: Path, target: Path, *, sheet_name: Any = None) -> tuple[str, int]:
    if source.suffix.lower() not in _EXCEL_EXTENSIONS:
        raise ValueError(f"file_convert only supports Excel .xlsx/.xlsm inputs, got: {source.name}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency is present in normal core installs.
        raise RuntimeError("file_convert requires openpyxl to convert Excel workbooks") from exc

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"file_convert could not read Excel workbook: {source.name}") from exc
    try:
        requested_sheet = str(sheet_name).strip() if sheet_name is not None else ""
        if requested_sheet:
            if requested_sheet not in workbook.sheetnames:
                raise ValueError(f"file_convert sheet not found: {requested_sheet}")
            worksheet = workbook[requested_sheet]
        else:
            worksheet = workbook.worksheets[0]
        row_count = 0
        with target.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
                row_count += 1
        return worksheet.title, row_count
    finally:
        workbook.close()


def _file_conversion_summary(files: list[dict[str, Any]]) -> str:
    lines = [f"Converted {len(files)} Excel file(s) to CSV:"]
    for item in files:
        lines.append(f"- {item['input']} -> {item['output']} ({item['rows']} rows from sheet {item['sheet']})")
    return "\n".join(lines)


def _run_dir(run: TaskRun) -> Path:
    return _workspace_dir(run) / "results" / run.batch_id / run.run_id


def _workspace_dir(run: TaskRun) -> Path:
    return workspace_layout().workspace_path(
        run.agent_instance_id,
        run.username,
        conversation_id=run.parent_conversation_id,
    )


def _workspace_relative_path(path: Path, workspace: Path) -> str:
    try:
        return path.resolve().relative_to(workspace.resolve()).as_posix()
    except ValueError:
        return ""


def _command_env(run: TaskRun) -> dict[str, str]:
    """Per-run env overlay for ``run_command``.

    Only SICO_* identity vars are declared; the local backend merges these over
    ``os.environ`` and container backends forward them as ``-e``/pod env, so the
    set is intentionally small (no host environment leakage into containers).
    """
    return {
        "SICO_TASK_RUN_ID": run.run_id,
        "SICO_AGENT_INSTANCE_ID": str(run.agent_instance_id),
        "SICO_PROJECT_ID": str(run.project_id),
        "SICO_APP_NAME": _sico_app_name(),
    }


def _sico_app_name() -> str:
    return os.getenv("SICO_APP_NAME", "sico").strip() or "sico"


def _command_timeout_seconds(run: TaskRun) -> int:
    requested = run.spec.args.get("timeout")
    if requested:
        with contextlib.suppress(TypeError, ValueError):
            return max(0, int(requested))
    return run.execution_policy.timeout_seconds


def _command_pod_name(run: TaskRun) -> str:
    return sanitize_dns_label(f"task-{run.run_id}", max_len=63)


def _command_summary(command: str, outcome: CommandResult, status: TaskStatus) -> str:
    head = command if len(command) <= _COMMAND_SUMMARY_HEAD else command[: _COMMAND_SUMMARY_HEAD - 3] + "..."
    if status == TaskStatus.COMPLETED:
        lines = [f"`{head}` finished with exit code {outcome.return_code}"]
    elif status == TaskStatus.TIMED_OUT:
        lines = [f"`{head}` timed out"]
    elif outcome.system_error:
        lines = [f"`{head}` failed to run: {outcome.system_error}"]
    else:
        lines = [f"`{head}` failed with exit code {outcome.return_code}"]
    # Fold the command's output into the summary so the caller sees it via the
    # batch digest (which only carries ``summary``), not just the exit code.
    stdout = truncate_stream(outcome.stdout, _COMMAND_STDOUT_HEAD)
    if stdout:
        lines.append(f"stdout:\n{stdout}")
    if status != TaskStatus.COMPLETED:
        stderr = truncate_stream(outcome.stderr, _COMMAND_STDERR_HEAD)
        if stderr:
            lines.append(f"stderr:\n{stderr}")
    return "\n".join(lines)


def _command_result_to_task_result(
    run: TaskRun,
    command: str,
    outcome: CommandResult,
    started_at: int,
) -> TaskResult:
    finished_at = _now_ms()
    timed_out = outcome.return_code == -1 and "timed out" in outcome.system_error.lower()
    if timed_out:
        status = TaskStatus.TIMED_OUT
    elif outcome.system_error or outcome.return_code != 0:
        status = TaskStatus.FAILED
    else:
        status = TaskStatus.COMPLETED
    error_class: ErrorClass | None = None
    error_message = ""
    if status == TaskStatus.TIMED_OUT:
        error_class = ErrorClass.TIMEOUT
        error_message = outcome.system_error or f"command timed out after {_command_timeout_seconds(run)}s"
    elif status == TaskStatus.FAILED:
        error_class = ErrorClass.TRANSIENT if outcome.system_error else ErrorClass.SKILL_RUNTIME
        error_message = outcome.system_error or outcome.stderr or f"command exited with {outcome.return_code}"
    return TaskResult(
        run_id=run.run_id,
        task_id=run.spec.task_id,
        status=status,
        title=run.spec.title,
        summary=_command_summary(command, outcome, status),
        output=outcome.stdout,
        error_class=error_class,
        error_message=error_message,
        sandbox=run.sandbox,
        started_at=started_at,
        ended_at=finished_at,
        duration_ms=max(0, finished_at - started_at),
    )
