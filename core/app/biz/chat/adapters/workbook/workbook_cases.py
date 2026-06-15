# Copyright (c) 2026 Sico Authors
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from __future__ import annotations

import contextlib
import csv
import re
from collections.abc import Iterator
from pathlib import Path
from typing import Any

WORKBOOK_HEADER_LIMIT = 12
EXCEL_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
CSV_WORKBOOK_SUFFIXES = {".csv"}
SUPPORTED_WORKBOOK_SUFFIXES = EXCEL_WORKBOOK_SUFFIXES | CSV_WORKBOOK_SUFFIXES

_CSV_ENCODING_CANDIDATES = ("utf-8-sig", "gbk")
_CSV_SNIFF_DELIMITERS = ",;\t|"
_CSV_SNIFF_SAMPLE_BYTES = 4096
_CSV_MAX_BYTES = 8 * 1024 * 1024  # 8 MiB safety cap for in-memory CSV parsing.

_CASE_ID_HEADERS = {
    "id",
    "case id",
    "case_id",
    "test case id",
    "testcase id",
    "tc id",
    "用例id",
    "用例编号",
}
_TITLE_HEADERS = {
    "title",
    "case title",
    "test case",
    "testcase",
    "test title",
    "name",
    "标题",
    "用例标题",
}
_CASE_ID_RE = re.compile(r"(?<![A-Z0-9])[A-Z][A-Z0-9]{1,20}-\d+(?![A-Z0-9])", re.IGNORECASE)


def workbook_manifest(path: Path) -> dict[str, Any] | None:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_WORKBOOK_SUFFIXES:
        return None
    workbook = _load_workbook_any(path)
    if workbook is None:
        return None
    try:
        with contextlib.suppress(Exception):
            sheets = [
                _worksheet_manifest(worksheet.title, worksheet.iter_rows(values_only=True))
                for worksheet in workbook.worksheets
            ]
            return build_workbook_manifest(sheets)
    finally:
        workbook.close()
    return None


def workbook_case_sources(path: Path, manifest: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_WORKBOOK_SUFFIXES:
        return []
    workbook = _load_workbook_any(path)
    if workbook is None:
        return []
    try:
        with contextlib.suppress(Exception):
            manifest_by_sheet = _manifest_by_sheet(manifest)
            sources: list[dict[str, Any]] = []
            for worksheet in workbook.worksheets:
                # Materialize rows once to avoid re-iterating a read-only worksheet.
                rows = list(worksheet.iter_rows(values_only=True))
                sheet_manifest = manifest_by_sheet.get(_normalize_sheet_name(worksheet.title))
                if sheet_manifest is None:
                    sheet_manifest = _worksheet_manifest(worksheet.title, iter(rows))
                if sheet_manifest.get("kind") not in {"data", "master"} or int(sheet_manifest.get("data_rows") or 0) <= 0:
                    continue
                cases = _cases_from_rows(worksheet.title, sheet_manifest, iter(rows))
                if not cases:
                    continue
                sources.append(
                    {
                        "sheet_name": worksheet.title,
                        "kind": sheet_manifest.get("kind") or "data",
                        "headers": sheet_manifest.get("headers") or [],
                        "case_count": len(cases),
                        "cases": cases,
                    }
                )
            return sources
    finally:
        workbook.close()
    return []


def extract_workbook_cases(
    path: Path,
    *,
    sheet_name: str = "",
    row_start: int | None = None,
    row_end: int | None = None,
    case_ids: list[str] | None = None,
    max_cases: int = 500,
) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_WORKBOOK_SUFFIXES:
        return {
            "error_message": "file is not a supported workbook (.xlsx/.xlsm/.csv)",
            "cases": [],
        }
    if suffix in EXCEL_WORKBOOK_SUFFIXES:
        try:
            from openpyxl import load_workbook
        except Exception:
            return {"error_message": "openpyxl is required to extract workbook cases", "cases": []}
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return {"error_message": str(exc), "cases": []}
    else:
        try:
            workbook = _load_csv_workbook(path)
        except Exception as exc:
            return {"error_message": str(exc), "cases": []}
        if workbook is None:
            return {"error_message": "csv file could not be decoded", "cases": []}

    try:
        return _extract_workbook_cases_from_loaded_workbook(
            workbook,
            path,
            sheet_name=sheet_name,
            row_start=row_start,
            row_end=row_end,
            case_ids=case_ids or [],
            max_cases=max_cases,
        )
    finally:
        workbook.close()


def _load_workbook_any(path: Path) -> Any | None:
    """Open path as a workbook-like object (xlsx via openpyxl, csv via adapter)."""

    suffix = path.suffix.lower()
    if suffix in EXCEL_WORKBOOK_SUFFIXES:
        try:
            from openpyxl import load_workbook
        except Exception:
            return None
        try:
            return load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return None
    if suffix in CSV_WORKBOOK_SUFFIXES:
        with contextlib.suppress(Exception):
            return _load_csv_workbook(path)
        return None
    return None


class _CsvWorksheet:
    """Minimal worksheet adapter exposing the surface used by ``workbook_cases``.

    Implements only ``title`` and ``iter_rows(values_only=True)`` so that the
    Excel-oriented manifest / case extraction pipeline can treat a CSV file as
    a single-sheet workbook without branching.
    """

    def __init__(self, title: str, rows: list[list[str]]) -> None:
        self.title = title
        self._rows = rows

    def iter_rows(self, *, values_only: bool = True) -> Iterator[list[str]]:
        # ``values_only`` is accepted for parity with openpyxl; CSV rows are
        # always plain string sequences, so the flag has no effect here.
        _ = values_only
        return iter(self._rows)


class _CsvWorkbook:
    """Single-sheet workbook adapter wrapping a ``_CsvWorksheet``."""

    def __init__(self, worksheet: _CsvWorksheet) -> None:
        self.worksheets: list[_CsvWorksheet] = [worksheet]

    def close(self) -> None:
        return None


def _load_csv_workbook(path: Path) -> _CsvWorkbook | None:
    rows = _read_csv_rows(path)
    if rows is None:
        return None
    title = _csv_sheet_title(path)
    return _CsvWorkbook(_CsvWorksheet(title, rows))


def _csv_sheet_title(path: Path) -> str:
    stem = (path.stem or "").strip()
    return stem or "csv"


def _read_csv_rows(path: Path) -> list[list[str]] | None:
    try:
        if path.stat().st_size > _CSV_MAX_BYTES:
            return None
    except OSError:
        return None
    for encoding in _CSV_ENCODING_CANDIDATES:
        try:
            with path.open("r", encoding=encoding, newline="") as fp:
                sample = fp.read(_CSV_SNIFF_SAMPLE_BYTES)
                fp.seek(0)
                dialect = _sniff_csv_dialect(sample)
                reader = csv.reader(fp, dialect)
                return [list(row) for row in reader]
        except UnicodeDecodeError:
            continue
        except OSError:
            return None
    # Last resort: re-read as UTF-8 replacing unencodable bytes so that
    # files with a handful of stray non-UTF-8 characters (e.g. Windows-1252
    # punctuation like × 0xD7) are still usable rather than rejected.
    try:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as fp:
            sample = fp.read(_CSV_SNIFF_SAMPLE_BYTES)
            fp.seek(0)
            dialect = _sniff_csv_dialect(sample)
            reader = csv.reader(fp, dialect)
            return [list(row) for row in reader]
    except OSError:
        pass
    return None


def _sniff_csv_dialect(sample: str) -> Any:
    if not sample.strip():
        return csv.excel
    try:
        sniffed = csv.Sniffer().sniff(sample, delimiters=_CSV_SNIFF_DELIMITERS)
        # The sniffer may misdetect doublequote as False when the sample
        # contains ``""""``-style escaped quotes (common in LLM-generated or
        # Excel-exported CSVs with multi-line cells).  When doublequote is
        # wrong the reader treats ``""`` as a field terminator, splitting
        # multi-line quoted cells into separate rows.  Always enforce the
        # standard quoting behaviour so multi-line fields parse correctly.
        sniffed.doublequote = True
        return sniffed
    except csv.Error:
        return csv.excel


def _extract_workbook_cases_from_loaded_workbook(
    workbook: Any,
    path: Path,
    *,
    sheet_name: str,
    row_start: int | None,
    row_end: int | None,
    case_ids: list[str],
    max_cases: int,
) -> dict[str, Any]:
    # Materialize all worksheet rows up-front so that read-only worksheets
    # are iterated exactly once.  Re-iterating a ReadOnlyWorksheet can
    # indeterministically lose trailing rows depending on the xlsx structure
    # and openpyxl's internal XML stream handling.
    cached_rows: dict[str, list[tuple[Any, ...]]] = {}
    for worksheet in workbook.worksheets:
        cached_rows[worksheet.title] = list(worksheet.iter_rows(values_only=True))

    manifest = build_workbook_manifest(
        [_worksheet_manifest(ws.title, iter(cached_rows[ws.title])) for ws in workbook.worksheets]
    )
    available_sheets = _available_sheet_summaries(manifest)
    selected = _select_sheets(workbook.worksheets, manifest, sheet_name, allow_multi_sheet_case_ids=bool(case_ids))
    if isinstance(selected, dict):
        return {**selected, "cases": [], "available_sheets": available_sheets}

    cases = _cases_from_sheets_cached(selected, manifest, cached_rows)
    cases = filter_cases(cases, row_start=row_start, row_end=row_end, case_ids=case_ids)
    cases = dedupe_cases_by_case_id(cases, case_ids)
    missing_case_ids = _missing_case_ids(cases, case_ids)
    if missing_case_ids:
        return {
            "error_message": f"no cases matched case_ids: {', '.join(missing_case_ids)}",
            "cases": [],
            "available_sheets": available_sheets,
        }
    if not cases and _has_selected_scope(sheet_name, row_start, row_end, case_ids):
        return {
            "error_message": "no cases matched the selected scope",
            "cases": [],
            "available_sheets": available_sheets,
        }
    if len(cases) > max_cases:
        return {
            "error_message": (
                f"selected {len(cases)} cases exceeds max_cases={max_cases}; choose a narrower sheet, range, or case list"
            ),
            "cases": [],
            "available_sheets": available_sheets,
        }
    return {
        "error_message": "",
        "source_type": "workbook",
        "file_path": path.name,
        "selected_sheets": _selected_sheet_names(selected, cases),
        "available_sheets": available_sheets,
        "case_count": len(cases),
        "cases": cases,
    }


def _cases_from_sheets(selected: list[Any], manifest: dict[str, Any]) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    manifest_by_sheet = _manifest_by_sheet(manifest)
    for worksheet in selected:
        sheet_manifest = manifest_by_sheet.get(_normalize_sheet_name(worksheet.title)) or {}
        cases.extend(_cases_from_worksheet(worksheet, sheet_manifest))
    return cases


def _cases_from_sheets_cached(
    selected: list[Any],
    manifest: dict[str, Any],
    cached_rows: dict[str, list[tuple[Any, ...]]],
) -> list[dict[str, Any]]:
    """Like ``_cases_from_sheets`` but reads from pre-materialized *cached_rows*."""
    cases: list[dict[str, Any]] = []
    manifest_by_sheet = _manifest_by_sheet(manifest)
    for worksheet in selected:
        sheet_manifest = manifest_by_sheet.get(_normalize_sheet_name(worksheet.title)) or {}
        rows = cached_rows.get(worksheet.title, [])
        cases.extend(_cases_from_rows(worksheet.title, sheet_manifest, iter(rows)))
    return cases


def dedupe_cases_by_case_id(cases: list[dict[str, Any]], case_ids: list[str] | None) -> list[dict[str, Any]]:
    wanted_order = [_normalize_case_id(case_id) for case_id in case_ids or [] if str(case_id).strip()]
    if not wanted_order:
        return cases
    selected_by_id: dict[str, dict[str, Any]] = {}
    for case in cases:
        case_id = _normalize_case_id(str(case.get("case_id") or ""))
        if not case_id:
            continue
        existing = selected_by_id.get(case_id)
        if existing is None or _case_source_rank(case) < _case_source_rank(existing):
            selected_by_id[case_id] = case
    ordered: list[dict[str, Any]] = []
    seen: set[str] = set()
    for case_id in wanted_order:
        if case_id in seen:
            continue
        seen.add(case_id)
        selected = selected_by_id.get(case_id)
        if selected is not None:
            ordered.append(selected)
    return ordered


def _case_source_rank(case: dict[str, Any]) -> int:
    kind = str(case.get("sheet_kind") or "").lower()
    if kind == "data":
        return 0
    if kind == "master":
        return 1
    return 2


def _has_selected_scope(
    sheet_name: str,
    row_start: int | None,
    row_end: int | None,
    case_ids: list[str],
) -> bool:
    return bool(sheet_name or row_start is not None or row_end is not None or case_ids)


def filter_cases(
    cases: list[dict[str, Any]],
    *,
    row_start: int | None = None,
    row_end: int | None = None,
    case_ids: list[str] | None = None,
) -> list[dict[str, Any]]:
    wanted_ids = {_normalize_case_id(case_id) for case_id in case_ids or [] if str(case_id).strip()}
    selected: list[dict[str, Any]] = []
    for case in cases:
        data_row_index = int(case.get("data_row_index") or 0)
        if row_start is not None and data_row_index < row_start:
            continue
        if row_end is not None and data_row_index > row_end:
            continue
        if wanted_ids and _normalize_case_id(str(case.get("case_id") or "")) not in wanted_ids:
            continue
        selected.append(case)
    return selected


def _missing_case_ids(cases: list[dict[str, Any]], case_ids: list[str]) -> list[str]:
    wanted_ids = [_normalize_case_id(case_id) for case_id in case_ids if str(case_id).strip()]
    if not wanted_ids:
        return []
    found_ids = {_normalize_case_id(str(case.get("case_id") or "")) for case in cases}
    return [case_id for case_id in wanted_ids if case_id not in found_ids]


def build_workbook_manifest(sheets: list[dict[str, Any]]) -> dict[str, Any]:
    data_sheets = [sheet for sheet in sheets if sheet["kind"] == "data"]
    master_sheets = [sheet for sheet in sheets if sheet["kind"] == "master"]
    summary_sheets = [sheet for sheet in sheets if sheet["kind"] == "summary"]
    executable_sheets = [sheet for sheet in sheets if sheet["kind"] in {"data", "master"} and sheet["data_rows"] > 0]
    source_data_rows = _sum_sheet_data_rows(data_sheets)
    master_data_rows = _sum_sheet_data_rows(master_sheets)
    summary_data_rows = _sum_sheet_data_rows(summary_sheets)
    total_data_rows = _sum_sheet_data_rows(sheets)
    runnable_data_rows = source_data_rows if source_data_rows > 0 else master_data_rows
    return {
        "type": "workbook",
        "sheet_count": len(sheets),
        "total_data_rows": total_data_rows,
        "runnable_data_rows": runnable_data_rows,
        "source_data_rows": source_data_rows,
        "master_data_rows": master_data_rows,
        "summary_data_rows": summary_data_rows,
        "data_sheet_count": len(data_sheets),
        "executable_sheet_count": len(executable_sheets),
        "multiple_data_sheets": len(data_sheets) > 1,
        "requires_scope_selection": len(executable_sheets) > 1,
        "contains_master_sheet": any(sheet["kind"] == "master" for sheet in sheets),
        "sheets": sheets,
        "scope_confirmation_hint": _workbook_scope_confirmation_hint(executable_sheets),
    }


def _worksheet_manifest(name: str, rows: Any) -> dict[str, Any]:
    non_empty_rows = 0
    headers: list[str] = []
    for row in rows:
        values = [_stringify_cell(cell) for cell in row]
        non_empty_values = [value for value in values if value]
        if not non_empty_values:
            continue
        non_empty_rows += 1
        if not headers:
            headers = non_empty_values[:WORKBOOK_HEADER_LIMIT]
    data_rows = max(0, non_empty_rows - 1)
    return {
        "name": name,
        "kind": classify_sheet(name, headers, data_rows),
        "non_empty_rows": non_empty_rows,
        "data_rows": data_rows,
        "headers": headers,
    }


def classify_sheet(name: str, headers: list[str], data_rows: int) -> str:
    if data_rows <= 0:
        return "empty"
    lowered_name = name.strip().lower()
    lowered_headers = {header.strip().lower() for header in headers if header.strip()}
    if lowered_name in {"summary", "readme", "overview"} or (lowered_headers and lowered_headers <= {"metric", "value"}):
        return "summary"
    if lowered_name in {"master", "all", "combined"} or {"source_file", "source_row"}.issubset(lowered_headers):
        return "master"
    return "data"


def _cases_from_worksheet(worksheet: Any, sheet_manifest: dict[str, Any]) -> list[dict[str, Any]]:
    return _cases_from_rows(worksheet.title, sheet_manifest, worksheet.iter_rows(values_only=True))


def _cases_from_rows(sheet_title: str, sheet_manifest: dict[str, Any], rows: Any) -> list[dict[str, Any]]:
    headers: list[str] = []
    cases: list[dict[str, Any]] = []
    data_row_index = 0
    for excel_row, row in enumerate(rows, start=1):
        values = [_stringify_cell(cell) for cell in row]
        if not any(values):
            continue
        if not headers:
            headers = _normalized_headers(values)
            continue
        data_row_index += 1
        row_values = _row_values(headers, values)
        case = _case_from_values(sheet_title, sheet_manifest, row_values, data_row_index, excel_row)
        cases.append(case)
    return cases


def _case_from_values(
    sheet_name: str,
    sheet_manifest: dict[str, Any],
    values: dict[str, str],
    data_row_index: int,
    excel_row: int,
) -> dict[str, Any]:
    case_id = _first_header_value(values, _CASE_ID_HEADERS) or _first_case_id_value(values)
    title = _first_header_value(values, _TITLE_HEADERS) or case_id or f"{sheet_name} row {data_row_index}"
    instructions = _case_instructions(sheet_name, case_id, title, data_row_index, excel_row, values)
    return {
        "sheet_name": sheet_name,
        "sheet_kind": sheet_manifest.get("kind") or "data",
        "data_row_index": data_row_index,
        "source_row": excel_row,
        "case_id": case_id,
        "title": title,
        "instructions": instructions,
        "values": values,
    }


def _case_instructions(
    sheet_name: str,
    case_id: str,
    title: str,
    data_row_index: int,
    excel_row: int,
    values: dict[str, str],
) -> str:
    lines = [
        f"Workbook sheet: {sheet_name}",
        f"Sheet data row: {data_row_index}",
        f"Excel row: {excel_row}",
    ]
    if case_id:
        lines.append(f"Case ID: {case_id}")
    if title:
        lines.append(f"Title: {title}")
    lines.append("Case fields:")
    for header, value in values.items():
        if value:
            lines.append(f"- {header}: {value}")
    return "\n".join(lines)


def _select_sheets(
    worksheets: list[Any],
    manifest: dict[str, Any],
    sheet_name: str,
    *,
    allow_multi_sheet_case_ids: bool = False,
) -> list[Any] | dict[str, str]:
    runnable_names = {
        _normalize_sheet_name(sheet["name"])
        for sheet in manifest.get("sheets") or []
        if sheet.get("kind") in {"data", "master"} and int(sheet.get("data_rows") or 0) > 0
    }
    if sheet_name:
        normalized = _normalize_sheet_name(sheet_name)
        matches = [worksheet for worksheet in worksheets if _normalize_sheet_name(worksheet.title) == normalized]
        if not matches:
            result: list[Any] | dict[str, str] = {"error_message": f"sheet not found: {sheet_name}"}
        elif _normalize_sheet_name(matches[0].title) not in runnable_names:
            result = {"error_message": f"sheet is not executable: {matches[0].title}"}
        else:
            result = matches
    else:
        runnable = [worksheet for worksheet in worksheets if _normalize_sheet_name(worksheet.title) in runnable_names]
        if len(runnable) == 1 or (allow_multi_sheet_case_ids and runnable):
            result = runnable
        elif not runnable:
            result = {"error_message": "no executable sheets found"}
        else:
            result = {"error_message": "multiple executable sheets found; provide sheet_name, row range, or case_ids"}
    return result


def _selected_sheet_names(selected: list[Any], cases: list[dict[str, Any]]) -> list[str]:
    if not cases:
        return [worksheet.title for worksheet in selected]
    names: list[str] = []
    for case in cases:
        name = str(case.get("sheet_name") or "")
        if name and name not in names:
            names.append(name)
    return names or [worksheet.title for worksheet in selected]


def _manifest_by_sheet(manifest: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    if not isinstance(manifest, dict):
        return {}
    result: dict[str, dict[str, Any]] = {}
    for sheet in manifest.get("sheets") or []:
        if isinstance(sheet, dict):
            result[_normalize_sheet_name(str(sheet.get("name") or ""))] = sheet
    return result


def _available_sheet_summaries(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "name": str(sheet.get("name") or ""),
            "kind": str(sheet.get("kind") or ""),
            "data_rows": int(sheet.get("data_rows") or 0),
        }
        for sheet in manifest.get("sheets") or []
        if isinstance(sheet, dict)
    ]


def _normalized_headers(values: list[str]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        header = value.strip() or f"column_{index}"
        count = seen.get(header, 0) + 1
        seen[header] = count
        headers.append(header if count == 1 else f"{header}_{count}")
    return headers


def _row_values(headers: list[str], values: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    width = max(len(headers), len(values))
    for index in range(width):
        header = headers[index] if index < len(headers) else f"column_{index + 1}"
        value = values[index] if index < len(values) else ""
        result[header] = value
    return result


def _first_header_value(values: dict[str, str], wanted_headers: set[str]) -> str:
    for header, value in values.items():
        if _normalize_header(header) in wanted_headers and value:
            return value
    return ""


def _first_case_id_value(values: dict[str, str]) -> str:
    for value in values.values():
        match = _CASE_ID_RE.search(value)
        if match:
            return match.group(0).upper()
    return ""


def _normalize_header(header: str) -> str:
    # Strip leading/trailing non-alphanumeric characters (e.g. stray '?' from
    # corrupted BOM or export artefacts) before normalizing.
    cleaned = re.sub(r"^[^\w]+|[^\w]+$", "", header.strip(), flags=re.UNICODE)
    return re.sub(r"\s+", " ", cleaned.lower().replace("_", " "))


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", name.strip().lower())


def _normalize_case_id(case_id: str) -> str:
    return case_id.strip().upper()


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _sum_sheet_data_rows(sheets: list[dict[str, Any]]) -> int:
    return sum(int(sheet.get("data_rows") or 0) for sheet in sheets)


def _workbook_scope_confirmation_hint(executable_sheets: list[dict[str, Any]]) -> str:
    if len(executable_sheets) <= 1:
        return ""
    return (
        "Workbook has multiple runnable sheets; ask the user which sheet(s), row range, or case IDs to execute before "
        "delegation."
    )
