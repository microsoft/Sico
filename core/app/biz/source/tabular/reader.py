"""Pure local-file reader for supported tabular formats."""

from __future__ import annotations

import csv
import hashlib
import json
from collections.abc import Iterable, Iterator
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from ..errors import SourceError
from ..models import TabularDocument, TabularRow, TabularSheet, normalize_header, table_id_for_source

SUPPORTED_TABULAR_SUFFIXES = frozenset((".xlsx", ".xlsm", ".csv", ".tsv", ".jsonl"))
_CSV_ENCODINGS = ("utf-8-sig", "gbk")
_CSV_MAX_BYTES = 8 * 1024 * 1024
_TABULAR_MAX_BYTES = 32 * 1024 * 1024
_TABULAR_MAX_SOURCE_ROWS = 50_000
_TABULAR_MAX_COLUMNS = 500
_TABULAR_MAX_DECODED_CELLS = 250_000
_TABULAR_MAX_CELL_CHARS = 100_000
_TABULAR_MAX_DISPLAY_CHARS = 32 * 1024 * 1024


@dataclass(slots=True)
class _DecodedBudget:
    cells: int = 0
    display_chars: int = 0

    def take(self, values: Iterable[Any], *, label: str) -> None:
        materialized = tuple(values)
        if len(materialized) > _TABULAR_MAX_COLUMNS:
            raise SourceError(
                f"tabular table {label!r} exceeds {_TABULAR_MAX_COLUMNS} columns",
                code="tabular_column_limit",
                details={"table": label, "columns": len(materialized), "max_columns": _TABULAR_MAX_COLUMNS},
            )
        rendered = tuple(display_value(value) for value in materialized)
        if oversized := next((len(value) for value in rendered if len(value) > _TABULAR_MAX_CELL_CHARS), None):
            raise SourceError(
                f"tabular table {label!r} contains an oversized cell",
                code="tabular_cell_limit",
                details={"table": label, "cell_chars": oversized, "max_cell_chars": _TABULAR_MAX_CELL_CHARS},
            )
        self.cells += len(materialized)
        self.display_chars += sum(len(value) for value in rendered)
        if self.cells > _TABULAR_MAX_DECODED_CELLS or self.display_chars > _TABULAR_MAX_DISPLAY_CHARS:
            raise SourceError(
                f"tabular source exceeds decoded-data limits at {label!r}",
                code="tabular_decoded_size_limit",
                details={
                    "table": label,
                    "decoded_cells": self.cells,
                    "max_decoded_cells": _TABULAR_MAX_DECODED_CELLS,
                    "display_chars": self.display_chars,
                    "max_display_chars": _TABULAR_MAX_DISPLAY_CHARS,
                },
            )


class TabularReader:
    def read(
        self,
        path: Path,
        source_ref: str,
        source_id: str,
        *,
        sheet_names: tuple[str, ...] = (),
    ) -> TabularDocument:
        _validate_source_size(path)
        budget = _DecodedBudget()
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_TABULAR_SUFFIXES:
            raise SourceError(
                f"unsupported tabular source format: {suffix or '<none>'}",
                code="tabular_unsupported_format",
                details={"source_ref": source_ref, "supported": sorted(SUPPORTED_TABULAR_SUFFIXES)},
            )
        if suffix in {".xlsx", ".xlsm"}:
            sheets = _excel_sheets(path, source_id, sheet_names, source_ref, budget)
        elif suffix in {".csv", ".tsv"}:
            sheets = (_delimited_sheet(path, suffix, source_id, budget),)
        else:
            sheets = _jsonl_sheets(path, source_id, source_ref, sheet_names, budget)
        if not any(sheet.rows for sheet in sheets):
            raise SourceError("no tabular rows matched the source", code="tabular_no_rows", details={"source_ref": source_ref})
        return TabularDocument(
            source_ref=source_ref,
            source_id=source_id,
            file_name=path.name,
            format=suffix.lstrip("."),
            sheets=sheets,
            materialized_ref=source_ref,
        )


def content_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _validate_source_size(path: Path) -> None:
    try:
        size = path.stat().st_size
    except OSError as exc:
        raise SourceError(f"failed to inspect tabular source: {exc}") from exc
    if size > _TABULAR_MAX_BYTES:
        raise SourceError(
            f"tabular source exceeds {_TABULAR_MAX_BYTES} bytes",
            code="tabular_source_too_large",
            details={"size_bytes": size, "max_bytes": _TABULAR_MAX_BYTES},
        )


def _excel_sheets(
    path: Path,
    source_id: str,
    sheet_names: tuple[str, ...],
    source_ref: str,
    budget: _DecodedBudget,
) -> tuple[TabularSheet, ...]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise SourceError(f"failed to open workbook: {exc}") from exc
    try:
        wanted = {name.strip().casefold() for name in sheet_names if name.strip()}
        worksheets = tuple(
            worksheet for worksheet in workbook.worksheets if not wanted or worksheet.title.strip().casefold() in wanted
        )
        found = {worksheet.title.strip().casefold() for worksheet in worksheets}
        if missing := sorted(wanted - found):
            raise SourceError(
                f"sheet(s) not found: {missing}",
                code="tabular_sheet_not_found",
                details={
                    "source_ref": source_ref,
                    "missing": missing,
                    "available": [worksheet.title for worksheet in workbook.worksheets],
                },
            )
        sheets: list[TabularSheet] = []
        source_row_count = 0
        for index, worksheet in enumerate(worksheets):
            def worksheet_rows() -> Iterator[tuple[Any, ...]]:
                nonlocal source_row_count
                for row in worksheet.iter_rows(values_only=True):
                    source_row_count += 1
                    if source_row_count > _TABULAR_MAX_SOURCE_ROWS:
                        raise SourceError(
                            f"tabular source exceeds {_TABULAR_MAX_SOURCE_ROWS} source rows",
                            code="tabular_source_row_limit",
                            details={"table": path.name, "max_rows": _TABULAR_MAX_SOURCE_ROWS},
                        )
                    yield row

            sheets.append(_sheet(source_id, worksheet.title, index, worksheet_rows(), budget))
        return tuple(sheets)
    finally:
        workbook.close()


def _delimited_sheet(path: Path, suffix: str, source_id: str, budget: _DecodedBudget) -> TabularSheet:
    try:
        size = path.stat().st_size
    except OSError as exc:
        raise SourceError(f"failed to inspect tabular source: {exc}") from exc
    if size > _CSV_MAX_BYTES:
        raise SourceError(
            f"tabular source exceeds {_CSV_MAX_BYTES} bytes",
            code="tabular_source_too_large",
            details={"size_bytes": size, "max_bytes": _CSV_MAX_BYTES},
        )
    delimiter = "\t" if suffix == ".tsv" else None
    for encoding in _CSV_ENCODINGS:
        checkpoint = budget.cells, budget.display_chars
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                sample = stream.read(4096)
                stream.seek(0)
                dialect = csv.excel_tab if delimiter == "\t" else _csv_dialect(sample)
                return _sheet(source_id, "table", 0, _bounded_rows(csv.reader(stream, dialect), path.name), budget)
        except UnicodeDecodeError:
            budget.cells, budget.display_chars = checkpoint
            continue
        except OSError as exc:
            raise SourceError(f"failed to read tabular source: {exc}") from exc
    try:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as stream:
            sample = stream.read(4096)
            stream.seek(0)
            dialect = csv.excel_tab if delimiter == "\t" else _csv_dialect(sample)
            return _sheet(source_id, "table", 0, _bounded_rows(csv.reader(stream, dialect), path.name), budget)
    except OSError as exc:
        raise SourceError(f"failed to read tabular source: {exc}") from exc


def _jsonl_sheets(
    path: Path,
    source_id: str,
    source_ref: str,
    sheet_names: tuple[str, ...],
    budget: _DecodedBudget,
) -> tuple[TabularSheet, ...]:
    records: dict[str, list[tuple[int, int, dict[str, Any]]]] = {}
    headers: dict[str, list[str]] = {}
    fallback_indexes: dict[str, int] = {}
    wanted = {name.strip().casefold() for name in sheet_names if name.strip()}
    try:
        with path.open("r", encoding="utf-8") as stream:
            for line_number, line in enumerate(stream, start=1):
                if line_number > _TABULAR_MAX_SOURCE_ROWS:
                    raise SourceError(
                        f"tabular source exceeds {_TABULAR_MAX_SOURCE_ROWS} source rows",
                        code="tabular_source_row_limit",
                        details={"table": path.name, "max_rows": _TABULAR_MAX_SOURCE_ROWS},
                    )
                if not line.strip():
                    continue
                value = json.loads(line)
                if not isinstance(value, dict):
                    raise ValueError(f"line {line_number} is not an object")
                sheet_name = str(value.get("sheet_name") or "table")
                if wanted and sheet_name.strip().casefold() not in wanted:
                    continue
                fallback_indexes[sheet_name] = fallback_indexes.get(sheet_name, 0) + 1
                data_index = int(value.get("data_row_index") or fallback_indexes[sheet_name])
                source_row = int(value.get("source_row") or data_index + 1)
                raw_values = value.get("values") if isinstance(value.get("values"), dict) else value
                clean_values = {
                    str(key): item
                    for key, item in raw_values.items()
                    if key not in {"sheet_name", "sheet_kind", "data_row_index", "source_row", "instructions"}
                }
                budget.take(clean_values.values(), label=sheet_name)
                sheet_headers = headers.setdefault(sheet_name, [])
                for header in clean_values:
                    if header not in sheet_headers:
                        sheet_headers.append(header)
                records.setdefault(sheet_name, []).append((source_row, data_index, clean_values))
    except SourceError:
        raise
    except (OSError, json.JSONDecodeError, TypeError, ValueError) as exc:
        raise SourceError(
            f"tabular source contains invalid row metadata: {exc}",
            code="tabular_invalid_rows",
            details={"source_ref": source_ref},
        ) from exc
    if missing := sorted(wanted - {name.strip().casefold() for name in records}):
        raise SourceError(
            f"sheet(s) not found: {missing}",
            code="tabular_sheet_not_found",
            details={"source_ref": source_ref, "missing": missing, "available": list(records)},
        )
    return tuple(
        _jsonl_sheet(source_id, name, index, tuple(headers[name]), rows)
        for index, (name, rows) in enumerate(records.items())
    )


def _jsonl_sheet(
    source_id: str,
    name: str,
    index: int,
    headers: tuple[str, ...],
    records: list[tuple[int, int, dict[str, Any]]],
) -> TabularSheet:
    rows = tuple(
        TabularRow(
            source_row=source_row,
            data_row_index=data_index,
            values={header: raw_values.get(header) for header in headers},
            display_values={header: display_value(raw_values.get(header)) for header in headers},
        )
        for source_row, data_index, raw_values in records
    )
    return TabularSheet(
        table_id=table_id_for_source(source_id, name),
        name=name,
        index=index,
        headers=headers,
        rows=rows,
        normalized_headers=_normalized_headers(headers),
    )


def _sheet(
    source_id: str,
    name: str,
    index: int,
    raw_rows: Iterable[Any],
    budget: _DecodedBudget,
) -> TabularSheet:
    headers: tuple[str, ...] | None = None
    rows: list[TabularRow] = []
    data_index = 0
    for source_row, raw in enumerate(raw_rows, start=1):
        values = list(raw)
        budget.take(values, label=name)
        if headers is None:
            display = [display_value(value).strip() for value in values]
            if any(display):
                headers = _unique_headers(display)
            continue
        if not any(display_value(value).strip() for value in values):
            continue
        data_index += 1
        mapped = {header: values[position] if position < len(values) else None for position, header in enumerate(headers)}
        rows.append(
            TabularRow(
                source_row=source_row,
                data_row_index=data_index,
                values=mapped,
                display_values={header: display_value(value) for header, value in mapped.items()},
            )
        )
    if headers is None:
        raise SourceError("tabular source has no header row", code="tabular_missing_headers")
    return TabularSheet(
        table_id=table_id_for_source(source_id, name),
        name=name,
        index=index,
        headers=headers,
        rows=tuple(rows),
        normalized_headers=_normalized_headers(headers),
    )


def _bounded_rows(rows: Iterable[Any], label: str) -> Iterator[Any]:
    for index, row in enumerate(rows, start=1):
        if index > _TABULAR_MAX_SOURCE_ROWS:
            raise SourceError(
                f"tabular table {label!r} exceeds {_TABULAR_MAX_SOURCE_ROWS} source rows",
                code="tabular_source_row_limit",
                details={"table": label, "max_rows": _TABULAR_MAX_SOURCE_ROWS},
            )
        yield row


def _csv_dialect(sample: str) -> Any:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        dialect.doublequote = True
        return dialect
    except csv.Error:
        return csv.excel


def _unique_headers(values: list[str]) -> tuple[str, ...]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = value or f"column_{index}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return tuple(headers)


def _normalized_headers(headers: tuple[str, ...]) -> dict[str, tuple[str, ...]]:
    grouped: dict[str, list[str]] = {}
    for header in headers:
        grouped.setdefault(normalize_header(header), []).append(header)
    return {key: tuple(values) for key, values in grouped.items() if key}
